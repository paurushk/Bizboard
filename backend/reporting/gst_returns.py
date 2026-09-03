"""
GSTR-1 / GSTR-3B / GSTR-9 builders from completed GST documents.

Offline CA aids — not GSTN portal upload schema unless feature flag enabled.
Builder version: gstr1@2.0.0 — rate-wise B2B/CDNR, line HSN/UQC snapshots only.
"""

from __future__ import annotations

import hashlib
import json
import zipfile
from calendar import monthrange
from collections import defaultdict
from datetime import date
from decimal import Decimal
from io import BytesIO

from django.utils import timezone

from core.services.billing import extract_state_code, is_intra_state, q2
from core.services.uqc import normalize_uqc
from purchases.models import PurchaseCreditNote, PurchaseDebitNote, PurchaseInvoice
from sales.models import SalesCreditNote, SalesDebitNote, SalesInvoice

from .models import GstReturnPeriod, GstReturnSnapshot
from .gst_returns_sections import (
    accumulate_hsn_line,
    append_b2_outward_rows,
    apply_after_tax_header_discount,
    build_note_rate_rows,
    new_hsn_buckets,
)

BUILDER_VERSION_GSTR1 = "gstr1@2.0.0"
BUILDER_VERSION_GSTR3B = "gstr3b@2.2.0"
BUILDER_VERSION_GSTR9 = "gstr9@2.2.0"

# Notification 12/2024-CT: interstate B2C large (B2CL) threshold ₹1,00,000 from 1 Aug 2024.
# Invoices dated before that date still use the prior ₹2.5 lakh cut.
B2CL_THRESHOLD = Decimal("100000")
B2CL_THRESHOLD_BEFORE_2024_08_01 = Decimal("250000")
B2CL_THRESHOLD_CHANGE_DATE = date(2024, 8, 1)


def b2cl_threshold_for(invoice_date) -> Decimal:
    if invoice_date is None:
        return B2CL_THRESHOLD
    if hasattr(invoice_date, "date") and not isinstance(invoice_date, date):
        invoice_date = invoice_date.date()
    if invoice_date < B2CL_THRESHOLD_CHANGE_DATE:
        return B2CL_THRESHOLD_BEFORE_2024_08_01
    return B2CL_THRESHOLD


GST_INVOICE_TYPES = {
    SalesInvoice.InvoiceType.GST,
    SalesInvoice.InvoiceType.TAX,
    SalesInvoice.InvoiceType.RETAIL,
}
# AATO default for e-Invoice mandatory alert (₹5 crore) — override via company.aato_turnover.
EINVOICE_AATO_THRESHOLD = Decimal("50000000")


def parse_period(period: str) -> tuple[date, date]:
    try:
        year_str, month_str = period.split("-", 1)
        year, month = int(year_str), int(month_str)
        if month < 1 or month > 12:
            raise ValueError
    except (ValueError, AttributeError) as exc:
        raise ValueError(f"Invalid period '{period}'. Expected YYYY-MM.") from exc
    last_day = monthrange(year, month)[1]
    return date(year, month, 1), date(year, month, last_day)


def _money(value: Decimal | None) -> str:
    return str(q2(value or Decimal("0")))


def _sum_fields(rows: list[dict], *keys: str) -> dict[str, str]:
    totals = {key: Decimal("0") for key in keys}
    for row in rows:
        for key in keys:
            totals[key] += Decimal(str(row.get(key, "0")))
    return {key: _money(val) for key, val in totals.items()}


def _party_pos(company, party_state: str, party_gstin: str, *, filing_pos: str = "") -> str:
    overlay = (filing_pos or "").strip()
    if overlay:
        mapped = extract_state_code(overlay)
        if mapped:
            return mapped
    code = extract_state_code(party_gstin)
    if code:
        return code
    mapped = extract_state_code(party_state)
    if mapped:
        return mapped
    if company.assume_local_state_for_blank_party:
        return (
            extract_state_code(company.gstin)
            or extract_state_code(company.state)
            or "NA"
        )
    return "NA"


def _filing_gstin(invoice: SalesInvoice) -> str:
    overlay = (getattr(invoice, "filing_party_gstin", None) or "").strip()
    if overlay:
        return overlay
    return (invoice.customer.gstin or "").strip()


def _filing_pos(invoice: SalesInvoice, company) -> str:
    overlay = (getattr(invoice, "filing_place_of_supply", None) or "").strip()
    return _party_pos(
        company,
        invoice.customer.state or "",
        _filing_gstin(invoice),
        filing_pos=overlay,
    )


def _is_b2b(party_gstin: str) -> bool:
    # R4-006: only a 15-char GSTIN with a recognised state code routes a supply
    # to the B2B table — stray free-text left in the GSTIN field must not.
    g = (party_gstin or "").strip()
    return len(g) == 15 and extract_state_code(g) is not None


def _inter_state(company, party_state: str, party_gstin: str, *, seller_gstin: str = "", seller_state: str = "") -> bool:
    return not is_intra_state(
        seller_state or company.state or "",
        party_state or "",
        company_gstin=seller_gstin or company.gstin or "",
        party_gstin=party_gstin or "",
    )


def invoice_value_mismatch(invoice) -> bool:
    """
    True when grand_total cannot equal taxable + taxes + charges ± round-off/discount.
    BB-000621: additional_charges must not silently drop invoices from GSTR sections.
    BB-000361: AFTER_TAX discount causes GST value mismatch for B2B.
    """
    taxable = Decimal(str(invoice.taxable_total or 0))
    tax = (
        Decimal(str(invoice.cgst_total or 0))
        + Decimal(str(invoice.sgst_total or 0))
        + Decimal(str(invoice.igst_total or 0))
        + Decimal(str(getattr(invoice, "cess_total", 0) or 0))
    )
    round_off = Decimal(str(invoice.round_off or 0))
    charges = Decimal(str(getattr(invoice, "additional_charges", 0) or 0))
    discount = Decimal(str(getattr(invoice, "invoice_discount", 0) or 0))
    mode = getattr(invoice, "invoice_discount_mode", "AFTER_TAX")

    if discount != 0 and str(mode).upper() == "AFTER_TAX":
        party_gstin = (
            (getattr(invoice, "filing_party_gstin", None) or "")
            or (getattr(getattr(invoice, "customer", None), "gstin", None) or "")
        ).strip()
        # B2B AFTER_TAX is a GSTN identity break (Complete already blocks it).
        # B2C cash discount is allowed — reconcile expected = taxable+tax+charges−discount.
        if party_gstin:
            return True

    expected = q2(taxable + tax + round_off)
    from core.services.charges import charges_are_taxable

    if not charges_are_taxable(invoice):
        expected = q2(expected + charges)
    if discount != 0 and str(mode).upper() == "AFTER_TAX":
        expected = q2(expected - discount)
    if getattr(invoice, "tcs_in_grand_total", False):
        expected = q2(expected + Decimal(str(getattr(invoice, "tcs_amount", 0) or 0)))
    grand = q2(Decimal(str(invoice.grand_total or 0)))
    # GST-01: an exact `!=` here dropped the entire invoice's tax from both the
    # GSTR-1 sections and the 3B liability for a single-paise drift on a legacy /
    # imported document. Only a materially unreconciled invoice (data actually
    # broken) should be pulled out; tolerate the ≤5 paise rounding band that
    # `post_sales_invoice`'s own drift guard already allows.
    return abs(expected - grand) > Decimal("0.05")


def note_value_mismatch(note) -> bool:
    """GST-06: `invoice_value_mismatch` is written for invoices (TCS-in-grand,
    invoice_discount_mode, additional_charges). A credit / debit note only ever
    foots as taxable + tax ± round-off, so check that directly instead of
    reusing the invoice formula (which, via getattr defaults, happens to work
    today but would silently misjudge a note the moment a note grows one of
    those fields).
    """
    taxable = Decimal(str(getattr(note, "taxable_total", 0) or 0))
    tax = (
        Decimal(str(getattr(note, "cgst_total", 0) or 0))
        + Decimal(str(getattr(note, "sgst_total", 0) or 0))
        + Decimal(str(getattr(note, "igst_total", 0) or 0))
        + Decimal(str(getattr(note, "cess_total", 0) or 0))
    )
    round_off = Decimal(str(getattr(note, "round_off", 0) or 0))
    tcs = Decimal(str(getattr(note, "tcs_amount", 0) or 0))
    expected = q2(taxable + tax + round_off + tcs)
    grand = q2(Decimal(str(getattr(note, "grand_total", 0) or 0)))
    return abs(expected - grand) > Decimal("0.05")


def _rate_buckets(items, invoice=None) -> dict[Decimal, dict]:
    buckets: dict[Decimal, dict] = defaultdict(
        lambda: {
            "taxable_value": Decimal("0"),
            "cgst": Decimal("0"),
            "sgst": Decimal("0"),
            "igst": Decimal("0"),
            "cess": Decimal("0"),
        }
    )
    restore_rcm = bool(invoice and getattr(invoice, "is_reverse_charge", False))
    rcm_igst = Decimal(str(getattr(invoice, "rcm_igst", 0) or 0)) if invoice else Decimal("0")
    rcm_cgst = Decimal(str(getattr(invoice, "rcm_cgst", 0) or 0)) if invoice else Decimal("0")
    rcm_sgst = Decimal(str(getattr(invoice, "rcm_sgst", 0) or 0)) if invoice else Decimal("0")
    # GST-04: don't infer intra/inter purely from `rcm_igst == 0` — a legacy row
    # that never populated rcm_igst would then rebuild an inter-state RCM supply
    # as CGST/SGST. Prefer the explicit memo split, then the invoice's own
    # is_intra_state, then the heuristic.
    if rcm_cgst > 0 or rcm_sgst > 0:
        intra = True
    elif rcm_igst > 0:
        intra = False
    else:
        _is_intra = getattr(invoice, "is_intra_state", None) if invoice else None
        intra = bool(_is_intra) if isinstance(_is_intra, bool) else (rcm_igst == 0)
    for item in items:
        nature = (getattr(item, "supply_nature", None) or "TAXABLE").upper()
        if nature in ("NIL", "EXEMPT", "NON_GST"):
            continue
        # GST-03: bucket by the date-frozen legally-in-force rate when we have it.
        rate = Decimal(str(getattr(item, "applied_rate", None) or item.gst_rate or 0))
        taxable = Decimal(str(item.taxable_amount or 0))
        cgst = Decimal(str(item.cgst or 0))
        sgst = Decimal(str(item.sgst or 0))
        igst = Decimal(str(item.igst or 0))
        cess = Decimal(str(getattr(item, "cess", 0) or 0))
        if restore_rcm and cgst + sgst + igst == 0 and rate > 0:
            from core.services.billing import q2

            tax = q2(taxable * rate / Decimal("100"))
            if intra:
                # GST-04: symmetric split so the rate bucket foots CGST == SGST
                # (the GSTN offline tool validates this per bucket — BILL-01).
                half = q2(tax / 2)
                cgst, sgst, igst = half, half, Decimal("0")
            else:
                cgst, sgst, igst = Decimal("0"), Decimal("0"), tax
            # R1-016 / R4-001: rebuild cess from BOTH the ad-valorem rate and the
            # specific per-unit amount (they are additive — see core.models
            # DocumentLineModel.cess_amount), matching _apply_line_tax.
            if cess == 0:
                cess_rate = Decimal(str(getattr(item, "cess_rate", 0) or 0))
                cess_specific = Decimal(str(getattr(item, "cess_amount", 0) or 0))
                qty = Decimal(str(getattr(item, "quantity", 0) or 0))
                ad_valorem = (
                    q2(taxable * cess_rate / Decimal("100")) if cess_rate > 0 else Decimal("0")
                )
                specific = q2(qty * cess_specific) if cess_specific > 0 else Decimal("0")
                cess = q2(ad_valorem + specific)
        buckets[rate]["taxable_value"] += taxable
        buckets[rate]["cgst"] += cgst
        buckets[rate]["sgst"] += sgst
        buckets[rate]["igst"] += igst
        buckets[rate]["cess"] += cess
    return buckets


def _gst_sales_invoices_base(company, date_from: date, date_to: date):
    return SalesInvoice.objects.filter(
        company=company,
        status__in=(SalesInvoice.Status.COMPLETED, SalesInvoice.Status.RETURNED),
        invoice_type__in=GST_INVOICE_TYPES,
        invoice_date__gte=date_from,
        invoice_date__lte=date_to,
        # BB-000335: Tally-migration opening balances are not real supplies —
        # they must never inflate GSTR-1/3B outward tax liability.
        is_opening_balance=False,
    )


def _gst_sales_gstin_stamps(company, date_from: date, date_to: date):
    """GST-09: distinct company_gstin_id values for the period without the
    heavy ``select_related`` / ``prefetch_related("items")`` payload."""
    return list(
        _gst_sales_invoices_base(company, date_from, date_to)
        .values_list("company_gstin_id", flat=True)
        .distinct()
    )


def _gst_sales_invoices(company, date_from: date, date_to: date, *, company_gstin_id=None):
    qs = (
        _gst_sales_invoices_base(company, date_from, date_to)
        .select_related("customer", "company_gstin")
        .prefetch_related("items")
    )
    if company_gstin_id is not None:
        qs = _filter_purchase_gstin(qs, company, company_gstin_id, field="company_gstin_id")
    return qs


def _gst_credit_notes(company, date_from: date, date_to: date, *, company_gstin_id=None):
    qs = (
        SalesCreditNote.objects.filter(
            company=company,
            status=SalesCreditNote.Status.COMPLETED,
            note_date__gte=date_from,
            note_date__lte=date_to,
            sales_invoice__invoice_type__in=GST_INVOICE_TYPES,
            # BB-000398: exclude notes against opening invoices.
            sales_invoice__is_opening_balance=False,
        )
        .select_related("customer", "sales_invoice", "sales_invoice__company_gstin")
        .prefetch_related("items")
    )
    if company_gstin_id is not None:
        qs = _filter_purchase_gstin(
            qs, company, company_gstin_id, field="sales_invoice__company_gstin_id"
        )
    return qs


def _gst_debit_notes(company, date_from: date, date_to: date, *, company_gstin_id=None):
    qs = (
        SalesDebitNote.objects.filter(
            company=company,
            status=SalesDebitNote.Status.COMPLETED,
            note_date__gte=date_from,
            note_date__lte=date_to,
            sales_invoice__invoice_type__in=GST_INVOICE_TYPES,
            sales_invoice__is_opening_balance=False,
        )
        .select_related("customer", "sales_invoice", "sales_invoice__company_gstin")
        .prefetch_related("items")
    )
    if company_gstin_id is not None:
        qs = _filter_purchase_gstin(
            qs, company, company_gstin_id, field="sales_invoice__company_gstin_id"
        )
    return qs


def _gst_purchase_invoices(company, date_from: date, date_to: date, *, company_gstin_id=None):
    qs = PurchaseInvoice.objects.filter(
        company=company,
        status__in=(PurchaseInvoice.Status.COMPLETED, PurchaseInvoice.Status.RETURNED),
        purchase_type=PurchaseInvoice.PurchaseType.GST,
        invoice_date__gte=date_from,
        invoice_date__lte=date_to,
        # BB-000335: opening balances are not real purchases — must never inflate
        # GSTR-2B-matching ITC claims or 3B inward liability.
        is_opening_balance=False,
    ).select_related("supplier", "company_gstin").prefetch_related("items")
    return _filter_purchase_gstin(qs, company, company_gstin_id, field="company_gstin_id")


def _filter_purchase_gstin(qs, company, company_gstin_id, *, field: str):
    if company_gstin_id is None:
        return qs
    from django.db.models import Q

    from accounts.models import CompanyGstin

    primary = CompanyGstin.objects.filter(company=company, is_primary=True, is_active=True).first()
    if primary is not None and primary.id == company_gstin_id:
        return qs.filter(Q(**{field: company_gstin_id}) | Q(**{field: None}))
    return qs.filter(**{field: company_gstin_id})


def _gst_purchase_credit_notes(company, date_from: date, date_to: date, *, company_gstin_id=None):
    """GST purchase CNs on non-RCM invoices — eligible for non-RCM ITC netting."""
    qs = PurchaseCreditNote.objects.filter(
        company=company,
        status=PurchaseCreditNote.Status.COMPLETED,
        note_date__gte=date_from,
        note_date__lte=date_to,
        purchase_invoice__purchase_type=PurchaseInvoice.PurchaseType.GST,
        purchase_invoice__is_reverse_charge=False,
        purchase_invoice__is_opening_balance=False,
    ).select_related("purchase_invoice")
    return _filter_purchase_gstin(qs, company, company_gstin_id, field="purchase_invoice__company_gstin_id")


def _gst_purchase_debit_notes(company, date_from: date, date_to: date, *, company_gstin_id=None):
    qs = PurchaseDebitNote.objects.filter(
        company=company,
        status=PurchaseDebitNote.Status.COMPLETED,
        note_date__gte=date_from,
        note_date__lte=date_to,
        purchase_invoice__purchase_type=PurchaseInvoice.PurchaseType.GST,
        purchase_invoice__is_reverse_charge=False,
        purchase_invoice__is_opening_balance=False,
    )
    return _filter_purchase_gstin(qs, company, company_gstin_id, field="purchase_invoice__company_gstin_id")


def _gst_purchase_credit_notes_rcm(company, date_from: date, date_to: date, *, company_gstin_id=None):
    """BB-000336: purchase CNs on RCM invoices — netted into 3.1(d), not 4(A)(3)."""
    qs = PurchaseCreditNote.objects.filter(
        company=company,
        status=PurchaseCreditNote.Status.COMPLETED,
        note_date__gte=date_from,
        note_date__lte=date_to,
        purchase_invoice__purchase_type=PurchaseInvoice.PurchaseType.GST,
        purchase_invoice__is_reverse_charge=True,
        purchase_invoice__is_opening_balance=False,
    )
    return _filter_purchase_gstin(qs, company, company_gstin_id, field="purchase_invoice__company_gstin_id")


def _gst_purchase_debit_notes_rcm(company, date_from: date, date_to: date, *, company_gstin_id=None):
    """BB-000336: purchase DNs on RCM invoices — netted into 3.1(d), not 4(A)(3)."""
    qs = PurchaseDebitNote.objects.filter(
        company=company,
        status=PurchaseDebitNote.Status.COMPLETED,
        note_date__gte=date_from,
        note_date__lte=date_to,
        purchase_invoice__purchase_type=PurchaseInvoice.PurchaseType.GST,
        purchase_invoice__is_reverse_charge=True,
        purchase_invoice__is_opening_balance=False,
    )
    return _filter_purchase_gstin(qs, company, company_gstin_id, field="purchase_invoice__company_gstin_id")


def assert_not_composition_for_regular_returns(company):
    from core.exceptions import BusinessRuleError
    from accounts.models import Company

    if company.registration_type == Company.RegistrationType.COMPOSITION:
        raise BusinessRuleError(
            "Composition dealers cannot export Regular GSTR-1/GSTR-3B packs. "
            "Use /api/v1/reports/cmp08/ and /api/v1/reports/gstr4/ instead."
        )


def _resolve_filing_gstin_id(company, invoices, *, company_gstin=None):
    """BB-000556: GSTR is per CompanyGstin stamp. Fail closed on mixed stamps.

    GST-09: ``invoices`` may be a materialised list OR an iterable of raw
    ``company_gstin_id`` values (ints) — ``build_gstr1`` passes the latter from
    a cheap ``.values_list(...).distinct()`` so the heavy prefetch query only
    runs once.
    """
    from accounts.models import CompanyGstin
    from core.exceptions import BusinessRuleError

    if company_gstin in (None, "", "all"):
        stamp_ids = {
            v if isinstance(v, int) else getattr(v, "company_gstin_id", None)
            for v in invoices
        }
        stamp_ids.discard(None)
        if len(stamp_ids) > 1:
            raise BusinessRuleError(
                "Period contains invoices stamped with multiple GSTINs. "
                "Pass company_gstin to file each registration separately."
            )
        if len(stamp_ids) == 1:
            return next(iter(stamp_ids))
        primary = CompanyGstin.objects.filter(
            company=company, is_primary=True, is_active=True
        ).first()
        return primary.id if primary else None

    if isinstance(company_gstin, int) or str(company_gstin).isdigit():
        row = CompanyGstin.objects.filter(company=company, pk=int(company_gstin)).first()
    else:
        row = CompanyGstin.objects.filter(
            company=company, gstin__iexact=str(company_gstin).strip()
        ).first()
    if row is None:
        raise BusinessRuleError("Unknown company_gstin for this company.")
    return row.id


def build_gstr1(company, period: str, *, company_gstin=None) -> dict:
    assert_not_composition_for_regular_returns(company)
    date_from, date_to = parse_period(period)
    stamp_id = _resolve_filing_gstin_id(
        company, _gst_sales_gstin_stamps(company, date_from, date_to), company_gstin=company_gstin
    )
    invoices = list(_gst_sales_invoices(company, date_from, date_to, company_gstin_id=stamp_id))
    credit_notes = list(_gst_credit_notes(company, date_from, date_to, company_gstin_id=stamp_id))
    debit_notes = list(_gst_debit_notes(company, date_from, date_to, company_gstin_id=stamp_id))

    issues: list[dict] = []
    b2b: list[dict] = []
    b2cl: list[dict] = []
    exp: list[dict] = []
    sez: list[dict] = []
    b2cs_buckets: dict[tuple, dict] = defaultdict(
        lambda: {
            "taxable_value": Decimal("0"),
            "cgst": Decimal("0"),
            "sgst": Decimal("0"),
            "igst": Decimal("0"),
            "cess": Decimal("0"),
        }
    )
    hsn_buckets = new_hsn_buckets()
    nil_bucket = {
        "taxable_value": Decimal("0"),
        "nil_rated": Decimal("0"),
        "exempt": Decimal("0"),
        "non_gst": Decimal("0"),
    }
    supecom_rows: list[dict] = []

    for inv in invoices:
        items = list(inv.items.all())
        for idx, item in enumerate(items, start=1):
            hsn_raw = (item.hsn_code or "").strip()
            uqc_raw = (getattr(item, "uqc_code", None) or "").strip()
            if not hsn_raw:
                issues.append({
                    "code": "HSN_MISSING",
                    "severity": "critical",
                    "document_type": "sales_invoice",
                    "document_id": inv.id,
                    "number": inv.number,
                    "message": f"Line {idx}: missing HSN snapshot.",
                })
            unit_name = (getattr(item, "unit_name", None) or "").strip()
            uqc_unmapped = not uqc_raw or (
                uqc_raw == "OTH" and unit_name and not normalize_uqc(unit_name)
            )
            if uqc_unmapped:
                issues.append({
                    "code": "UQC_UNMAPPED",
                    "severity": "warning",
                    "document_type": "sales_invoice",
                    "document_id": inv.id,
                    "number": inv.number,
                    "message": f"Line {idx}: no GSTN UQC mapped.",
                })

        if invoice_value_mismatch(inv):
            issues.append({
                "code": "INVOICE_VALUE_MISMATCH",
                "severity": "critical",
                "document_type": "sales_invoice",
                "document_id": inv.id,
                "number": inv.number,
                "message": "grand_total does not reconcile to taxable+tax (charges / AFTER_TAX discount).",
            })
            # Exclude from GSTN-shaped sections; HSN/UQC issues already recorded above.
            continue

        # R4-002: an unresolvable place of supply ("NA") would produce a GSTN
        # JSON the portal rejects — surface it as a blocking issue and keep the
        # invoice out of the section buckets rather than emitting "NA" rows.
        if _filing_pos(inv, company) == "NA":
            issues.append({
                "code": "PLACE_OF_SUPPLY_UNRESOLVED",
                "severity": "critical",
                "document_type": "sales_invoice",
                "document_id": inv.id,
                "number": inv.number,
                "message": (
                    "Place of supply could not be resolved to a GST state code — "
                    "set the customer's GSTIN or a valid state before filing."
                ),
            })
            continue
        charges = Decimal(str(getattr(inv, "additional_charges", 0) or 0))
        from core.services.charges import charges_are_taxable

        if charges > 0 and not charges_are_taxable(inv):
            issues.append({
                "code": "ADDITIONAL_CHARGES_NONTAXABLE",
                "severity": "warning",
                "document_type": "sales_invoice",
                "document_id": inv.id,
                "number": inv.number,
                "message": (
                    f"Additional charges of {charges} are treated as non-taxable. "
                    "Set charges HSN and GST rate if freight/packing should carry GST."
                ),
            })
        # Missing HSN is recorded as an issue above — still include the invoice in
        # B2 sections so the worksheet remains useful for remediation.
        ecom = (getattr(inv, "ecommerce_operator_gstin", None) or "").strip()
        if ecom:
            party_gstin = (inv.filing_party_gstin or getattr(inv.customer, "gstin", "") or "").strip()
            supecom_rows.append({
                "invoice_number": inv.number,
                "invoice_date": inv.invoice_date.isoformat() if inv.invoice_date else "",
                "ecommerce_operator_gstin": ecom,
                "taxable_value": _money(inv.taxable_total),
                "invoice_value": _money(inv.grand_total),
                "section": "15A" if party_gstin else "15B",
                "party_gstin": party_gstin,
            })
            continue

        from core.services.charges import charge_line as _charge_line

        stamp = getattr(inv, "company_gstin", None)
        intra = not _inter_state(
            company,
            _filing_pos(inv, company),
            _filing_gstin(inv),
            seller_gstin=(stamp.gstin if stamp is not None else "") or "",
            seller_state=(stamp.state if stamp is not None else "") or "",
        )
        ch = _charge_line(inv, intra_state=intra)
        if ch is not None:
            items = list(items) + [ch]
        buckets = _rate_buckets(items, invoice=inv)
        for item in items:
            nature = (getattr(item, "supply_nature", None) or "TAXABLE").upper()
            amt = Decimal(str(getattr(item, "taxable_amount", 0) or 0))
            if nature == "NIL":
                nil_bucket["nil_rated"] += amt
                nil_bucket["taxable_value"] += amt
            elif nature == "EXEMPT":
                nil_bucket["exempt"] += amt
                nil_bucket["taxable_value"] += amt
            elif nature == "NON_GST":
                nil_bucket["non_gst"] += amt
                nil_bucket["taxable_value"] += amt

        for item in items:
            accumulate_hsn_line(hsn_buckets, item)
        apply_after_tax_header_discount(hsn_buckets, inv, items)

        append_b2_outward_rows(
            company=company,
            inv=inv,
            items=items,
            buckets=buckets,
            b2b=b2b,
            b2cl=b2cl,
            b2cs_buckets=b2cs_buckets,
            exp=exp,
            sez=sez,
            nil_bucket=nil_bucket,
            issues=issues,
        )

    cdnr: list[dict] = []
    cdnur: list[dict] = []

    for note in credit_notes:
        parent = getattr(note, "sales_invoice", None)
        if parent is not None and (getattr(parent, "ecommerce_operator_gstin", None) or "").strip():
            continue
        build_note_rate_rows(
            company=company,
            note=note,
            note_kind="CREDIT",
            original_number=note.sales_invoice.number,
            issues=issues,
            hsn_buckets=hsn_buckets,
            cdnr=cdnr,
            cdnur=cdnur,
            b2cs_buckets=b2cs_buckets,
        )
    for note in debit_notes:
        parent = getattr(note, "sales_invoice", None)
        if parent is not None and (getattr(parent, "ecommerce_operator_gstin", None) or "").strip():
            continue
        build_note_rate_rows(
            company=company,
            note=note,
            note_kind="DEBIT",
            original_number=note.sales_invoice.number,
            issues=issues,
            hsn_buckets=hsn_buckets,
            cdnr=cdnr,
            cdnur=cdnur,
            b2cs_buckets=b2cs_buckets,
        )

    b2cs = [
        {
            "place_of_supply": pos,
            "rate": rate,
            "taxable_value": _money(vals["taxable_value"]),
            "cgst": _money(vals["cgst"]),
            "sgst": _money(vals["sgst"]),
            "igst": _money(vals["igst"]),
            "cess": _money(vals.get("cess", Decimal("0"))),
        }
        for (pos, rate), vals in sorted(b2cs_buckets.items())
        if vals["taxable_value"] or vals["cgst"] or vals["sgst"] or vals["igst"] or vals.get("cess")
    ]

    cdnr.sort(key=lambda r: (r["note_date"], r["note_number"], r["rate"]))
    cdnur.sort(key=lambda r: (r["note_date"], r["note_number"], r["rate"]))

    hsn = [
        {
            "hsn": hsn_code,
            "rate": rate,
            "uqc": uqc,
            "quantity": _money(vals["quantity"]),
            "taxable_value": _money(vals["taxable_value"]),
            "cgst": _money(vals["cgst"]),
            "sgst": _money(vals["sgst"]),
            "igst": _money(vals["igst"]),
            "cess": _money(vals.get("cess", Decimal("0"))),
        }
        for (hsn_code, rate, uqc), vals in sorted(hsn_buckets.items())
    ]

    cancelled_qs = SalesInvoice.objects.filter(
        company=company,
        status=SalesInvoice.Status.CANCELLED,
        invoice_date__gte=date_from,
        invoice_date__lte=date_to,
        invoice_type__in=GST_INVOICE_TYPES,
    )
    if stamp_id is not None:
        cancelled_qs = cancelled_qs.filter(company_gstin_id=stamp_id)
    cancelled_invoices = list(cancelled_qs)
    cancelled_inv = len(cancelled_invoices)

    # GST-05: an invoice whose invoice_date is in a *prior* period but which was
    # cancelled in this one. It has dropped out of outward supplies, yet if that
    # prior period's GSTR-1 was filed the supply is still on the portal — it
    # needs a 9A amendment, not a silent disappearance.
    prior_cancel_qs = SalesInvoice.objects.filter(
        company=company,
        status=SalesInvoice.Status.CANCELLED,
        invoice_type__in=GST_INVOICE_TYPES,
        invoice_date__lt=date_from,
        cancelled_at__date__gte=date_from,
        cancelled_at__date__lte=date_to,
    )
    if stamp_id is not None:
        prior_cancel_qs = prior_cancel_qs.filter(company_gstin_id=stamp_id)
    for inv in prior_cancel_qs:
        orig_period = inv.invoice_date.strftime("%Y-%m")
        filed = GstReturnSnapshot.objects.filter(
            company=company,
            period=orig_period,
            return_type=GstReturnSnapshot.ReturnType.GSTR1,
        ).exists()
        issues.append({
            "code": "PRIOR_PERIOD_INVOICE_CANCELLED",
            "severity": "critical" if filed else "warning",
            "document_type": "sales_invoice",
            "document_id": inv.pk,
            "number": inv.number or str(inv.pk),
            "message": (
                f"Invoice {inv.number or inv.pk} (dated {inv.invoice_date}, period "
                f"{orig_period}) was cancelled on {inv.cancelled_at.date() if inv.cancelled_at else '?'}. "
                + (
                    f"GSTR-1 for {orig_period} was already filed — file a Table 9A "
                    "amendment to reverse this outward supply."
                    if filed
                    else f"Amend or refile GSTR-1 for {orig_period} to drop this supply."
                )
            ),
        })

    supecom_invoices = [
        inv
        for inv in invoices
        if (getattr(inv, "ecommerce_operator_gstin", None) or "").strip()
    ]
    doc_invoices = list(invoices)
    cancelled_cn_qs = SalesCreditNote.objects.filter(
        company=company,
        status=SalesCreditNote.Status.CANCELLED,
        note_date__gte=date_from,
        note_date__lte=date_to,
    )
    cancelled_dn_qs = SalesDebitNote.objects.filter(
        company=company,
        status=SalesDebitNote.Status.CANCELLED,
        note_date__gte=date_from,
        note_date__lte=date_to,
    )
    if stamp_id is not None:
        cancelled_cn_qs = cancelled_cn_qs.filter(company_gstin_id=stamp_id)
        cancelled_dn_qs = cancelled_dn_qs.filter(company_gstin_id=stamp_id)
    cancelled_credit_notes = list(cancelled_cn_qs)
    cancelled_debit_notes = list(cancelled_dn_qs)
    regular_invoices = [
        inv for inv in invoices
        if not (getattr(inv, "ecommerce_operator_gstin", None) or "").strip()
    ]
    docs = {
        "invoices_issued": len(regular_invoices),
        "invoices_cancelled": cancelled_inv,
        "credit_notes_issued": len(credit_notes),
        "debit_notes_issued": len(debit_notes),
        "credit_notes_cancelled": len(cancelled_credit_notes),
        "debit_notes_cancelled": len(cancelled_debit_notes),
        "supecom_table15": len(supecom_invoices),
    }
    doc_table = _gstr1_doc_table(
        company,
        date_from,
        date_to,
        regular_invoices,
        cancelled_invoices,
        credit_notes,
        debit_notes,
        cancelled_credit_notes,
        cancelled_debit_notes,
    )
    at_table = _gstr1_at_table(company, date_from, date_to, company_gstin_id=stamp_id)
    atadj_table = _gstr1_atadj_table(company, date_from, date_to, company_gstin_id=stamp_id)
    txpd_table = _gstr1_txpd_table(company, date_from, date_to, company_gstin_id=stamp_id)

    # Liability / register totals: exclude INVOICE_VALUE_MISMATCH from outward so
    # GSTR-3B aligns with GSTR-1 section coverage (BB-000038 / G16). Reverse-charge
    # *sales* are shown in the B2B section (rchrg=Y) for disclosure but carry no
    # output-tax liability for the supplier, so they are kept out of the outward
    # liability totals.
    def _rcm(doc) -> bool:
        return bool(
            getattr(doc, "is_reverse_charge", False)
            or getattr(getattr(doc, "sales_invoice", None), "is_reverse_charge", False)
        )

    matched_invoices = [
        inv for inv in invoices
        if not invoice_value_mismatch(inv)
        and not (getattr(inv, "ecommerce_operator_gstin", None) or "").strip()
        and not _rcm(inv)
    ]
    def _note_parent_is_ecom(note) -> bool:
        parent = getattr(note, "sales_invoice", None) or getattr(note, "invoice", None)
        return bool(parent and (getattr(parent, "ecommerce_operator_gstin", None) or "").strip())

    matched_credit_notes = [
        n for n in credit_notes
        if not note_value_mismatch(n) and not _note_parent_is_ecom(n) and not _rcm(n)
    ]
    matched_debit_notes = [
        n for n in debit_notes
        if not note_value_mismatch(n) and not _note_parent_is_ecom(n) and not _rcm(n)
    ]
    outward_taxable = sum((inv.taxable_total for inv in matched_invoices), Decimal("0"))
    outward_cgst = sum((inv.cgst_total for inv in matched_invoices), Decimal("0"))
    outward_sgst = sum((inv.sgst_total for inv in matched_invoices), Decimal("0"))
    outward_igst = sum((inv.igst_total for inv in matched_invoices), Decimal("0"))
    outward_cess = sum((Decimal(str(getattr(inv, "cess_total", 0) or 0)) for inv in matched_invoices), Decimal("0"))
    for note in matched_credit_notes:
        outward_taxable -= note.taxable_total
        outward_cgst -= note.cgst_total
        outward_sgst -= note.sgst_total
        outward_igst -= note.igst_total
        outward_cess -= Decimal(str(getattr(note, "cess_total", 0) or 0))
    for note in matched_debit_notes:
        outward_taxable += note.taxable_total
        outward_cgst += note.cgst_total
        outward_sgst += note.sgst_total
        outward_igst += note.igst_total
        outward_cess += Decimal(str(getattr(note, "cess_total", 0) or 0))

    section_taxable = (
        sum(Decimal(r["taxable_value"]) for r in b2b)
        + sum(Decimal(r["taxable_value"]) for r in b2cl)
        + sum(Decimal(r["taxable_value"]) for r in b2cs)
        + sum(Decimal(r["taxable_value"]) for r in exp)
        + sum(Decimal(r["taxable_value"]) for r in sez)
        + nil_bucket["taxable_value"]
        - sum(Decimal(r["taxable_value"]) for r in cdnr if r["note_kind"] == "CREDIT")
        - sum(Decimal(r["taxable_value"]) for r in cdnur if r["note_kind"] == "CREDIT")
        + sum(Decimal(r["taxable_value"]) for r in cdnr if r["note_kind"] == "DEBIT")
        + sum(Decimal(r["taxable_value"]) for r in cdnur if r["note_kind"] == "DEBIT")
    )

    # Wave 17B: amendments from filing identity audit trail when present.
    amendments = _gstr1_amendments(company, date_from, date_to, company_gstin_id=stamp_id)

    footing_delta = outward_taxable - section_taxable
    footing_mismatch = abs(footing_delta) > Decimal("0.01")
    if footing_mismatch:
        # GST-07: a sub-rupee delta is rounding drift (warn); anything larger is
        # a real bucketing bug that must block the file, not just caution.
        footing_severity = "warning" if abs(footing_delta) <= Decimal("1") else "critical"
        issues.append({
            "code": "OUTWARD_FOOTING_MISMATCH",
            "severity": footing_severity,
            "document_type": "gstr1",
            "document_id": None,
            "number": period,
            "message": (
                f"Header outward_taxable {_money(outward_taxable)} differs from "
                f"section_net_taxable {_money(section_taxable)} "
                f"(delta {_money(footing_delta)})."
            ),
        })

    totals = {
        "outward_taxable": _money(outward_taxable),
        "outward_cgst": _money(outward_cgst),
        "outward_sgst": _money(outward_sgst),
        "outward_igst": _money(outward_igst),
        "outward_cess": _money(outward_cess),
        "section_net_taxable": _money(section_taxable),
        "footing_discrepancy": footing_mismatch,
        "footing_delta": _money(footing_delta) if footing_mismatch else "0.00",
        "b2b": _sum_fields(b2b, "taxable_value", "cgst", "sgst", "igst", "cess"),
        "b2cl": _sum_fields(b2cl, "taxable_value", "igst", "cess"),
        "b2cs": _sum_fields(b2cs, "taxable_value", "cgst", "sgst", "igst", "cess"),
        "exp": _sum_fields(exp, "taxable_value", "igst", "cess"),
        "sez": _sum_fields(sez, "taxable_value", "cgst", "sgst", "igst", "cess"),
        "cdnr": _sum_fields(cdnr, "taxable_value", "cgst", "sgst", "igst", "cess"),
        "cdnur": _sum_fields(cdnur, "taxable_value", "cgst", "sgst", "igst", "cess"),
    }

    from accounts.models import CompanyGstin

    stamp_gstin = company.gstin or ""
    stamp_state = company.state or ""
    if stamp_id:
        stamp_row = CompanyGstin.objects.filter(pk=stamp_id).values("gstin", "state").first()
        if stamp_row:
            stamp_gstin = stamp_row.get("gstin") or stamp_gstin
            stamp_state = stamp_row.get("state") or stamp_state

    return {
        "return_type": "GSTR-1",
        "builder_version": BUILDER_VERSION_GSTR1,
        "period": period,
        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
        "company_gstin_id": stamp_id,
        "company": {
            "name": company.name,
            "gstin": stamp_gstin,
            "state": stamp_state,
        },
        "b2b": sorted(b2b, key=lambda r: (r["invoice_date"], r["invoice_number"], r["rate"])),
        "b2cl": sorted(b2cl, key=lambda r: (r["invoice_date"], r["invoice_number"], r["rate"])),
        "b2cs": b2cs,
        "exp": sorted(exp, key=lambda r: (r["invoice_date"], r["invoice_number"], r["rate"])),
        "sez": sorted(sez, key=lambda r: (r["invoice_date"], r["invoice_number"], r["rate"])),
        "cdnr": cdnr,
        "cdnur": cdnur,
        "hsn": hsn,
        "nil": {
            "taxable_value": _money(nil_bucket["taxable_value"]),
            "nil_rated": _money(nil_bucket["nil_rated"]),
            "exempt": _money(nil_bucket["exempt"]),
            "non_gst": _money(nil_bucket["non_gst"]),
            "aid_kind": "nil_split" if nil_bucket["taxable_value"] else "nil_unsplit",
            "note": (
                "Nil / exempt / non-GST use line supply_nature. 0% TAXABLE lines still "
                "accumulate as nil-rated aid."
            ),
        },
        "amendments": amendments,
        "docs": docs,
        "doc": doc_table,
        "at": at_table,
        "atadj": atadj_table,
        "txpd": txpd_table,
        "supecom": {
            "supported": False,
            "table": "15",
            "15A": [r for r in supecom_rows if r.get("section") == "15A"],
            "15B": [r for r in supecom_rows if r.get("section") == "15B"],
            "rows": supecom_rows,
            "note": (
                "Table 15 aid only — not a portal-complete SUPECOM engine. "
                "15A = supplies through e-commerce to registered buyers; 15B = unregistered. "
                "These invoices are excluded from B2 / 3.1(a) in this worksheet so Table 15 "
                "and B2 are not double-filed. Confirm with your CA before portal upload."
            ),
        },
        "totals": totals,
        "issues": issues,
        "disclaimer": "Offline export for CA review — not a GSTN portal upload file.",
    }


def _gstr1_doc_table(
    company,
    date_from,
    date_to,
    invoices,
    cancelled_invoices,
    credit_notes,
    debit_notes,
    cancelled_credit_notes=None,
    cancelled_debit_notes=None,
) -> list[dict]:
    """GSTR-1 DOC: invoice series summary with natural sorting (prefix from document number)."""
    import re
    from collections import defaultdict

    def _doc_sort_key(num: str):
        raw = (num or "").strip()
        m = re.search(r"(\d+)\s*$", raw)
        num_val = int(m.group(1)) if m else 0
        return (num_val, raw)

    def _prefix(num: str) -> str:
        return "".join(ch for ch in (num or "") if not ch.isdigit()) or "INV"

    def _summarize_docs(doc_list, cancelled_list, default_prefix="INV"):
        series: dict[str, dict] = defaultdict(lambda: {"nums": [], "total": 0, "cancelled": 0})
        for item in doc_list:
            num = getattr(item, "number", "") or ""
            if not num:
                continue
            p = _prefix(num)
            bucket = series[p]
            bucket["total"] += 1
            bucket["nums"].append(num)
        for item in cancelled_list:
            num = getattr(item, "number", "") or ""
            if not num:
                continue
            p = _prefix(num)
            bucket = series[p]
            bucket["total"] += 1
            bucket["cancelled"] += 1
            bucket["nums"].append(num)
        results = []
        for p, bucket in sorted(series.items()):
            sorted_nums = sorted(bucket["nums"], key=_doc_sort_key)
            sr_from = sorted_nums[0] if sorted_nums else ""
            sr_to = sorted_nums[-1] if sorted_nums else ""
            results.append({
                "sr_from": sr_from,
                "sr_to": sr_to,
                "total_number": bucket["total"],
                "cancelled": bucket["cancelled"],
                "series": p,
            })
        return results

    rows = []
    inv_summaries = _summarize_docs(invoices, cancelled_invoices, default_prefix="INV")
    for s in inv_summaries:
        rows.append({
            "nature": "Invoices for outward supply",
            "sr_from": s["sr_from"],
            "sr_to": s["sr_to"],
            "total_number": s["total_number"],
            "cancelled": s["cancelled"],
            "series": s["series"],
        })
    if not inv_summaries:
        rows.append({
            "nature": "Invoices for outward supply",
            "sr_from": "",
            "sr_to": "",
            "total_number": 0,
            "cancelled": 0,
            "series": "",
        })

    cancelled_credit_notes = cancelled_credit_notes or []
    cancelled_debit_notes = cancelled_debit_notes or []
    cancelled_cn = len(cancelled_credit_notes)
    cancelled_dn = len(cancelled_debit_notes)

    cn_summaries = _summarize_docs(credit_notes, cancelled_credit_notes, default_prefix="SCN")
    if cn_summaries:
        for s in cn_summaries:
            rows.append({
                "nature": "Credit notes",
                "sr_from": s["sr_from"],
                "sr_to": s["sr_to"],
                "total_number": s["total_number"],
                "cancelled": s["cancelled"],
                "series": s["series"],
            })
    else:
        rows.append({
            "nature": "Credit notes",
            "sr_from": "",
            "sr_to": "",
            "total_number": len(credit_notes) + cancelled_cn,
            "cancelled": cancelled_cn,
            "series": "",
        })

    dn_summaries = _summarize_docs(debit_notes, cancelled_debit_notes, default_prefix="SDN")
    if dn_summaries:
        for s in dn_summaries:
            rows.append({
                "nature": "Debit notes",
                "sr_from": s["sr_from"],
                "sr_to": s["sr_to"],
                "total_number": s["total_number"],
                "cancelled": s["cancelled"],
                "series": s["series"],
            })
    else:
        rows.append({
            "nature": "Debit notes",
            "sr_from": "",
            "sr_to": "",
            "total_number": len(debit_notes) + cancelled_dn,
            "cancelled": cancelled_dn,
            "series": "",
        })
    return rows


def _gstr1_at_table(company, date_from, date_to, *, company_gstin_id=None) -> list[dict]:
    """GSTR-1 AT aid: unallocated customer receipts in period (advances).

    Advances are company-level until allocated to a stamped invoice. When a
    non-primary ``company_gstin_id`` is requested, return empty (ATADJ covers
    stamp-scoped allocations). Primary / unset stamp includes all unallocated.
    """
    from django.db.models import Sum as DjSum

    from accounts.models import CompanyGstin
    from core.services.billing import extract_state_code
    from payments.models import CustomerReceipt, PaymentAllocation, ReceiptStatus

    if company_gstin_id is not None:
        primary = CompanyGstin.objects.filter(
            company=company, is_primary=True, is_active=True
        ).first()
        if primary is not None and primary.id != company_gstin_id:
            return []

    rows = []
    receipts = CustomerReceipt.objects.filter(
        company=company,
        receipt_date__gte=date_from,
        receipt_date__lte=date_to,
        status=ReceiptStatus.POSTED,
    ).select_related("customer")
    for rec in receipts:
        allocated = (
            PaymentAllocation.objects.filter(receipt=rec, reversed_at__isnull=True).aggregate(t=DjSum("amount"))["t"]
            or Decimal("0")
        )
        unalloc = Decimal(str(rec.amount or 0)) - Decimal(str(allocated or 0))
        if unalloc <= 0:
            continue
        customer = rec.customer
        pos = (
            extract_state_code(getattr(customer, "gstin", "") or "")
            or extract_state_code(getattr(customer, "state", "") or "")
            or ""
        )
        rows.append({
            "receipt_number": rec.number,
            "receipt_date": rec.receipt_date.isoformat() if rec.receipt_date else "",
            "place_of_supply": pos,
            "rate": "0.00",
            "cgst": "0.00",
            "sgst": "0.00",
            "igst": "0.00",
            "gross_advance": _money(unalloc),
            "tax_status": "rate_unknown",
            "note": "Unallocated receipt is an advance aid pending CA rate — not GSTN AT tax.",
            "aid_kind": "unallocated_receipt",
            "honesty": "rate_unknown_do_not_file_as_at",
        })
    return rows


def _gstr1_atadj_table(company, date_from, date_to, *, company_gstin_id=None) -> list[dict]:
    """BB-000621: ATADJ aid when advances are allocated to invoices."""
    from payments.models import PaymentAllocation

    rows = []
    allocs = (
        PaymentAllocation.objects.filter(
            company=company,
            receipt__isnull=False,
            sales_invoice__isnull=False,
            reversed_at__isnull=True,
            sales_invoice__invoice_date__gte=date_from,
            sales_invoice__invoice_date__lte=date_to,
        )
        .select_related("receipt", "sales_invoice", "receipt__customer")
    )
    if company_gstin_id is not None:
        allocs = allocs.filter(sales_invoice__company_gstin_id=company_gstin_id)
    for alloc in allocs:
        rec = alloc.receipt
        inv = alloc.sales_invoice
        if rec is None or inv is None:
            continue
        if rec.receipt_date and inv.invoice_date and rec.receipt_date > inv.invoice_date:
            continue
        rows.append({
            "receipt_number": rec.number,
            "receipt_date": rec.receipt_date.isoformat() if rec.receipt_date else "",
            "invoice_number": inv.number,
            "invoice_date": inv.invoice_date.isoformat() if inv.invoice_date else "",
            "adjusted_amount": _money(alloc.amount),
            "aid_kind": "atadj_allocation",
            "note": "Advance allocation aid — rate unknown; do not double-count with AT.",
        })
    return rows


def _gstr1_txpd_table(company, date_from, date_to, *, company_gstin_id=None) -> list[dict]:
    """TXPD aid mirrors AT advances — tax rate unknown until invoiced."""
    at_rows = _gstr1_at_table(company, date_from, date_to, company_gstin_id=company_gstin_id)
    if not at_rows:
        return [{
            "aid_kind": "txpd_none",
            "supported": False,
            "note": "No unallocated advances in period — TXPD is nil for this aid.",
        }]
    return [
        {
            **row,
            "aid_kind": "txpd_from_at",
            "note": (
                "TXPD aid copies unallocated advances (AT). Tax rate is unknown — "
                "do not file as GSTN TXPD without CA rate split."
            ),
        }
        for row in at_rows
    ]


def _gstr1_amendments(company, date_from: date, date_to: date, *, company_gstin_id=None) -> list[dict]:
    """Build amendment rows from AuditEvent filing-identity changes when available."""
    from core.models import AuditEvent
    from sales.models import SalesInvoice

    qs = AuditEvent.objects.filter(
        company=company,
        entity_type="salesinvoice",
        description__icontains="amend_filing_identity",
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
    ).order_by("created_at")[:500]
    allowed_ids = None
    if company_gstin_id is not None:
        allowed_ids = set(
            SalesInvoice.objects.filter(
                company=company, company_gstin_id=company_gstin_id
            ).values_list("id", flat=True)
        )
    rows = []
    for row in qs:
        if allowed_ids is not None and row.entity_id not in allowed_ids:
            continue
        meta = row.metadata if isinstance(getattr(row, "metadata", None), dict) else {}
        rows.append({
            "document_id": row.entity_id,
            "changed_at": row.created_at.isoformat() if row.created_at else "",
            "old_gstin": meta.get("old_gstin", ""),
            "new_gstin": meta.get("new_gstin", ""),
            "old_pos": meta.get("old_pos", ""),
            "new_pos": meta.get("new_pos", ""),
            "aid_kind": "amendment_audit",
        })
    return rows


def build_gstr3b(company, period: str, gstr1: dict | None = None, *, company_gstin=None) -> dict:
    assert_not_composition_for_regular_returns(company)
    date_from, date_to = parse_period(period)
    if gstr1 is None:
        gstr1 = build_gstr1(company, period, company_gstin=company_gstin)
    # BB-000697: never resolve stamp from []; reuse GSTR-1 stamp or explicit param.
    stamp_id = gstr1.get("company_gstin_id")
    if stamp_id is None and company_gstin is not None:
        stamp_id = getattr(company_gstin, "id", company_gstin)
    if stamp_id is None:
        stamp_id = _resolve_filing_gstin_id(
            company,
            gstr1.get("_stamp_invoices") or [],
            company_gstin=company_gstin,
        )
    purchases = list(_gst_purchase_invoices(company, date_from, date_to, company_gstin_id=stamp_id))
    purchase_cns = list(_gst_purchase_credit_notes(company, date_from, date_to, company_gstin_id=stamp_id))
    purchase_dns = list(_gst_purchase_debit_notes(company, date_from, date_to, company_gstin_id=stamp_id))
    purchase_cns_rcm = list(_gst_purchase_credit_notes_rcm(company, date_from, date_to, company_gstin_id=stamp_id))
    purchase_dns_rcm = list(_gst_purchase_debit_notes_rcm(company, date_from, date_to, company_gstin_id=stamp_id))

    non_rcm = [p for p in purchases if not getattr(p, "is_reverse_charge", False)]
    rcm = [p for p in purchases if getattr(p, "is_reverse_charge", False)]
    itc_eligible = [
        p
        for p in non_rcm
        if getattr(p, "itc_eligibility", "") == PurchaseInvoice.ItcEligibility.CLAIMABLE
    ]

    def _note_claimable(note) -> bool:
        inv = getattr(note, "purchase_invoice", None)
        return inv is not None and getattr(inv, "itc_eligibility", "") == PurchaseInvoice.ItcEligibility.CLAIMABLE

    # GST-10: ITC parked as UNREVIEWED is neither claimed in table 4 nor
    # surfaced anywhere — a user reconciling against 2B has no line for it.
    unreviewed_itc_rows = [
        p
        for p in non_rcm
        if getattr(p, "itc_eligibility", "") == PurchaseInvoice.ItcEligibility.UNREVIEWED
    ]
    unreviewed_itc = {
        "count": len(unreviewed_itc_rows),
        "cgst": _money(sum((inv.cgst_total for inv in unreviewed_itc_rows), Decimal("0"))),
        "sgst": _money(sum((inv.sgst_total for inv in unreviewed_itc_rows), Decimal("0"))),
        "igst": _money(sum((inv.igst_total for inv in unreviewed_itc_rows), Decimal("0"))),
        "cess": _money(
            sum(
                (Decimal(str(getattr(inv, "cess_total", 0) or 0)) for inv in unreviewed_itc_rows),
                Decimal("0"),
            )
        ),
        "note": (
            "ITC on these purchases is parked pending IMS/2B review — not included "
            "in table 4(A) claimable ITC above. Accept or reject each before filing."
        ),
    }

    inward_taxable = sum((inv.taxable_total for inv in non_rcm), Decimal("0"))
    inward_cgst = sum((inv.cgst_total for inv in itc_eligible), Decimal("0"))
    inward_sgst = sum((inv.sgst_total for inv in itc_eligible), Decimal("0"))
    inward_igst = sum((inv.igst_total for inv in itc_eligible), Decimal("0"))
    inward_cess = sum(
        (Decimal(str(getattr(inv, "cess_total", 0) or 0)) for inv in itc_eligible),
        Decimal("0"),
    )
    for note in purchase_cns:
        inward_taxable -= note.taxable_total
        if _note_claimable(note):
            inward_cgst -= note.cgst_total
            inward_sgst -= note.sgst_total
            inward_igst -= note.igst_total
            inward_cess -= Decimal(str(getattr(note, "cess_total", 0) or 0))
    for note in purchase_dns:
        inward_taxable += note.taxable_total
        if _note_claimable(note):
            inward_cgst += note.cgst_total
            inward_sgst += note.sgst_total
            inward_igst += note.igst_total
            inward_cess += Decimal(str(getattr(note, "cess_total", 0) or 0))

    # SYS-03 / R4-005: the memoised rcm_* fields are authoritative once *any* of
    # them is set. Only fall back to taxable_total / cess_total for genuinely
    # legacy rows that carry no RCM memo at all — not when the memo legitimately
    # says zero (a `X or fallback` chain would silently swap in taxable_total).
    def _has_rcm_memo(row) -> bool:
        return any(
            Decimal(str(getattr(row, f, 0) or 0)) != 0
            for f in ("rcm_taxable", "rcm_cgst", "rcm_sgst", "rcm_igst", "rcm_cess")
        )

    def _rcm_taxable(row) -> Decimal:
        if _has_rcm_memo(row):
            return Decimal(str(getattr(row, "rcm_taxable", 0) or 0))
        return Decimal(str(getattr(row, "taxable_total", 0) or 0))

    rcm_taxable = sum((_rcm_taxable(p) for p in rcm), Decimal("0"))
    rcm_cgst = sum((Decimal(str(getattr(p, "rcm_cgst", 0) or 0)) for p in rcm), Decimal("0"))
    rcm_sgst = sum((Decimal(str(getattr(p, "rcm_sgst", 0) or 0)) for p in rcm), Decimal("0"))
    rcm_igst = sum((Decimal(str(getattr(p, "rcm_igst", 0) or 0)) for p in rcm), Decimal("0"))
    rcm_cess = sum((Decimal(str(getattr(p, "rcm_cess", 0) or 0)) for p in rcm), Decimal("0"))
    # BB-000336: RCM credit/debit notes reduce/increase the RCM liability itself
    # (3.1(d)) — they must never net into the non-RCM ITC block above.

    def _rcm_cess(note):
        if _has_rcm_memo(note):
            return Decimal(str(getattr(note, "rcm_cess", 0) or 0))
        return Decimal(str(getattr(note, "cess_total", 0) or 0))

    for note in purchase_cns_rcm:
        rcm_taxable -= _rcm_taxable(note)
        rcm_cgst -= Decimal(str(getattr(note, "rcm_cgst", 0) or 0))
        rcm_sgst -= Decimal(str(getattr(note, "rcm_sgst", 0) or 0))
        rcm_igst -= Decimal(str(getattr(note, "rcm_igst", 0) or 0))
        rcm_cess -= _rcm_cess(note)
    for note in purchase_dns_rcm:
        rcm_taxable += _rcm_taxable(note)
        rcm_cgst += Decimal(str(getattr(note, "rcm_cgst", 0) or 0))
        rcm_sgst += Decimal(str(getattr(note, "rcm_sgst", 0) or 0))
        rcm_igst += Decimal(str(getattr(note, "rcm_igst", 0) or 0))
        rcm_cess += _rcm_cess(note)

    outward = gstr1["totals"]
    itc_available = {
        "igst": _money(inward_igst),
        "cgst": _money(inward_cgst),
        "sgst": _money(inward_sgst),
        "cess": _money(inward_cess),
        "total_tax": _money(inward_igst + inward_cgst + inward_sgst + inward_cess),
    }
    # BB-000279: RCM books also post Dr Input ITC — surface as provisional until 2B.
    rcm_itc_provisional = {
        "label": "provisional",
        "provisional": True,
        "claimable": False,
        "count": len(rcm),
        "igst": _money(rcm_igst),
        "cgst": _money(rcm_cgst),
        "sgst": _money(rcm_sgst),
        "cess": _money(rcm_cess),
        "total_tax": _money(rcm_igst + rcm_cgst + rcm_sgst + rcm_cess),
        "note": (
            "Provisional RCM Input ITC from books (matches RCM liability above) — "
            "not GSTR-2B matched. Do not auto-claim."
        ),
    }
    manual_review = [
        {
            "section": "4(A)(5) ITC on imports",
            "status": "manual_review",
            "note": "Import of goods/services ITC is not tracked — manual review required.",
        },
        {
            "section": "4(B) ITC reversed / ineligible",
            "status": "manual_review",
            "note": (
                "Purchase/2B ITC eligibility flags exclude ineligible/reversed from claimable ITC. "
                "Confirm Sec 17 / other reversals offline."
            ),
        },
    ]
    if not rcm:
        manual_review.insert(
            0,
            {
                "section": "3.1(d) Inward supplies liable to reverse charge",
                "status": "manual_review",
                "note": "No RCM-flagged purchases in period — confirm none offline.",
            },
        )

    # Wave 16D: claimable ITC only from matched GSTR-2B rows.
    from reporting.gstr2b import claimable_itc_from_2b

    itc_2b = claimable_itc_from_2b(company, period, company_gstin_id=stamp_id)
    from reporting.models import Gstr2bIngest

    has_2b_rows = Gstr2bIngest.objects.filter(company=company, period=period).exists()
    has_matched_2b = Gstr2bIngest.objects.filter(
        company=company,
        period=period,
        match_status=Gstr2bIngest.MatchStatus.MATCHED,
    ).exists()
    has_2b = has_matched_2b
    itc_block = {
        "provisional": not has_2b_rows,
        "claimable": bool(itc_2b.get("claimable_rows")),
        "available_from_purchases": itc_available,
        "from_gstr2b_matched": {
            "igst": _money(itc_2b["igst"]),
            "cgst": _money(itc_2b["cgst"]),
            "sgst": _money(itc_2b["sgst"]),
            "cess": _money(itc_2b.get("cess") or 0),
            "total_tax": _money(
                itc_2b["igst"] + itc_2b["cgst"] + itc_2b["sgst"] + Decimal(str(itc_2b.get("cess") or 0))
            ),
            "claimable": bool(itc_2b.get("claimable")),
            "source": itc_2b["source"],
        },
        "rcm_provisional": rcm_itc_provisional,
        "unreviewed_itc": unreviewed_itc,
        "manual_review": manual_review,
        # R4-004: the safe amount to actually claim is the lower of books ITC and
        # 2B-matched ITC per head — surface it explicitly rather than leaving the
        # "which number do I use" decision entirely to the UI.
        "books_itc": {
            "igst": _money(inward_igst),
            "cgst": _money(inward_cgst),
            "sgst": _money(inward_sgst),
            "cess": _money(inward_cess),
        },
        "recommended_claimable": (
            {
                "igst": _money(min(inward_igst, itc_2b["igst"])),
                "cgst": _money(min(inward_cgst, itc_2b["cgst"])),
                "sgst": _money(min(inward_sgst, itc_2b["sgst"])),
                "cess": _money(min(inward_cess, Decimal(str(itc_2b.get("cess") or 0)))),
                "basis": "min(books, gstr2b_matched)",
            }
            if has_2b
            else {
                "igst": _money(0),
                "cgst": _money(0),
                "sgst": _money(0),
                "cess": _money(0),
                "basis": "books_provisional_no_2b",
            }
        ),
        "note": (
            "ITC claimable only after GSTR-2B match (Wave 16D). "
            "Books provisional ITC is informational until matched."
            if not has_2b
            else "ITC claimable from matched GSTR-2B rows."
        ),
        "disclaimer": (
            "Provisional ITC: claimable only after GSTR-2B match. "
            "Books ITC is informational until matched."
            if not has_2b
            else "ITC claimable from matched GSTR-2B rows."
        ),
    }

    # Table 3.1 (a)/(b)/(c)/(d) split from GSTR-1 sections (honest labels).
    def _sec_money(sec: dict | None, key: str) -> Decimal:
        if not sec:
            return Decimal("0")
        return Decimal(str(sec.get(key, "0") or 0))

    b2b_t = outward.get("b2b") or {}
    b2cl_t = outward.get("b2cl") or {}
    b2cs_t = outward.get("b2cs") or {}
    exp_t = outward.get("exp") or {}
    sez_t = outward.get("sez") or {}
    nil_payload = gstr1.get("nil") or {}

    # Signed note nets: export-typed CDNUR → zero-rated; else taxable (a).
    a_note_taxable = Decimal("0")
    a_note_igst = Decimal("0")
    a_note_cgst = Decimal("0")
    a_note_sgst = Decimal("0")
    a_note_cess = Decimal("0")
    b_note_taxable = Decimal("0")
    b_note_igst = Decimal("0")
    b_note_cgst = Decimal("0")
    b_note_sgst = Decimal("0")
    b_note_cess = Decimal("0")
    for row in gstr1.get("cdnr") or []:
        if str(row.get("rchrg") or "N").upper() == "Y":
            continue
        sign = Decimal("-1") if row.get("note_kind") == "CREDIT" else Decimal("1")
        a_note_taxable += sign * Decimal(str(row.get("taxable_value") or 0))
        a_note_igst += sign * Decimal(str(row.get("igst") or 0))
        a_note_cgst += sign * Decimal(str(row.get("cgst") or 0))
        a_note_sgst += sign * Decimal(str(row.get("sgst") or 0))
        a_note_cess += sign * Decimal(str(row.get("cess") or 0))
    for row in gstr1.get("cdnur") or []:
        if str(row.get("rchrg") or "N").upper() == "Y":
            continue
        sign = Decimal("-1") if row.get("note_kind") == "CREDIT" else Decimal("1")
        supply = (row.get("supply_type") or "").upper()
        bucket_taxable = sign * Decimal(str(row.get("taxable_value") or 0))
        bucket_igst = sign * Decimal(str(row.get("igst") or 0))
        bucket_cgst = sign * Decimal(str(row.get("cgst") or 0))
        bucket_sgst = sign * Decimal(str(row.get("sgst") or 0))
        bucket_cess = sign * Decimal(str(row.get("cess") or 0))
        if supply in ("SEZWP", "SEZWOP", "EXPWP", "EXPWOP", "DEXP"):
            b_note_taxable += bucket_taxable
            b_note_igst += bucket_igst
            b_note_cgst += bucket_cgst
            b_note_sgst += bucket_sgst
            b_note_cess += bucket_cess
        else:
            a_note_taxable += bucket_taxable
            a_note_igst += bucket_igst
            a_note_cgst += bucket_cgst
            a_note_sgst += bucket_sgst
            a_note_cess += bucket_cess

    # (a) taxable other than zero-rated: B2B/B2CL/B2CS ± non-export notes.
    # Outward RCM supplies are included in taxable turnover, while taxes (CGST/SGST/IGST/Cess)
    # exclude RCM because liability is paid by the recipient under section 9(3).
    def _non_rcm_sum(rows, key: str) -> Decimal:
        total = Decimal("0")
        for row in rows or []:
            if str(row.get("rchrg") or "N").upper() == "Y":
                continue
            total += Decimal(str(row.get(key) or 0))
        return total

    def _all_sum(rows, key: str) -> Decimal:
        total = Decimal("0")
        for row in rows or []:
            total += Decimal(str(row.get(key) or 0))
        return total

    # Taxable turnover includes outward RCM; tax columns exclude it (recipient pays).
    a_taxable_value = (
        _all_sum(gstr1.get("b2b"), "taxable_value")
        + _sec_money(b2cl_t, "taxable_value")
        + _sec_money(b2cs_t, "taxable_value")
        + a_note_taxable
    )
    a_igst = (
        _non_rcm_sum(gstr1.get("b2b"), "igst")
        + _sec_money(b2cl_t, "igst")
        + _sec_money(b2cs_t, "igst")
        + a_note_igst
    )
    a_cgst = _non_rcm_sum(gstr1.get("b2b"), "cgst") + _sec_money(b2cs_t, "cgst") + a_note_cgst
    a_sgst = _non_rcm_sum(gstr1.get("b2b"), "sgst") + _sec_money(b2cs_t, "sgst") + a_note_sgst
    a_cess = (
        _non_rcm_sum(gstr1.get("b2b"), "cess")
        + _sec_money(b2cl_t, "cess")
        + _sec_money(b2cs_t, "cess")
        + a_note_cess
    )

    # (b) zero-rated (exports / SEZ)
    b_taxable_value = (
        _sec_money(exp_t, "taxable_value") + _sec_money(sez_t, "taxable_value") + b_note_taxable
    )
    b_igst = _sec_money(exp_t, "igst") + _sec_money(sez_t, "igst") + b_note_igst
    b_cgst = _sec_money(sez_t, "cgst") + b_note_cgst
    b_sgst = _sec_money(sez_t, "sgst") + b_note_sgst
    b_cess = _sec_money(exp_t, "cess") + _sec_money(sez_t, "cess") + b_note_cess

    # (c) nil-rated / exempt
    c_taxable_value = Decimal(str(nil_payload.get("nil_rated") or 0)) + Decimal(
        str(nil_payload.get("exempt") or 0)
    )
    # (e) non-GST outward
    d_taxable_value = Decimal(str(nil_payload.get("non_gst") or 0))

    # Table 3.2 — inter-state supplies made to unregistered persons (B2CL + IGST B2CS).
    table_32_by_pos: dict[str, dict] = {}
    for row in gstr1.get("b2cl") or []:
        pos = row.get("place_of_supply") or ""
        bucket = table_32_by_pos.setdefault(
            pos,
            {"place_of_supply": pos, "taxable_value": Decimal("0"), "igst": Decimal("0")},
        )
        bucket["taxable_value"] += Decimal(str(row.get("taxable_value") or 0))
        bucket["igst"] += Decimal(str(row.get("igst") or 0))
    for row in gstr1.get("b2cs") or []:
        igst = Decimal(str(row.get("igst") or 0))
        if igst <= 0:
            continue
        pos = row.get("place_of_supply") or ""
        bucket = table_32_by_pos.setdefault(
            pos,
            {"place_of_supply": pos, "taxable_value": Decimal("0"), "igst": Decimal("0")},
        )
        bucket["taxable_value"] += Decimal(str(row.get("taxable_value") or 0))
        bucket["igst"] += igst
    table_32 = [
        {
            "place_of_supply": pos,
            "taxable_value": _money(vals["taxable_value"]),
            "igst": _money(vals["igst"]),
        }
        for pos, vals in sorted(table_32_by_pos.items())
        if vals["taxable_value"] or vals["igst"]
    ]

    return {
        "return_type": "GSTR-3B",
        "builder_version": BUILDER_VERSION_GSTR3B,
        "period": period,
        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
        "company": gstr1["company"],
        "outward_supplies": {
            "a_taxable_other_than_zero_rated": {
                "label": "3.1(a) Outward taxable supplies (other than zero rated, nil rated and exempted)",
                "taxable_value": _money(a_taxable_value),
                "igst": _money(a_igst),
                "cgst": _money(a_cgst),
                "sgst": _money(a_sgst),
                "cess": _money(a_cess),
            },
            "b_zero_rated": {
                "label": "3.1(b) Outward taxable supplies (zero rated)",
                "taxable_value": _money(b_taxable_value),
                "igst": _money(b_igst),
                "cgst": _money(b_cgst),
                "sgst": _money(b_sgst),
                "cess": _money(b_cess),
            },
            "c_nil_rated_exempt": {
                "label": "3.1(c) Other outward supplies (nil rated, exempted)",
                "taxable_value": _money(c_taxable_value),
                "igst": "0.00",
                "cgst": "0.00",
                "sgst": "0.00",
                "cess": "0.00",
                "note": (
                    "3.1(c) = line supply_nature NIL + EXEMPT, plus 0% TAXABLE aid. "
                    "3.1(e) uses NON_GST lines."
                ),
            },
            "d_non_gst": {
                "label": "3.1(e) Non-GST outward supplies",
                "taxable_value": _money(d_taxable_value),
                "igst": "0.00",
                "cgst": "0.00",
                "sgst": "0.00",
                "cess": "0.00",
                "note": "Labeled 3.1(e) per GSTN form; (d) on the form is inward RCM.",
            },
            # Compatibility rollup (sum of a–d outward buckets).
            "taxable_value": outward["outward_taxable"],
            "igst": outward["outward_igst"],
            "cgst": outward["outward_cgst"],
            "sgst": outward["outward_sgst"],
            "cess": outward.get("outward_cess", "0.00"),
        },
        "table_3_2_inter_state_unregistered": {
            "label": "3.2 Supplies made to unregistered persons (inter-state)",
            "rows": table_32,
            "taxable_value": _money(sum((r["taxable_value"] for r in table_32_by_pos.values()), Decimal("0"))),
            "igst": _money(sum((r["igst"] for r in table_32_by_pos.values()), Decimal("0"))),
        },
        "inward_supplies": {
            "from_purchases": {
                "count": len(non_rcm),
                "taxable_value": _money(inward_taxable),
                "igst": _money(inward_igst),
                "cgst": _money(inward_cgst),
                "sgst": _money(inward_sgst),
                "cess": _money(inward_cess),
            },
            "reverse_charge": {
                "count": len(rcm),
                "taxable_value": _money(rcm_taxable),
                "igst": _money(rcm_igst),
                "cgst": _money(rcm_cgst),
                "sgst": _money(rcm_sgst),
                "cess": _money(rcm_cess),
            },
        },
        "itc": itc_block,
        "tax_on_advances": {
            "txpd": gstr1.get("txpd") or [],
            "at": gstr1.get("at") or [],
            "atadj": gstr1.get("atadj") or [],
            "fileable": False,
            "note": (
                "TXPD aid is copied from GSTR-1 unallocated advances (AT). "
                "Tax rate is unknown — do not file as GSTN TXPD without CA rate split."
            ),
        },
        "tax_payable_summary": {
            "outward_tax": _money(
                Decimal(outward["outward_igst"])
                + Decimal(outward["outward_cgst"])
                + Decimal(outward["outward_sgst"])
                + Decimal(outward.get("outward_cess", "0") or 0)
            ),
            "itc_available": itc_available["total_tax"],
            "itc_provisional": not has_2b,
            "itc_claimable": has_2b,
            # Wave 16D: subtract matched 2B ITC when present; else do not subtract provisional.
            "net_payable_hint": _money(
                Decimal(outward["outward_igst"])
                + Decimal(outward["outward_cgst"])
                + Decimal(outward["outward_sgst"])
                + Decimal(outward.get("outward_cess", "0") or 0)
                + (rcm_cgst + rcm_sgst + rcm_igst + rcm_cess)
                - (
                    itc_2b["igst"] + itc_2b["cgst"] + itc_2b["sgst"] + Decimal(str(itc_2b.get("cess") or 0))
                    if has_2b
                    else Decimal("0")
                )
            ),
            "note": (
                "Net payable subtracts matched GSTR-2B ITC when present; "
                "otherwise excludes provisional books ITC."
            ),
        },
        "issues": gstr1.get("issues", []),
        "disclaimer": "Offline export for CA review — not a GSTN portal upload file.",
    }


def build_gstr9(company, fy_label: str, *, company_gstin=None) -> dict:
    """
    FY outward + minimal inward aid (not a full GSTR-9 engine).
    fy_label like '2025-26' (April–March when fy_start_month=4).
    Prefer summing monthly GSTR-1 snapshots; fallback re-agg.
    BB-000698: company_gstin scopes monthly builders and purchase ITC.
    """
    assert_not_composition_for_regular_returns(company)
    try:
        start_y, end_yy = fy_label.split("-")
        start_year = int(start_y)
        end_year = int(end_yy) if len(end_yy) == 4 else 2000 + int(end_yy)
    except (ValueError, AttributeError) as exc:
        raise ValueError(f"Invalid FY '{fy_label}'. Expected e.g. 2025-26.") from exc

    fy_start = company.fy_start_month or 4
    months = []
    y, m = start_year, fy_start
    for _ in range(12):
        months.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            m = 1
            y += 1

    # Resolve stamp once for the FY (empty preview + explicit param).
    stamp_id = _resolve_filing_gstin_id(company, [], company_gstin=company_gstin) if company_gstin is not None else None
    if stamp_id is None and company_gstin is None:
        # Prefer primary when multi-GSTIN without explicit param — still scoped.
        from accounts.models import CompanyGstin

        primary = CompanyGstin.objects.filter(company=company, is_primary=True, is_active=True).first()
        stamp_id = primary.id if primary else None

    monthly = []
    outward_taxable = Decimal("0")
    outward_tax = Decimal("0")
    inward_taxable = Decimal("0")
    inward_tax = Decimal("0")
    hsn_fy: dict[tuple, dict] = defaultdict(
        lambda: {
            "quantity": Decimal("0"),
            "taxable_value": Decimal("0"),
            "cgst": Decimal("0"),
            "sgst": Decimal("0"),
            "igst": Decimal("0"),
            "cess": Decimal("0"),
        }
    )
    hsn_inward = new_hsn_buckets()
    nil_fy = Decimal("0")
    itc6_taxable = Decimal("0")
    itc6_cgst = Decimal("0")
    itc6_sgst = Decimal("0")
    itc6_igst = Decimal("0")
    itc6_cess = Decimal("0")
    itc7_taxable = Decimal("0")
    itc7_cgst = Decimal("0")
    itc7_sgst = Decimal("0")
    itc7_igst = Decimal("0")
    itc7_cess = Decimal("0")
    itc8a_tax = Decimal("0")
    itc8_import = Decimal("0")
    for period in months:
        snap = (
            GstReturnSnapshot.objects.filter(
                company=company,
                return_type=GstReturnSnapshot.ReturnType.GSTR1,
                period=period,
            )
            .order_by("-generated_at")
            .first()
        )
        period_ctrl = GstReturnPeriod.objects.filter(company=company, period=period).first()
        dirty = bool(period_ctrl and period_ctrl.dirty_after_snapshot)
        # Prefer snapshot only when clean; rebuild live when period was dirtied.
        if snap and stamp_id is None and not dirty:
            payload = dict(snap.payload or {})
            payload.pop("gsp_upload", None)
        else:
            payload = build_gstr1(company, period, company_gstin=stamp_id or company_gstin)
        totals = payload.get("totals", {})
        nil_fy += Decimal(str((payload.get("nil") or {}).get("taxable_value", "0") or 0))
        for hrow in payload.get("hsn", []) or []:
            key = (hrow.get("hsn", "NA"), str(hrow.get("rate", "0")), hrow.get("uqc", "OTH"))
            hsn_fy[key]["quantity"] += Decimal(str(hrow.get("quantity", "0") or 0))
            hsn_fy[key]["taxable_value"] += Decimal(str(hrow.get("taxable_value", "0") or 0))
            hsn_fy[key]["cgst"] += Decimal(str(hrow.get("cgst", "0") or 0))
            hsn_fy[key]["sgst"] += Decimal(str(hrow.get("sgst", "0") or 0))
            hsn_fy[key]["igst"] += Decimal(str(hrow.get("igst", "0") or 0))
            hsn_fy[key]["cess"] += Decimal(str(hrow.get("cess", "0") or 0))
        ot = Decimal(str(totals.get("outward_taxable", "0")))
        tax = (
            Decimal(str(totals.get("outward_cgst", "0")))
            + Decimal(str(totals.get("outward_sgst", "0")))
            + Decimal(str(totals.get("outward_igst", "0")))
            + Decimal(str(totals.get("outward_cess", "0")))
        )
        outward_taxable += ot
        outward_tax += tax

        date_from, date_to = parse_period(period)
        purchases = list(_gst_purchase_invoices(company, date_from, date_to, company_gstin_id=stamp_id))
        for inv in purchases:
            if getattr(inv, "is_reverse_charge", False):
                continue
            for item in inv.items.all():
                accumulate_hsn_line(hsn_inward, item)
        non_rcm = [p for p in purchases if not getattr(p, "is_reverse_charge", False)]
        for inv in non_rcm:
            if getattr(inv, "itc_eligibility", "") != PurchaseInvoice.ItcEligibility.CLAIMABLE:
                continue
            itc6_taxable += Decimal(str(inv.taxable_total or 0))
            itc6_cgst += Decimal(str(inv.cgst_total or 0))
            itc6_sgst += Decimal(str(inv.sgst_total or 0))
            itc6_igst += Decimal(str(inv.igst_total or 0))
            itc6_cess += Decimal(str(getattr(inv, "cess_total", 0) or 0))
        from reporting.models import Gstr2bIngest

        itc8_qs = Gstr2bIngest.objects.filter(
            company=company,
            period=period,
            match_status=Gstr2bIngest.MatchStatus.MATCHED,
            itc_eligibility=Gstr2bIngest.ItcEligibility.CLAIMABLE,
        )
        if stamp_id is not None:
            itc8_qs = itc8_qs.filter(purchase_invoice__company_gstin_id=stamp_id)
        for row in itc8_qs:
            itc8a_tax += (
                Decimal(str(row.cgst or 0))
                + Decimal(str(row.sgst or 0))
                + Decimal(str(row.igst or 0))
                + Decimal(str(getattr(row, "cess", 0) or 0))
            )
        for inv in non_rcm:
            supplier_gstin = (getattr(getattr(inv, "supplier", None), "gstin", None) or "").strip()
            notes = (getattr(inv, "notes", "") or "").upper()
            if (not supplier_gstin and Decimal(str(inv.igst_total or 0)) > 0) or "IMPORT" in notes:
                itc8_import += Decimal(str(inv.igst_total or 0)) + Decimal(str(getattr(inv, "cess_total", 0) or 0))
        period_inward_taxable = sum((inv.taxable_total for inv in non_rcm), Decimal("0"))
        period_inward_tax = sum(
            (
                Decimal(str(inv.cgst_total or 0))
                + Decimal(str(inv.sgst_total or 0))
                + Decimal(str(inv.igst_total or 0))
                for inv in non_rcm
            ),
            Decimal("0"),
        )
        purchase_cns = list(_gst_purchase_credit_notes(company, date_from, date_to, company_gstin_id=stamp_id))
        purchase_dns = list(_gst_purchase_debit_notes(company, date_from, date_to, company_gstin_id=stamp_id))
        for note in purchase_cns:
            parent = getattr(note, "purchase_invoice", None)
            if parent is not None and getattr(parent, "itc_eligibility", "") == PurchaseInvoice.ItcEligibility.CLAIMABLE:
                itc7_taxable += Decimal(str(note.taxable_total or 0))
                itc7_cgst += Decimal(str(note.cgst_total or 0))
                itc7_sgst += Decimal(str(note.sgst_total or 0))
                itc7_igst += Decimal(str(note.igst_total or 0))
                itc7_cess += Decimal(str(getattr(note, "cess_total", 0) or 0))
            period_inward_taxable -= note.taxable_total
            period_inward_tax -= (
                Decimal(str(note.cgst_total or 0))
                + Decimal(str(note.sgst_total or 0))
                + Decimal(str(note.igst_total or 0))
            )
        for note in purchase_dns:
            period_inward_taxable += note.taxable_total
            period_inward_tax += (
                Decimal(str(note.cgst_total or 0))
                + Decimal(str(note.sgst_total or 0))
                + Decimal(str(note.igst_total or 0))
            )
        inward_taxable += period_inward_taxable
        inward_tax += period_inward_tax

        monthly.append({
            "period": period,
            "outward_taxable": _money(ot),
            "outward_tax": _money(tax),
            "inward_taxable": _money(period_inward_taxable),
            "inward_tax": _money(period_inward_tax),
            "from_snapshot": bool(snap) and stamp_id is None,
        })

    return {
        "return_type": "GSTR-9",
        "aid_kind": "gstr9_worksheet_mvp",
        "title": "GSTR-9 worksheet aid (tables 4–8 + HSN 17/18)",
        "builder_version": BUILDER_VERSION_GSTR9,
        "fy": fy_label,
        "fy_end_year": end_year,
        "company_gstin_id": stamp_id,
        "company": {"name": company.name, "gstin": company.gstin or "", "state": company.state or ""},
        "monthly": monthly,
        "annual": {
            "outward_taxable": _money(outward_taxable),
            "outward_tax": _money(outward_tax),
            "inward_taxable": _money(inward_taxable),
            "inward_tax": _money(inward_tax),
        },
        "tables": {
            "4": {
                "aid_kind": "outward_supplies",
                "taxable_value": _money(outward_taxable),
                "tax": _money(outward_tax),
                "note": "Aggregated from monthly GSTR-1 aids.",
            },
            "5": {
                "aid_kind": "outward_nil_exempt_monthly_rollup",
                "taxable_value": _money(nil_fy),
                "note": "Sum of monthly GSTR-1 nil taxable aid (nil/exempt/zero-rated unsplit).",
            },
            "6": {
                "aid_kind": "itc_claimable_fy_books",
                "status": "WORKSHEET",
                "taxable_value": _money(itc6_taxable),
                "cgst": _money(itc6_cgst),
                "sgst": _money(itc6_sgst),
                "igst": _money(itc6_igst),
                "cess": _money(itc6_cess),
                "tax": _money(itc6_cgst + itc6_sgst + itc6_igst + itc6_cess),
                "note": (
                    "FY ITC from non-RCM purchases marked CLAIMABLE. Worksheet aid only — "
                    "not a filed GSTR-9 Table 6 engine."
                ),
            },
            "7": {
                "aid_kind": "itc_reversal_purchase_cn_fy",
                "status": "WORKSHEET",
                "taxable_value": _money(itc7_taxable),
                "cgst": _money(itc7_cgst),
                "sgst": _money(itc7_sgst),
                "igst": _money(itc7_igst),
                "cess": _money(itc7_cess),
                "tax": _money(itc7_cgst + itc7_sgst + itc7_igst + itc7_cess),
                "note": (
                    "FY ITC reversal aid from completed purchase credit notes on claimable "
                    "non-RCM invoices. Not a full Table 7 (rules 37/42/43) engine."
                ),
            },
            "8": {
                "aid_kind": "itc_2b_vs_books_fy",
                "status": "WORKSHEET",
                "itc_as_per_2b": _money(itc8a_tax),
                "itc_as_per_books": _money(itc6_cgst + itc6_sgst + itc6_igst + itc6_cess),
                "variance": _money(
                    (itc6_cgst + itc6_sgst + itc6_igst + itc6_cess) - itc8a_tax
                ),
                "imports_igst": _money(itc8_import),
                "tax": _money(itc8a_tax + itc8_import),
                "note": (
                    "Table 8 worksheet: MATCHED+CLAIMABLE 2B ingest vs claimable books ITC for the FY, plus "
                    "IGST on purchases without supplier GSTIN (import-like). Not GSTR-2A live."
                ),
            },
            "17": {
                "aid_kind": "hsn_outward",
                "rows": [
                    {
                        "hsn": hsn_code,
                        "rate": rate,
                        "uqc": uqc,
                        "quantity": _money(vals["quantity"]),
                        "taxable_value": _money(vals["taxable_value"]),
                        "cgst": _money(vals["cgst"]),
                        "sgst": _money(vals["sgst"]),
                        "igst": _money(vals["igst"]),
                        "cess": _money(vals.get("cess", Decimal("0"))),
                    }
                    for (hsn_code, rate, uqc), vals in sorted(hsn_fy.items())
                ],
                "note": "Aggregated from monthly GSTR-1 HSN sections.",
            },
            "18": {
                "aid_kind": "hsn_inward",
                "rows": [
                    {
                        "hsn": hsn_code,
                        "rate": rate,
                        "uqc": uqc,
                        "quantity": _money(vals["quantity"]),
                        "taxable_value": _money(vals["taxable_value"]),
                        "cgst": _money(vals["cgst"]),
                        "sgst": _money(vals["sgst"]),
                        "igst": _money(vals["igst"]),
                        "cess": _money(vals.get("cess", Decimal("0"))),
                    }
                    for (hsn_code, rate, uqc), vals in sorted(hsn_inward.items())
                ],
                "note": "Aggregated from completed GST purchase HSN lines (non-RCM).",
            },
        },
        "disclaimer": (
            "books worksheet, not filing pack. GSTR-9 worksheet aid for CA — "
            "tables 4–8 best-effort from books/2B ingest; 17/18 HSN from books. "
            "Not a complete portal upload."
        ),
        "watermark": "books worksheet, not filing pack",
        "supported": False,
    }


def to_gstn_json(payload: dict) -> dict:
    """Map BizBoard worksheet payloads to a GSTN-shaped fixture (not a live upload)."""
    rt = (payload.get("return_type") or "").upper().replace("_", "-")
    company = payload.get("company") or {}
    gstin = company.get("gstin") or ""
    period = payload.get("period") or payload.get("fy") or ""
    fp = ""
    if len(period) == 7 and period[4] == "-":
        year, month = period.split("-")
        fp = f"{month}{year}"
    # GST-02: `build_gstr1` already keeps unresolved-POS invoices out of the
    # section buckets, but strip any "NA"/"N/A"/blank place-of-supply row here as
    # a last line of defence — the portal rejects those and this JSON must never
    # carry one.
    def _drop_unresolved_pos(rows):
        if not isinstance(rows, list):
            return rows
        cleaned = []
        for row in rows:
            if isinstance(row, dict) and "place_of_supply" in row:
                pos = str(row.get("place_of_supply") or "").strip().upper()
                if pos in ("", "NA", "N/A"):
                    continue
            cleaned.append(row)
        return cleaned

    shaped = {
        "gstin": gstin,
        "fp": fp,
        "fy": payload.get("fy") or "",
        "return_type": rt,
        "builder_version": payload.get("builder_version"),
        "b2b": _drop_unresolved_pos(payload.get("b2b") or []),
        "b2cl": _drop_unresolved_pos(payload.get("b2cl") or []),
        "b2cs": _drop_unresolved_pos(payload.get("b2cs") or []),
        "cdnr": _drop_unresolved_pos(payload.get("cdnr") or []),
        "cdnur": _drop_unresolved_pos(payload.get("cdnur") or []),
        "exp": payload.get("exp") or [],
        "hsn": payload.get("hsn") or payload.get("tables", {}).get("17", {}).get("rows") or [],
        "supecom": payload.get("supecom") or {},
        "tables": payload.get("tables") or {},
        "disclaimer": (
            "GSTN-shaped fixture for CA tooling — not a GSTN portal upload file. "
            "Do not submit this JSON to GSTN."
        ),
    }
    return shaped


def canonical_payload_bytes(payload: dict) -> bytes:
    return json.dumps(payload, sort_keys=True, separators=(",", ":"), default=str).encode("utf-8")


def content_hash(payload: dict) -> str:
    return hashlib.sha256(canonical_payload_bytes(payload)).hexdigest()


def persist_snapshot(company, return_type: str, period: str, payload: dict, user=None) -> GstReturnSnapshot:
    rt_map = {
        "GSTR-1": GstReturnSnapshot.ReturnType.GSTR1,
        "GSTR1": GstReturnSnapshot.ReturnType.GSTR1,
        "GSTR-3B": GstReturnSnapshot.ReturnType.GSTR3B,
        "GSTR3B": GstReturnSnapshot.ReturnType.GSTR3B,
        "GSTR-9": GstReturnSnapshot.ReturnType.GSTR9,
        "GSTR9": GstReturnSnapshot.ReturnType.GSTR9,
    }
    rt = rt_map.get(return_type, return_type)
    payload = dict(payload or {})
    payload.pop("gsp_upload", None)
    version = payload.get("builder_version") or BUILDER_VERSION_GSTR1
    h = content_hash(payload)
    # BB-000064: replace existing row for same company+type+period (one snapshot).
    existing = GstReturnSnapshot.objects.filter(
        company=company, return_type=rt, period=period,
    ).first()
    if existing and existing.content_hash == h:
        return existing
    obj, _ = GstReturnSnapshot.objects.update_or_create(
        company=company,
        return_type=rt,
        period=period,
        defaults={
            "payload": payload,
            "content_hash": h,
            "builder_version": version,
            "generated_at": timezone.now(),
            "generated_by": user,
        },
    )
    return obj


def _sheet_from_rows(workbook, title: str, rows: list[dict]):
    sheet = workbook.create_sheet(title=title[:31])
    if not rows:
        sheet.append(["No records"])
        return
    headers = list(rows[0].keys())
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])


def gstr_return_to_xlsx(payload: dict) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.remove(workbook.active)
    meta = workbook.create_sheet("Summary")
    meta.append(["Return", payload.get("return_type", "")])
    meta.append(["Period", payload.get("period") or payload.get("fy", "")])
    meta.append(["Builder", payload.get("builder_version", "")])
    meta.append(["Company", payload.get("company", {}).get("name", "")])
    meta.append(["GSTIN", payload.get("company", {}).get("gstin", "")])
    meta.append(["Disclaimer", payload.get("disclaimer", "")])

    rtype = payload.get("return_type")
    if rtype == "GSTR-1":
        _sheet_from_rows(workbook, "B2B", payload.get("b2b", []))
        _sheet_from_rows(workbook, "B2CL", payload.get("b2cl", []))
        _sheet_from_rows(workbook, "B2CS", payload.get("b2cs", []))
        _sheet_from_rows(workbook, "CDNR", payload.get("cdnr", []))
        _sheet_from_rows(workbook, "CDNUR", payload.get("cdnur", []))
        _sheet_from_rows(workbook, "HSN", payload.get("hsn", []))
        _sheet_from_rows(workbook, "Issues", payload.get("issues", []))
        docs = payload.get("docs", {})
        doc_sheet = workbook.create_sheet("DOCS")
        for key, value in docs.items():
            doc_sheet.append([key, value])
        totals = payload.get("totals", {})
        total_sheet = workbook.create_sheet("Totals")
        for section, values in totals.items():
            if isinstance(values, dict):
                total_sheet.append([section])
                for k, v in values.items():
                    total_sheet.append(["", k, v])
            else:
                total_sheet.append([section, values])
    elif rtype == "GSTR-9":
        _sheet_from_rows(workbook, "Monthly", payload.get("monthly", []))
        annual = payload.get("annual", {})
        sheet = workbook.create_sheet("Annual")
        for k, v in annual.items():
            sheet.append([k, v])
    else:
        outward = payload.get("outward_supplies", {})
        outward_sheet = workbook.create_sheet("Outward")
        for key, value in outward.items():
            if isinstance(value, dict):
                outward_sheet.append([key])
                for k, v in value.items():
                    outward_sheet.append(["", k, v])
            else:
                outward_sheet.append([key, value])
        table32 = payload.get("table_3_2_inter_state_unregistered") or {}
        t32_sheet = workbook.create_sheet("Table32")
        t32_sheet.append(["place_of_supply", "taxable_value", "igst"])
        for row in table32.get("rows") or []:
            t32_sheet.append([
                row.get("place_of_supply"),
                row.get("taxable_value"),
                row.get("igst"),
            ])
        inward = payload.get("inward_supplies", {}).get("from_purchases", {})
        inward_sheet = workbook.create_sheet("Inward")
        for key, value in inward.items():
            inward_sheet.append([key, value])
        rcm = payload.get("inward_supplies", {}).get("reverse_charge", {})
        rcm_sheet = workbook.create_sheet("RCM")
        for key, value in rcm.items():
            rcm_sheet.append([key, value])
        itc_sheet = workbook.create_sheet("ITC")
        for key, value in payload.get("itc", {}).get("available_from_purchases", {}).items():
            itc_sheet.append([key, value])
        review_sheet = workbook.create_sheet("ManualReview")
        review_sheet.append(["section", "status", "note"])
        for row in payload.get("itc", {}).get("manual_review", []):
            review_sheet.append([row.get("section"), row.get("status"), row.get("note")])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_ca_pack_zip(company, period: str, gstr1: dict | None = None) -> bytes:
    from accounts.models import Company
    from reporting.gstr2b import build_cmp08, build_gstr4, claimable_itc_from_2b, match_gstr2b_to_purchases
    from reporting.models import Gstr2bIngest

    match_summary = match_gstr2b_to_purchases(company, period, persist=False)
    itc_2b = claimable_itc_from_2b(company, period)
    unmatched = Gstr2bIngest.objects.filter(
        company=company, period=period, match_status=Gstr2bIngest.MatchStatus.UNMATCHED
    ).count()
    summary = {
        "period": period,
        "gstr2b_match": match_summary,
        "gstr2b_unmatched": unmatched,
        "claimable_itc": {k: str(v) for k, v in itc_2b.items()},
    }
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        if company.registration_type == Company.RegistrationType.COMPOSITION:
            cmp08 = build_cmp08(company, period)
            y = int(period[:4])
            gstr4 = build_gstr4(company, f"{y}-{str(y + 1)[-2:]}")
            zf.writestr(f"cmp08-{period}.json", json.dumps(cmp08, indent=2, default=str))
            zf.writestr(f"gstr4-{period}.json", json.dumps(gstr4, indent=2, default=str))
        else:
            g1 = gstr1 if gstr1 is not None else build_gstr1(company, period)
            g3 = build_gstr3b(company, period, gstr1=g1)
            zf.writestr(f"gstr1-{period}.xlsx", gstr_return_to_xlsx(g1))
            zf.writestr(f"gstr3b-{period}.xlsx", gstr_return_to_xlsx(g3))
            zf.writestr(f"gstr1-{period}.json", json.dumps(g1, indent=2, default=str))
            zf.writestr(f"gstr3b-{period}.json", json.dumps(g3, indent=2, default=str))
        zf.writestr(f"gstr2b-match-{period}.json", json.dumps(summary, indent=2, default=str))
    return buf.getvalue()
