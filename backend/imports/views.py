from django.http import HttpResponse, StreamingHttpResponse
from rest_framework import mixins, status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

import csv
import io
from django.db import transaction

from core.exceptions import BusinessRuleError
from core.csv_utils import csv_safe
from core.idempotency import (
    begin_record,
    release_record,
    store_record,
)
from core.models import FileAsset
from core.permissions import CanImport, HasCompany, IsOwner, get_company_user
from core.services.files import FileService
from masters.models import Customer, Supplier

from .models import ImportJob
from .serializers import ImportJobSerializer
from .services import (
    BillImportService,
    ImportService,
    OPENING_LOTS_COLUMNS,
    OPENING_SERIALS_COLUMNS,
    products_item_columns,
    products_template_csv,
)

BILL_IMAGE_PDF_TYPES = {
    "application/pdf",
    "image/jpeg",
    "image/jpg",
    "image/png",
    "image/webp",
    "image/gif",
}
BILL_STRUCTURED_TYPES = {
    "text/csv",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}
BILL_CONTENT_TYPES = BILL_IMAGE_PDF_TYPES | BILL_STRUCTURED_TYPES
BILL_STRUCTURED_EXTENSIONS = (".csv", ".xlsx", ".xlsm")


class ImportJobViewSet(
    mixins.CreateModelMixin, mixins.ListModelMixin, mixins.RetrieveModelMixin,
    viewsets.GenericViewSet,
):
    serializer_class = ImportJobSerializer
    queryset = ImportJob.objects.all()
    permission_classes = [IsAuthenticated, HasCompany, CanImport]

    @property
    def company(self):
        return get_company_user(self.request).company

    def get_queryset(self):
        return self.queryset.filter(company=self.company)

    def create(self, request, *args, **kwargs):
        """Upload + validate (CSV masters) or start bill ingestion (purchase/sales bill)."""
        kind = (request.data.get("kind") or "").upper()
        if kind not in ImportJob.Kind.values:
            raise BusinessRuleError(f"kind must be one of {', '.join(ImportJob.Kind.values)}.")
        uploaded = request.FILES.get("file")
        if not uploaded:
            raise BusinessRuleError("A file is required.")

        supplier = None
        supplier_id = request.data.get("supplier_id") or request.data.get("supplier")
        if supplier_id:
            try:
                supplier = Supplier.objects.get(pk=supplier_id, company=self.company)
            except (Supplier.DoesNotExist, ValueError, TypeError):
                raise BusinessRuleError("Invalid supplier.")

        customer = None
        customer_id = request.data.get("customer_id") or request.data.get("customer")
        if customer_id:
            try:
                customer = Customer.objects.get(pk=customer_id, company=self.company)
            except (Customer.DoesNotExist, ValueError, TypeError):
                raise BusinessRuleError("Invalid customer.")

        raw_key = (request.headers.get("Idempotency-Key") or "").strip()
        # A bill *upload* is never idempotency-deduped — re-uploading the same
        # photo deliberately starts a fresh extraction job (the paired
        # store_record below is likewise skipped for BILL_KINDS). Only structured
        # master imports claim the key.
        use_idempotency = bool(raw_key) and kind not in ImportJob.BILL_KINDS
        claimed = None
        if use_idempotency:
            claimed = begin_record(
                company=self.company, scope="import_job_create", raw_key=raw_key
            )
            if isinstance(claimed, Response):
                return claimed

        created_ok = False
        try:
            if kind in ImportJob.BILL_KINDS:
                content_type = (getattr(uploaded, "content_type", "") or "").lower()
                name = (uploaded.name or "").lower()
                is_structured = content_type in BILL_STRUCTURED_TYPES or name.endswith(BILL_STRUCTURED_EXTENSIONS)
                is_image_pdf = content_type in BILL_IMAGE_PDF_TYPES or name.endswith(
                    (".pdf", ".png", ".jpg", ".jpeg", ".webp", ".gif")
                )
                if not (is_structured or is_image_pdf):
                    raise BusinessRuleError(
                        "Bill must be a PDF, image, CSV, or XLSX export from your supplier's system."
                    )
                with transaction.atomic():
                    asset = FileService.store_upload(
                        company=self.company, uploaded_file=uploaded,
                        kind=FileAsset.Kind.IMPORT, user=request.user,
                    )
                    job = ImportJob.objects.create(
                        company=self.company, kind=kind, file=asset, supplier=supplier, customer=customer,
                        created_by=request.user, updated_by=request.user,
                    )
                    if is_structured:
                        BillImportService.parse_structured_file(job)
                    else:
                        BillImportService.start_extraction(job)
                    job.refresh_from_db()
                    response = Response(self.get_serializer(job).data, status=status.HTTP_201_CREATED)
            else:
                # Master CSV/XLSX: atomic so a validate failure leaves no orphan job/asset.
                with transaction.atomic():
                    asset = FileService.store_upload(
                        company=self.company, uploaded_file=uploaded,
                        kind=FileAsset.Kind.IMPORT, user=request.user,
                    )
                    job = ImportJob.objects.create(
                        company=self.company, kind=kind, file=asset,
                        created_by=request.user, updated_by=request.user,
                    )
                    ImportService.validate(job)
                response = Response(self.get_serializer(job).data, status=status.HTTP_201_CREATED)

            if use_idempotency:
                store_record(
                    company=self.company,
                    scope="import_job_create",
                    raw_key=raw_key,
                    response=response,
                    resource_id=str(job.pk),
                )
                created_ok = True
            return response
        finally:
            if (
                use_idempotency
                and claimed is not None
                and not isinstance(claimed, Response)
                and not created_ok
            ):
                release_record(company=self.company, scope="import_job_create", raw_key=raw_key)
    @action(detail=True, methods=["post"], url_path="retry-extract")
    def retry_extract(self, request, pk=None):
        """Re-run LLM extraction for FAILED (or UPLOADED) bill jobs."""
        job = self.get_object()
        BillImportService.start_extraction(job)
        job.refresh_from_db()
        return Response(self.get_serializer(job).data)

    @action(detail=True, methods=["post"], url_path="clarify")
    def clarify(self, request, pk=None):
        """Answer document-level clarification questions (§4.3) and move the
        job from NEEDS_CLARIFICATION to PREVIEWED."""
        job = self.get_object()
        answers = request.data.get("answers")
        if not isinstance(answers, dict):
            raise BusinessRuleError("answers must be an object of {field: value}.")
        BillImportService.answer_clarifications(job, answers, user=request.user)
        job.refresh_from_db()
        return Response(self.get_serializer(job).data)

    @action(detail=True, methods=["post", "patch"], url_path="preview")
    def update_preview(self, request, pk=None):
        """Edit extracted bill preview lines / party before commit."""
        job = self.get_object()
        BillImportService.update_preview(job, request.data, user=request.user)
        return Response(self.get_serializer(job).data)

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated, HasCompany, CanImport])
    def commit(self, request, pk=None):
        """Commit requires Owner or explicit import permission (§5.5)."""
        raw_key = (request.headers.get("Idempotency-Key") or "").strip()
        if raw_key:
            # B3-013: claim an in-flight placeholder (not a bare get_record) so a
            # concurrent retry while the first commit is still running replays
            # instead of committing a second draft invoice.
            claimed = begin_record(
                company=self.company, scope="import_job_commit", raw_key=raw_key
            )
            if isinstance(claimed, Response):
                return claimed

        try:
            job = self.get_object()
            if job.kind in ImportJob.BILL_KINDS and request.data:
                # Allow last-minute supplier/customer / line tweaks on commit body.
                BillImportService.update_preview(job, request.data, user=request.user)
                job.refresh_from_db()
            result = ImportService.commit(job, request.user)
        except Exception:
            if raw_key:
                release_record(
                    company=self.company, scope="import_job_commit", raw_key=raw_key
                )
            raise
        job.refresh_from_db()
        if isinstance(result, dict):
            response = Response({
                "created": result.get("created", 0),
                "products_created": result.get("products_created", 0),
                "purchase_invoice_id": result.get("purchase_invoice_id"),
                "sales_invoice_id": result.get("sales_invoice_id"),
                "status": job.status,
                "error_rows": job.error_rows,
            })
        else:
            response = Response(
                {"created": result, "status": job.status, "error_rows": job.error_rows}
            )
        if raw_key:
            store_record(
                company=self.company,
                scope="import_job_commit",
                raw_key=raw_key,
                response=response,
                resource_id=str(job.pk),
            )
        return response

    @action(detail=True, methods=["post"], url_path="void")
    def void_import(self, request, pk=None):
        """Owner recovery: reverse a committed PRODUCTS/OPENING_STOCK import."""
        if not IsOwner().has_permission(request, self):
            raise BusinessRuleError("Owner role required to void an import.")
        job = self.get_object()
        ImportService.void(job, request.user)
        job.refresh_from_db()
        return Response(self.get_serializer(job).data)

    @action(detail=True, methods=["post"], url_path="void-rows")
    def void_rows(self, request, pk=None):
        """Reverse opening-stock (and product cleanup) for specific SKUs only."""
        if not IsOwner().has_permission(request, self):
            raise BusinessRuleError("Owner role required to void import rows.")
        job = self.get_object()
        skus = request.data.get("skus") or []
        if not isinstance(skus, list):
            raise BusinessRuleError("skus must be an array.")
        result = ImportService.void_rows(job, request.user, skus=[str(s) for s in skus])
        job = result["job"]
        job.refresh_from_db()
        payload = dict(self.get_serializer(job).data)
        payload["void_results"] = {"voided": result["voided"], "blocked": result["blocked"]}
        return Response(payload)

    @action(detail=True, methods=["get"], url_path="errors")
    def error_report(self, request, pk=None):
        job = self.get_object()
        fmt = (request.query_params.get("as") or request.query_params.get("format") or "").lower()
        if fmt == "csv":
            def generate():
                buf = io.StringIO()
                writer = csv.writer(buf)
                writer.writerow(["row", "errors", "data"])
                yield buf.getvalue()
                for err in job.errors or []:
                    if not isinstance(err, dict):
                        continue
                    buf.seek(0)
                    buf.truncate(0)
                    msgs = err.get("errors") or []
                    msg = "; ".join(msgs) if isinstance(msgs, list) else str(msgs)
                    data = err.get("data") or {}
                    data_str = (
                        " | ".join(f"{k}={v}" for k, v in data.items() if v)
                        if isinstance(data, dict)
                        else str(data)
                    )
                    writer.writerow([err.get("row", ""), csv_safe(msg), csv_safe(data_str)])
                    yield buf.getvalue()

            filename = f"{(job.kind or 'import').lower()}_import_errors.csv"
            response = StreamingHttpResponse(generate(), content_type="text/csv")
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response
        return Response({"errors": job.errors, "error_rows": job.error_rows})

    @action(detail=False, methods=["get"], url_path="template")
    def template(self, request):
        """Download the PRODUCTS template (xlsx with extra sheets, or CSV with the same item columns)."""
        kind = (request.query_params.get("kind") or "PRODUCTS").upper()
        if kind != "PRODUCTS":
            raise BusinessRuleError("An item template is currently published for PRODUCTS only.")
        company = get_company_user(request).company
        as_fmt = (request.query_params.get("as") or "").strip().lower()
        item_columns = products_item_columns(company)
        if as_fmt == "csv":
            response = HttpResponse(products_template_csv(company), content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="products_template.csv"'
            return response
        from openpyxl import Workbook

        workbook = Workbook()
        items = workbook.active
        items.title = "items"
        items.append(list(item_columns))
        items.append([""] * len(item_columns))
        items.freeze_panes = "A2"
        lots = workbook.create_sheet("opening_lots")
        lots.append(list(OPENING_LOTS_COLUMNS))
        lots.freeze_panes = "A2"
        serials = workbook.create_sheet("opening_serials")
        serials.append(list(OPENING_SERIALS_COLUMNS))
        serials.freeze_panes = "A2"
        notes = workbook.create_sheet("notes")
        notes.append(["Item import notes"])
        notes.append(["Only name is required. Other columns match Create Item: Basic, Stock, Pricing, Custom."])
        notes.append(["product_type: Goods or Service. Service rows cannot have opening stock, godown, batch, or serials."])
        notes.append(["opening_stock on this sheet posts to the default godown unless godown is filled."])
        notes.append(["track_batch Yes (or batch_no + expiry_date) posts a lot. Use opening_lots for multiple godowns/lots per SKU."])
        notes.append(["track_serial Yes with opening stock requires serial_no (comma-separated) or an opening_serials sheet. Quantity must equal the number of serials."])
        notes.append(["Blank godown uses the company default. Unknown godown names fail the whole job."])
        notes.append(["Custom field columns use the labels from Item Settings. Only active fields are included."])
        buffer = io.BytesIO()
        workbook.save(buffer)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="products_import_template.xlsx"'
        return response
