"""Import Service — CSV pipeline + purchase-bill LLM extraction."""

import csv
import hashlib
import io
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation

from django.conf import settings
from django.db import IntegrityError, transaction
from django.db.models import Q
from django.utils import timezone
from django.utils.dateparse import parse_date

from accounts.models import CompanyGstin
from core.exceptions import BusinessRuleError
from core.help_codes import HelpCode
from core.services.audit import AuditService
from core.services.files import CSV_UTF8_HINT
from core.validators import ALLOWED_GST_RATES, GSTIN_RE, HSN_RE, validate_gst_rate
from inventory.models import MovementType, StockMovement
from inventory.services import InventoryService
from masters.models import Brand, Category, Customer, Product, Supplier, Unit
from purchases.models import PurchaseInvoice
from purchases.services import PurchaseService
from sales.models import SalesInvoice
from sales.services import SalesService

from .models import ImportJob, SupplierBillTemplate
from .qty_formula import (
    apply_qty_formula,
    collect_extras,
    count_pool,
    detect_qty_clarifications,
    infer_qty_formula,
)

OCR_BILL_MIN_CONFIDENCE = float(getattr(settings, "OCR_BILL_MIN_CONFIDENCE", 0.7) or 0.7)
BILL_TEMPLATE_SIGNATURE_MATCH_THRESHOLD = 0.7
MAX_IMPORT_ROWS = 20_000
MAX_IMPORT_CELLS = 500_000

REQUIRED_COLUMNS = {
    ImportJob.Kind.CUSTOMERS: ["name"],
    ImportJob.Kind.SUPPLIERS: ["name"],
    ImportJob.Kind.PRODUCTS: ["name"],
    ImportJob.Kind.OPENING_STOCK: ["sku", "quantity"],
}

# Fuzzy header aliases for master CSV/XLSX imports (Tally / POS / Excel exports).
MASTER_COLUMN_ALIASES = {
    "name": ["name", "product name", "item", "item name", "item name*", "item description", "product"],
    "sku": ["sku", "item code", "product code", "pcode", "code", "itemcode"],
    "barcode": ["barcode", "ean", "bar code"],
    "hsn_code": ["hsn_code", "hsn", "hsn code", "hsn/sac", "hsn_sac"],
    "description": ["description", "item desc", "product description"],
    "gst_rate": ["gst_rate", "gst", "gst%", "tax_rate", "tax%", "gst rate", "gst tax rate(%)", "gst tax rate"],
    "purchase_price": ["purchase_price", "purchase price", "cost", "buy price", "purchase rate"],
    "selling_price": ["selling_price", "selling price", "sale price", "sell price", "rate", "sales price"],
    "mrp": ["mrp"],
    "wholesale_price": ["wholesale_price", "wholesale price", "wholesale rate", "wholesale"],
    "default_discount_percent": ["default_discount_percent", "discount", "discount %", "default discount"],
    "reorder_level": ["reorder_level", "reorder level", "reorder", "min stock", "low stock alert quantity"],
    "quantity": ["quantity", "qty", "pcs", "pieces", "stock", "stock qty"],
    "unit_cost": ["unit_cost", "unit cost", "cost price", "opening cost"],
    "opening_stock": ["opening_stock", "opening stock", "opening qty", "opening quantity", "current stock"],
    "unit": ["unit", "unit_name", "uom", "uqc"],
    "alternate_unit": ["alternate_unit", "alternate unit", "alt unit", "alt_uom"],
    "conversion_rate": ["conversion_rate", "conversion rate", "conversion"],
    "product_type": ["product_type", "item type", "item_type", "type"],
    "track_inventory": ["track_inventory", "track inventory", "track stock"],
    "track_batch": ["track_batch", "track batch", "batch tracking"],
    "track_serial": ["track_serial", "track serial", "serial tracking"],
    "tracking": ["tracking", "tracking mode", "tracking_mode"],
    "selling_tax_inclusive": ["selling_tax_inclusive", "sales tax inclusive", "sales tax inclusive?"],
    "purchase_tax_inclusive": ["purchase_tax_inclusive", "purchase tax inclusive"],
    "category": ["category"],
    "godown": ["godown", "warehouse", "godown name", "warehouse name"],
    "batch_no": ["batch_no", "batch no", "batch", "lot", "lot no"],
    "expiry_date": ["expiry_date", "expiry date", "expiry", "exp date"],
    "manufacturing_date": ["manufacturing_date", "manufacturing date", "mfg", "mfg date"],
    "as_of": ["as_of", "as of", "as of date", "opening date"],
    "serial_no": ["serial_no", "serial", "serial number", "serial numbers"],
    "phone": ["phone", "mobile", "contact", "phone number"],
    "email": ["email", "e-mail"],
    "gstin": ["gstin", "gst no", "gstin number"],
    "state": ["state"],
    "address": ["address", "billing_address", "billing address"],
}

# Canonical PRODUCTS template — same headers for CSV and the xlsx `items` sheet.
PRODUCTS_ITEM_COLUMNS = [
    "name", "sku", "barcode", "hsn_code", "description", "category", "unit",
    "alternate_unit", "conversion_rate", "gst_rate", "purchase_price",
    "purchase_tax_inclusive", "selling_price", "selling_tax_inclusive", "mrp",
    "wholesale_price", "default_discount_percent", "reorder_level", "product_type",
    "track_inventory", "track_batch", "track_serial", "godown", "opening_stock",
    "unit_cost", "batch_no", "expiry_date", "manufacturing_date", "as_of", "serial_no",
]
PRODUCTS_ITEM_SAMPLE_ROWS = [
    [""] * len(PRODUCTS_ITEM_COLUMNS),
]
OPENING_LOTS_COLUMNS = [
    "sku", "godown", "quantity", "as_of", "batch_no", "expiry_date", "manufacturing_date", "unit_cost",
]
OPENING_SERIALS_COLUMNS = ["sku", "godown", "serial_no", "as_of", "unit_cost"]


def products_item_columns(company=None) -> list[str]:
    from masters.custom_fields import active_defs

    columns = list(PRODUCTS_ITEM_COLUMNS)
    if company is None:
        return columns
    seen = {col.casefold() for col in columns}
    for row in active_defs(company):
        label = (row.get("label") or row.get("key") or "").strip()
        if label and label.casefold() not in seen:
            columns.append(label)
            seen.add(label.casefold())
    return columns


def products_template_csv(company=None) -> str:
    columns = products_item_columns(company)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(columns)
    writer.writerow([""] * len(columns))
    return buf.getvalue()


def _optional_bool(value, default=None):
    raw = str(value or "").strip().casefold()
    if not raw:
        return default
    if raw in {"yes", "true", "1", "y", "on"}:
        return True
    if raw in {"no", "false", "0", "n", "off"}:
        return False
    return default


def _is_service_type(row) -> bool:
    return str(row.get("product_type") or row.get("item_type") or "").strip().casefold() in {
        "service",
        "services",
    }


def _tracking_from_row(row) -> tuple[bool, bool]:
    tracking = str(row.get("tracking") or row.get("tracking_mode") or "").strip().casefold()
    track_batch = _optional_bool(row.get("track_batch"), default=None)
    track_serial = _optional_bool(row.get("track_serial"), default=None)
    if track_batch is None:
        track_batch = tracking in {"batch", "expiry", "batch / expiry", "batch/expiry"} or bool(
            str(row.get("batch_no") or "").strip()
        )
    if track_serial is None:
        track_serial = tracking in {"serial"} or bool(str(row.get("serial_no") or "").strip())
    return bool(track_batch), bool(track_serial)


def _custom_field_row_errors(row, header_map, defs) -> list[str]:
    from masters.custom_fields import values_from_row

    errors = []
    values = values_from_row(row, header_map)
    by_key = {d["key"]: d for d in defs if d.get("active")}
    for key, text in values.items():
        spec = by_key.get(key)
        if not spec or spec.get("type") != "list":
            continue
        options = spec.get("options") or []
        if not any(str(opt).casefold() == text.casefold() for opt in options):
            errors.append(f"'{text}' is not a valid option for '{spec.get('label') or key}'.")
    return errors


def _format_validation_detail(detail) -> str:
    if isinstance(detail, dict):
        for value in detail.values():
            return _format_validation_detail(value)
        return "Invalid custom field values."
    if isinstance(detail, (list, tuple)):
        for value in detail:
            text = _format_validation_detail(value)
            if text:
                return text
        return "Invalid custom field values."
    return str(detail)


def _custom_fields_for_commit(job, row, existing=None) -> dict:
    from rest_framework.exceptions import ValidationError as DRFValidationError

    from masters.custom_fields import coerce_values, values_from_row

    header_map = job.custom_field_header_map or {}
    snapshot = job.custom_field_defs_snapshot or []
    if not header_map:
        return dict(existing or {})
    job.company.refresh_from_db(fields=["item_custom_field_defs"])
    live = {
        str(spec.get("key") or "").casefold()
        for spec in (job.company.item_custom_field_defs or [])
        if isinstance(spec, dict) and spec.get("key")
    }
    for dest in header_map.values():
        if str(dest).casefold() not in live:
            raise BusinessRuleError(f"Custom field '{dest}' no longer exists.")
    values = values_from_row(row, header_map)
    try:
        return coerce_values(values, snapshot, existing, replace_active=False)
    except DRFValidationError as exc:
        raise BusinessRuleError(_format_validation_detail(exc.detail)) from exc


def _parse_serial_numbers(value) -> list[str]:
    raw = str(value or "").strip()
    if not raw:
        return []
    return [part.strip() for part in raw.replace(";", ",").split(",") if part.strip()]


def _extra_serial_counts(extra_sheets) -> dict[str, int]:
    counts: dict[str, int] = {}
    for raw in (extra_sheets or {}).get("opening_serials") or []:
        row = {str(k).strip().lower(): v for k, v in (raw or {}).items()}
        sku = str(row.get("sku") or row.get("item code") or "").strip().casefold()
        serial = str(row.get("serial_no") or row.get("serial") or "").strip()
        if sku and serial:
            counts[sku] = counts.get(sku, 0) + 1
    return counts


def _lots_from_item_rows(preview: list[dict]) -> list[dict]:
    lots = []
    for row in preview:
        _track_batch, track_serial = _tracking_from_row(row)
        if track_serial or str(row.get("serial_no") or "").strip():
            continue
        lot_used = any(
            str(row.get(key) or "").strip()
            for key in ("godown", "batch_no", "expiry_date", "manufacturing_date", "as_of")
        )
        qty = str(row.get("opening_stock") or row.get("quantity") or "").strip()
        if not lot_used or not qty:
            continue
        lots.append(
            {
                "sku": row.get("sku") or "",
                "godown": row.get("godown") or "",
                "quantity": qty,
                "as_of": row.get("as_of") or "",
                "batch_no": row.get("batch_no") or "",
                "expiry_date": row.get("expiry_date") or "",
                "manufacturing_date": row.get("manufacturing_date") or "",
                "unit_cost": row.get("unit_cost") or "",
            }
        )
    return lots


def _serials_from_item_rows(preview: list[dict]) -> list[dict]:
    serials = []
    for row in preview:
        numbers = _parse_serial_numbers(row.get("serial_no"))
        if not numbers:
            continue
        for serial in numbers:
            serials.append(
                {
                    "sku": row.get("sku") or "",
                    "godown": row.get("godown") or "",
                    "serial_no": serial,
                    "as_of": row.get("as_of") or "",
                    "unit_cost": row.get("unit_cost") or "",
                }
            )
    return serials

ALLOWED_GST = {Decimal(r) for r in ALLOWED_GST_RATES}

BULK_BATCH = 500


def _decode_csv_text(raw: bytes) -> str:
    """UTF-8 (with BOM) first, then Windows-1252 — Excel's default CSV export."""
    try:
        return raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        pass
    try:
        return raw.decode("cp1252")
    except UnicodeDecodeError:
        raise BusinessRuleError(CSV_UTF8_HINT)


def resolve_master_column_mappings(fieldnames: list[str]) -> list[dict]:
    """Non-identity alias resolutions: original header → canonical field."""
    present = {(f or "").strip().lower() for f in fieldnames if f}
    mappings = []
    claimed = set()
    for field, aliases in MASTER_COLUMN_ALIASES.items():
        if field in present or field in claimed:
            continue
        for alias in aliases:
            if alias != field and alias in present:
                mappings.append({"source": alias, "target": field})
                claimed.add(field)
                break
    return mappings


def _line_get(line: dict, *keys: str, default: str = "") -> str:
    for key in keys:
        if key in line and line[key] is not None:
            return str(line[key]).strip()
    return default


def _map_master_row(row: dict) -> dict:
    """Map fuzzy / aliased headers onto canonical master field names."""
    normalized = {(k or "").strip().lower(): ("" if v is None else str(v)).strip() for k, v in row.items()}
    mapped = dict(normalized)
    for field, aliases in MASTER_COLUMN_ALIASES.items():
        if mapped.get(field):
            continue
        for alias in aliases:
            if alias in normalized and normalized[alias]:
                mapped[field] = normalized[alias]
                break
    return mapped


def _released_barcodes(rows, products_by_sku) -> set[str]:
    """Barcodes existing products are giving up because the file assigns a different one."""
    released = set()
    for raw_row in rows:
        row = _map_master_row(raw_row)
        sku = (row.get("sku") or "").strip()
        barcode = (row.get("barcode") or "").strip()
        if not sku or not barcode:
            continue
        product = products_by_sku.get(sku.casefold())
        if product is None:
            continue
        old = (product.barcode or "").strip().casefold()
        new = barcode.casefold()
        if old and old != new:
            released.add(old)
    return released


def _validate_row(
    kind,
    row,
    company,
    *,
    seen_skus=None,
    seen_barcodes=None,
    existing_skus=None,
    existing_barcodes=None,
    products_by_sku=None,
    skus_with_opening=None,
    extra_serial_counts=None,
):
    """Validate one CSV row. seen_* sets track within-file duplicates for products."""
    errors = []
    if kind in (ImportJob.Kind.CUSTOMERS, ImportJob.Kind.SUPPLIERS, ImportJob.Kind.PRODUCTS):
        if not (row.get("name") or "").strip():
            errors.append("name is required")
        gstin = (row.get("gstin") or "").strip()
        if gstin and not GSTIN_RE.match(gstin):
            errors.append("invalid GSTIN format")
    if kind == ImportJob.Kind.PRODUCTS:
        hsn = (row.get("hsn_code") or "").strip()
        if hsn and not HSN_RE.match(hsn):
            errors.append("invalid HSN code")
        for field in (
            "gst_rate", "purchase_price", "selling_price", "mrp", "wholesale_price",
            "default_discount_percent", "conversion_rate", "reorder_level",
            "opening_stock", "quantity", "unit_cost",
        ):
            value = (row.get(field) or "").strip()
            if value:
                try:
                    num = Decimal(value)
                    if not num.is_finite():
                        # B3-014: Infinity parses and passes every >= 0 check.
                        errors.append(f"{field} must be a finite number")
                    elif field in (
                        "purchase_price", "selling_price", "mrp", "wholesale_price",
                        "default_discount_percent", "reorder_level", "unit_cost",
                    ) and num < 0:
                        errors.append(f"{field} must be >= 0")
                    elif field == "conversion_rate" and num <= 0:
                        errors.append("conversion_rate must be > 0")
                    elif field in ("opening_stock", "quantity") and num < 0:
                        # 0 means catalog-only (no opening lot), same as a blank cell.
                        errors.append(f"{field} must be >= 0")
                except InvalidOperation:
                    errors.append(f"{field} must be a number")
        gst_rate_raw = (row.get("gst_rate") or "").strip()
        if gst_rate_raw:
            try:
                validate_gst_rate(gst_rate_raw)
            except Exception as exc:
                from django.core.exceptions import ValidationError as DjangoValidationError

                if isinstance(exc, DjangoValidationError):
                    errors.extend(list(exc.messages))
                else:
                    errors.append(str(exc))
        sku = (row.get("sku") or "").strip()
        barcode = (row.get("barcode") or "").strip()
        item_type = (row.get("product_type") or row.get("item_type") or "").strip().casefold()
        opening_raw = (row.get("opening_stock") or row.get("quantity") or "").strip()
        if item_type in {"service", "services"} and opening_raw:
            try:
                if Decimal(opening_raw) > 0:
                    errors.append("service items cannot have opening stock")
            except InvalidOperation:
                pass
        track_batch, track_serial = _tracking_from_row(row)
        if track_batch and track_serial:
            errors.append("batch and serial tracking cannot both be on")
        serials = _parse_serial_numbers(row.get("serial_no"))
        extra_serial_count = 0
        if sku:
            extra_serial_count = int((extra_serial_counts or {}).get(sku.casefold(), 0))
        opening_qty = None
        if opening_raw:
            try:
                opening_qty = Decimal(opening_raw)
            except InvalidOperation:
                opening_qty = None
        if extra_serial_count == 0:
            if track_serial and opening_qty is not None and opening_qty > 0:
                if not serials:
                    errors.append("Serial numbers are required for serial-tracked opening stock.")
                elif Decimal(len(serials)) != opening_qty:
                    errors.append("Opening quantity must equal the number of serials.")
            elif serials and not track_serial:
                errors.append("Serial numbers are only allowed when serial tracking is on.")
        if serials and len(serials) != len({value.casefold() for value in serials}):
            errors.append("Duplicate serial numbers supplied")
        batch_no = (row.get("batch_no") or "").strip()
        expiry = (row.get("expiry_date") or "").strip()
        if (track_batch or batch_no or expiry) and opening_raw:
            try:
                if Decimal(opening_raw) > 0 and (not batch_no or not expiry):
                    errors.append("Batch tracking is on but this row has no Batch no or Expiry.")
            except InvalidOperation:
                pass
        godown = (row.get("godown") or "").strip()
        if item_type in {"service", "services"} and (godown or batch_no or expiry or track_batch or track_serial):
            errors.append("service items cannot have godown, batch, or serial tracking")
        elif godown:
            from inventory.item_stock import match_warehouse

            try:
                match_warehouse(company, godown)
            except BusinessRuleError as exc:
                errors.append(str(getattr(exc, "detail", None) or exc))
        if sku:
            key = sku.casefold()
            if seen_skus is not None and key in seen_skus:
                errors.append(f"duplicate sku '{sku}' in file")
            elif (existing_skus is not None and key in existing_skus) or (
                existing_skus is None and Product.objects.filter(company=company, sku__iexact=sku).exists()
            ):
                if seen_skus is not None:
                    seen_skus.add(key)
                if opening_raw:
                    try:
                        if Decimal(opening_raw) > 0:
                            product = products_by_sku.get(key) if products_by_sku is not None else Product.objects.filter(
                                company=company, sku__iexact=sku
                            ).first()
                            already = False
                            if product is not None:
                                if skus_with_opening is not None:
                                    already = product.pk in skus_with_opening
                                else:
                                    already = StockMovement.objects.filter(
                                        company=company,
                                        product=product,
                                        movement_type=MovementType.OPENING_STOCK,
                                    ).exclude(reference_type="import_voided").exists()
                            if already:
                                errors.append(f"opening stock already recorded for '{product.name}'")
                    except InvalidOperation:
                        pass
            elif seen_skus is not None:
                seen_skus.add(key)
        if barcode:
            key = barcode.casefold()
            sku_key = sku.casefold() if sku else ""
            owner = products_by_sku.get(sku_key) if products_by_sku is not None and sku_key else None
            own_barcode = owner is not None and (owner.barcode or "").casefold() == key
            if seen_barcodes is not None and key in seen_barcodes:
                errors.append(f"duplicate barcode '{barcode}' in file")
            elif own_barcode:
                if seen_barcodes is not None:
                    seen_barcodes.add(key)
            elif existing_barcodes is not None and key in existing_barcodes:
                errors.append(f"barcode '{barcode}' already exists")
            elif existing_barcodes is None and Product.objects.filter(company=company, barcode__iexact=barcode).exclude(
                sku__iexact=sku or ""
            ).exists():
                errors.append(f"barcode '{barcode}' already exists")
            elif seen_barcodes is not None:
                seen_barcodes.add(key)
    if kind == ImportJob.Kind.OPENING_STOCK:
        sku = (row.get("sku") or "").strip()
        if not sku:
            errors.append("sku is required")
        else:
            key = sku.casefold()
            if seen_skus is not None and key in seen_skus:
                errors.append(f"duplicate sku '{sku}' in file")
            elif seen_skus is not None:
                seen_skus.add(key)
            if products_by_sku is not None:
                product = products_by_sku.get(key)
            else:
                product = Product.objects.filter(company=company, sku__iexact=sku).first()
            if not product:
                errors.append(f"no product with sku '{sku}'")
            else:
                if skus_with_opening is not None:
                    already_has_opening = product.pk in skus_with_opening
                else:
                    already_has_opening = StockMovement.objects.filter(
                        company=company,
                        product=product,
                        movement_type=MovementType.OPENING_STOCK,
                    ).exclude(reference_type="import_voided").exists()
                if already_has_opening:
                    errors.append(f"opening stock already recorded for '{product.name}'")
                serials = _parse_serial_numbers(row.get("serial_no"))
                if product.track_serial:
                    if not serials:
                        errors.append("Serial numbers are required for serial-tracked opening stock.")
                    elif len(serials) != len({value.casefold() for value in serials}):
                        errors.append("Duplicate serial numbers supplied")
                elif serials:
                    errors.append("Serial numbers are only allowed when serial tracking is on.")
        qty_raw = (row.get("quantity") or "").strip()
        if not qty_raw:
            errors.append("quantity is required")
        else:
            try:
                qty = Decimal(qty_raw)
                if not qty.is_finite():
                    errors.append("quantity must be a finite number")
                elif qty <= 0:
                    errors.append("quantity must be > 0")
                else:
                    serials = _parse_serial_numbers(row.get("serial_no"))
                    if serials and Decimal(len(serials)) != qty:
                        errors.append("Opening quantity must equal the number of serials.")
            except InvalidOperation:
                errors.append("quantity must be a number")
        unit_cost_raw = (row.get("unit_cost") or "").strip()
        if unit_cost_raw:
            try:
                if Decimal(unit_cost_raw) < 0:
                    errors.append("unit_cost must be >= 0")
            except InvalidOperation:
                errors.append("unit_cost must be a number")
    return errors


def _as_decimal(value, default="0"):
    try:
        result = Decimal(str(value if value not in (None, "") else default).strip())
    except (InvalidOperation, AttributeError):
        raise BusinessRuleError(f"Invalid number: {value!r}")
    # B3-014: Decimal("Infinity") / Decimal("NaN") construct fine and slip past
    # the >= 0 guards; reject them before they reach bulk_create / post_opening.
    if not result.is_finite():
        raise BusinessRuleError(f"Invalid number: {value!r}")
    return result


def _parse_extraction_confidence(value) -> float:
    try:
        confidence = float(str(value if value not in (None, "") else "1").strip())
    except (TypeError, ValueError):
        return 1.0
    return max(0.0, min(1.0, confidence))


def _normalize_gst_rate(
    value, *, warnings: list | None = None, row: int | None = None, required: bool = False,
) -> Decimal:
    """Snap OCR GST rates to an allowed slab; optionally record a warning (PUR-09)."""
    if required and value in (None, ""):
        prefix = f"Row {row}: " if row is not None else ""
        raise BusinessRuleError(f"{prefix}GST rate is required and cannot be invented.")
    # B3-001: supplier exports very commonly write the GST cell as "18%",
    # "18 %", "18.00%". Strip a trailing percent sign / whitespace before
    # parsing so a formatting quirk doesn't abort the whole upload.
    if isinstance(value, str):
        value = value.strip().rstrip("%").strip()
    rate = _as_decimal(value, "18")
    if rate not in ALLOWED_GST:
        # Snap common OCR noise to nearest allowed rate, else 18.
        nearest = min(ALLOWED_GST, key=lambda allowed: abs(allowed - rate))
        if abs(nearest - rate) <= Decimal("0.5"):
            snapped = nearest
            reason = f"GST rate {rate} snapped to nearest allowed rate {snapped}"
        else:
            snapped = Decimal("18")
            reason = f"GST rate {rate} defaulted to 18 (not near an allowed slab)"
        if warnings is not None:
            prefix = f"Row {row}: " if row is not None else ""
            warnings.append(f"{prefix}{reason}")
        return snapped
    return rate


def _preview_bill_line(raw_line: dict, *, index: int, rate_warnings: list[str]) -> dict:
    """Build a preview line without inventing qty=1 or GST=18 for unread OCR fields."""
    line_warnings: list[str] = []
    qty = _line_get(raw_line, "quantity")
    gst_raw = _line_get(raw_line, "gst_rate", "gstRate")
    gst_unparseable = False
    if gst_raw:
        try:
            gst_rate = str(_normalize_gst_rate(gst_raw, warnings=line_warnings, row=index))
        except (BusinessRuleError, InvalidOperation):
            # B3-001: one bad GST cell must not abort the whole job — exclude
            # just this line, like every other unreadable field.
            gst_rate = ""
            gst_unparseable = True
            line_warnings.append(
                f"Row {index}: GST rate {gst_raw!r} is not a number — line excluded."
            )
    else:
        gst_rate = ""
    # Prefer explicit include from LLM normalize; otherwise require readable qty+GST.
    if "include" in raw_line and not gst_unparseable:
        include = bool(raw_line.get("include"))
    else:
        include = bool(qty) and bool(gst_rate)
    extras = collect_extras(raw_line)
    line = {
        "name": _line_get(raw_line, "name"),
        "sku": _line_get(raw_line, "sku"),
        "hsn_code": _line_get(raw_line, "hsn_code", "hsnCode", "hsn"),
        "quantity": qty,
        "unit_price": _line_get(raw_line, "unit_price", "unitPrice", "rate", default="0") or "0",
        "gst_rate": gst_rate,
        "mrp": _line_get(raw_line, "mrp", default="0") or "0",
        "include": include,
        "cs": extras.get("cs") or _line_get(raw_line, "cs"),
        "upc": extras.get("upc") or _line_get(raw_line, "upc"),
        "printed_gross_amt": _line_get(raw_line, "printed_gross_amt"),
        "printed_taxable_amt": _line_get(raw_line, "printed_taxable_amt"),
        "extras": extras,
        "flags": list(raw_line.get("flags") or []) if isinstance(raw_line.get("flags"), list) else [],
        "si": _line_get(raw_line, "si", "sl", "s_no", "sno") or str(index),
        "pcs": _line_get(raw_line, "pcs") or qty,
    }
    if line_warnings:
        line["warnings"] = line_warnings
        rate_warnings.extend(line_warnings)
    return line


def _bill_line_errors(line: dict) -> list[str]:
    errors = []
    if not str(line.get("name") or "").strip():
        errors.append("name is required")
    try:
        if _as_decimal(line.get("quantity"), "0") <= 0:
            errors.append("quantity must be > 0")
    except BusinessRuleError:
        errors.append("quantity must be a number")
    try:
        _as_decimal(line.get("unit_price"), "0")
    except BusinessRuleError:
        errors.append("unit_price must be a number")
    hsn = str(line.get("hsn_code") or "").strip()
    if hsn and not HSN_RE.match(hsn):
        errors.append("invalid HSN code")
    return errors


def _safe_decimal(value) -> Decimal:
    try:
        text = str(value if value not in (None, "") else "0").strip()
        return Decimal(text) if text else Decimal("0")
    except InvalidOperation:
        return Decimal("0")


def _detect_clarifications(payload: dict) -> list[dict]:
    """Document-level qty-formula ambiguity (Redesign Plan §4.2/§4.3).

    Options are derived from whatever count/pack columns this bill actually
    has — not a hardcoded Cs/Pcs/UPC questionnaire.
    """
    return detect_qty_clarifications(payload.get("lines") or [])


def _infer_qty_answers(lines: list[dict], *, tolerance: Decimal) -> dict | None:
    expr = infer_qty_formula(lines, tolerance=tolerance)
    if not expr:
        return None
    return {"qty_formula": expr}


def _apply_cross_check(
    lines: list[dict],
    answers: dict,
    *,
    tolerance: Decimal,
    reconcile_print: bool = True,
) -> str:
    return apply_qty_formula(
        lines, answers, tolerance=tolerance, reconcile_print=reconcile_print
    )


def _refresh_line_validity(lines: list[dict]) -> list[dict]:
    """Re-evaluate include/errors after quantity recombination.

    Case-only DMS rows (Cs>0, Pcs=0) fail `quantity must be > 0` on the first
    pass; once qty is rewritten as (Cs×UPC)+Pcs they must be brought back in.
    """
    errors = []
    for index, line in enumerate(lines, start=1):
        line_errors = _bill_line_errors(line)
        if line_errors:
            line["include"] = False
            errors.append({"row": index, "errors": line_errors, "data": line})
            continue
        gst_ok = bool(str(line.get("gst_rate") or "").strip())
        try:
            qty_ok = _as_decimal(line.get("quantity"), "0") > 0
        except BusinessRuleError:
            qty_ok = False
        if qty_ok and gst_ok:
            line["include"] = True
    return errors


def _match_bill_template(company, gstin: str, column_headers: list[str]) -> SupplierBillTemplate | None:
    """GSTIN is the primary key; column signature (order-independent) guards
    against applying a stale template after the vendor changes their layout."""
    if not gstin:
        return None
    template = SupplierBillTemplate.objects.filter(company=company, gstin__iexact=gstin).first()
    if template is None:
        return None
    stored = {str(h).strip().upper() for h in (template.column_signature or [])}
    seen = {str(h).strip().upper() for h in (column_headers or [])}
    if not stored or not seen:
        return template
    overlap = len(stored & seen) / max(len(stored), 1)
    if overlap < BILL_TEMPLATE_SIGNATURE_MATCH_THRESHOLD:
        return None
    return template


def _template_answers(template: SupplierBillTemplate) -> dict:
    mapping = template.column_mapping if isinstance(template.column_mapping, dict) else {}
    if mapping.get("qty_formula"):
        return {"qty_formula": mapping["qty_formula"]}
    if template.line_total_formula == SupplierBillTemplate.LineTotalFormula.CASE_UNITS_PLUS_LOOSE:
        return {"qty_formula": "cs*upc+quantity"}
    return {"qty_formula": "quantity"}


def _lines_have_pack_columns(lines: list[dict]) -> bool:
    for line in lines:
        keys = set(count_pool(line))
        keys.discard("quantity")
        if keys:
            return True
    return False


def _skip_simple_qty_template(template: SupplierBillTemplate | None, lines: list[dict]) -> bool:
    """A saved SIMPLE layout is from a prior extract that never saw Cs/UPC.

    If this bill actually has pack-size columns, ignore SIMPLE so inference
    can recombine billed qty instead of treating Pcs as the full quantity.
    """
    if template is None:
        return False
    if _template_answers(template).get("qty_formula") != "quantity":
        return False
    return _lines_have_pack_columns(lines)


def _resolve_import_company_gstin(company, preview: dict, *, kind: str):
    """Stamp filing identity from buyer GSTIN (sales) or primary CompanyGstin."""
    wanted = ""
    if kind == ImportJob.Kind.SALES_BILL:
        wanted = str(preview.get("buyer_gstin") or "").strip().upper()
    else:
        wanted = str(preview.get("company_gstin") or preview.get("buyer_gstin") or "").strip().upper()
        known = {
            (g or "").upper()
            for g in CompanyGstin.objects.filter(company=company, is_active=True).values_list("gstin", flat=True)
        }
        if getattr(company, "gstin", ""):
            known.add(company.gstin.upper())
        if wanted and wanted not in known:
            wanted = ""
    qs = CompanyGstin.objects.filter(company=company, is_active=True)
    if wanted:
        hit = qs.filter(gstin__iexact=wanted).first()
        if hit:
            return hit
    return qs.filter(is_primary=True).order_by("-id").first() or qs.order_by("id").first()


def _infer_direction_warning(kind: str, payload: dict, company) -> str:
    """Bill Import Redesign Plan §4.5: GSTIN-based sanity check, surfaced as a
    warning (not a block) since the user already picked purchase vs. sales by
    which upload flow they used."""
    company_gstins = {
        g.upper() for g in CompanyGstin.objects.filter(company=company, is_active=True)
        .values_list("gstin", flat=True)
    }
    if getattr(company, "gstin", ""):
        company_gstins.add(company.gstin.upper())
    if not company_gstins:
        return ""
    seller = (payload.get("supplier_gstin") or "").upper()
    buyer = (payload.get("buyer_gstin") or "").upper()
    if kind == ImportJob.Kind.PURCHASE_BILL and seller and seller in company_gstins:
        return (
            "This bill's seller GSTIN matches your company — it looks like a sales "
            "invoice you issued, not a purchase bill. Upload it from Sales instead?"
        )
    if kind == ImportJob.Kind.SALES_BILL and buyer and buyer in company_gstins:
        return (
            "This bill's buyer GSTIN matches your company — it looks like a purchase "
            "bill you received, not a sales invoice. Upload it from Purchases instead?"
        )
    return ""


BILL_COLUMN_ALIASES = {
    "si": ["si", "sl", "s.no", "s no", "sno", "sr no"],
    "name": ["name", "item", "item description", "description", "product", "product name"],
    "sku": ["sku", "pcode", "item code", "product code", "code"],
    "hsn_code": ["hsn", "hsn_code", "hsn code", "hsn/sac"],
    "quantity": ["quantity", "qty", "pcs", "pieces"],
    "unit_price": ["unit_price", "rate", "price", "pc price", "pcs price", "unit rate"],
    "gst_rate": ["gst_rate", "gst", "gst%", "tax_rate", "tax%", "gst %"],
    "mrp": ["mrp"],
    "cs": ["cs", "cases", "ctn", "cartons"],
    "upc": ["upc", "units per case", "pcs/cs"],
    "printed_gross_amt": ["printed_gross_amt", "gross amt", "gross amount", "gross", "gross value"],
}

_INVOICE_META_ALIASES = {
    "supplier_name": ("registered name", "supplier name", "seller name", "vendor name"),
    "supplier_gstin": ("gstin", "supplier gstin", "vendor gstin"),
    "buyer_name": ("customer name", "buyer name", "bill to"),
    "buyer_gstin": ("customer gstin", "buyer gstin"),
    "bill_number": ("invoice no", "invoice number", "bill no", "bill number"),
    "bill_date": ("invoice date", "bill date"),
}


def _xlsx_kv_meta(rows: list) -> dict:
    meta: dict[str, str] = {}
    for row in rows:
        if not row or row[0] in (None, ""):
            continue
        label = str(row[0]).strip().lower().rstrip(".:")
        value = row[1] if len(row) > 1 and row[1] not in (None, "") else ""
        if not value:
            continue
        text = str(value).strip()
        for field, aliases in _INVOICE_META_ALIASES.items():
            if label in aliases and field not in meta:
                meta[field] = text
                break
    return meta


def _xlsx_best_bill_rows(workbook) -> tuple[list[dict], dict]:
    """Read every sheet once (read-only iterators are single-pass) and pick
    the one whose header looks like a bill line-item table.

    ChatGPT/DMS exports often put invoice meta on the first sheet and the
    actual Cs/Pcs/UPC rows on a later 'Line Items' sheet — `active` would
    miss them.
    """
    candidates: list[tuple[int, int, list[dict]]] = []
    meta: dict[str, str] = {}
    for sheet in workbook.worksheets:
        # B3-004: cap what we materialise into worker memory, matching the
        # master-import path's guards. A crafted / huge bill sheet otherwise
        # loads every row and JSON-serialises it onto the job.
        materialized = []
        cells = 0
        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if idx > MAX_IMPORT_ROWS:
                raise BusinessRuleError(
                    f"Bill import exceeds {MAX_IMPORT_ROWS} rows. Split the file and retry."
                )
            row = tuple(row)
            cells += len(row)
            if cells > MAX_IMPORT_CELLS:
                raise BusinessRuleError(
                    f"Bill import exceeds {MAX_IMPORT_CELLS} cells. Split the file and retry."
                )
            materialized.append(row)
        if not materialized:
            continue
        header = [str(h or "").strip().lower() for h in materialized[0]]
        if not any(header):
            continue
        header_set = set(header)
        score = sum(
            1
            for token in (
                "item", "item description", "description", "particulars", "product",
                "qty", "quantity", "pcs", "hsn", "rate", "amount", "taxable",
                "sku", "code", "pcode", "mrp", "cs", "upc", "gross amt",
                "pc price", "pcs price",
            )
            if token in header_set or any(token in h for h in header_set)
        )
        data = []
        for raw_row in materialized[1:]:
            if raw_row is None or all(v in (None, "") for v in raw_row):
                continue
            data.append({header[i]: raw_row[i] for i in range(min(len(header), len(raw_row)))})
        candidates.append((score, len(data), data))
        if score < 2:
            for key, value in _xlsx_kv_meta(materialized).items():
                meta.setdefault(key, value)
    if not candidates:
        return [], meta
    scored = [c for c in candidates if c[0] >= 2]
    pool = scored or candidates
    pool.sort(key=lambda item: (item[0], item[1]), reverse=True)
    return pool[0][2], meta


def _read_structured_rows(raw: bytes, filename: str) -> list[dict]:
    rows, _meta = _read_structured_bill(raw, filename)
    return rows


def _normalize_sheet_name(name: str) -> str:
    return (name or "").strip().lower().replace(" ", "_")


def _cell_to_import_text(value) -> str:
    """Coerce Excel cell values to JSON-safe text (dates, ints, floats)."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.tzinfo is None and value.time() == time.min:
            return value.date().isoformat()
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "yes" if value else "no"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value != value or value in (float("inf"), float("-inf")):
            return ""
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def _sku_keys_from_rows(rows) -> set[str]:
    keys = set()
    for raw in rows or []:
        row = {str(k).strip().lower(): v for k, v in (raw or {}).items()}
        sku = _cell_to_import_text(row.get("sku") or row.get("item code") or "")
        if sku:
            keys.add(sku.casefold())
    return keys


def _prune_unmatched_extra_sheets(extra_sheets, item_rows) -> list[dict]:
    """Drop extra sheets whose SKUs match none of the item rows (mis-filled templates)."""
    notes = []
    item_skus = _sku_keys_from_rows(item_rows)
    for key in ("opening_lots", "opening_serials"):
        sheet = extra_sheets.get(key) or []
        if not sheet:
            continue
        sheet_skus = _sku_keys_from_rows(sheet)
        if sheet_skus and item_skus.isdisjoint(sheet_skus):
            extra_sheets[key] = []
            notes.append({"source": key, "target": "ignored (no matching SKUs)"})
    return notes


def _rows_from_worksheet(sheet) -> list[dict]:
    materialized = []
    cells = 0
    for idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if idx > MAX_IMPORT_ROWS:
            raise BusinessRuleError(
                f"Import exceeds {MAX_IMPORT_ROWS} rows. Split the file and retry."
            )
        cells += len(row or ())
        if cells > MAX_IMPORT_CELLS:
            raise BusinessRuleError("Import spreadsheet is too large (cell cap).")
        materialized.append(tuple(row))
    if not materialized:
        return []
    header = [str(h or "").strip().lower() for h in materialized[0]]
    if not any(header):
        return []
    rows = []
    for raw_row in materialized[1:]:
        if raw_row is None or all(v in (None, "") for v in raw_row):
            continue
        rows.append(
            {
                header[i]: _cell_to_import_text(raw_row[i])
                for i in range(min(len(header), len(raw_row)))
            }
        )
    return rows


def _read_named_sheets(raw: bytes) -> dict[str, list[dict]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheets = {}
    for sheet in workbook.worksheets:
        sheets[_normalize_sheet_name(sheet.title)] = _rows_from_worksheet(sheet)
    return sheets


def _read_structured_bill(raw: bytes, filename: str) -> tuple[list[dict], dict]:
    """Deterministic CSV/XLSX row reader for the bill-import 'I have an
    export' path (§4.1) — no LLM involved, exact by construction."""
    name = (filename or "").lower()
    head = raw[:8]
    # B3-006: decide by content, not just the filename. A .xls (OLE2) file — or
    # any upload whose browser content-type is application/vnd.ms-excel — that
    # doesn't end in .xlsx falls through to the CSV branch, where cp1252
    # "decodes" the binary into garbage rows. Sniff the magic bytes first.
    is_zip = head.startswith(b"PK\x03\x04")  # xlsx/xlsm is a zip container
    is_ole2 = head.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1")  # legacy .xls
    if is_ole2:
        raise BusinessRuleError(
            "Legacy .xls files aren't supported — re-save as .xlsx or export CSV."
        )
    if is_zip or name.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        return _xlsx_best_bill_rows(workbook)

    try:
        text = _decode_csv_text(raw)
    except BusinessRuleError:
        raise
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    cells = 0
    for idx, row in enumerate(reader):
        if idx > MAX_IMPORT_ROWS:
            raise BusinessRuleError(
                f"Bill import exceeds {MAX_IMPORT_ROWS} rows. Split the file and retry."
            )
        cells += len(row)
        if cells > MAX_IMPORT_CELLS:
            raise BusinessRuleError(
                f"Bill import exceeds {MAX_IMPORT_CELLS} cells. Split the file and retry."
            )
        rows.append({(k or "").strip().lower(): v for k, v in row.items()})
    return rows, {}


def _map_structured_row(row: dict) -> dict:
    mapped = {}
    claimed = set()
    for field, aliases in BILL_COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in row and row[alias] not in (None, ""):
                mapped[field] = row[alias]
                claimed.add(alias)
                break
    leftovers = {k: v for k, v in row.items() if k not in claimed}
    extras = collect_extras({**leftovers, "raw_columns": leftovers})
    if extras:
        mapped["extras"] = extras
        if "cs" not in mapped and extras.get("cs") is not None:
            mapped["cs"] = extras["cs"]
        if "upc" not in mapped and extras.get("upc") is not None:
            mapped["upc"] = extras["upc"]
    return mapped


def _inclusive_flag(value) -> bool:
    return str(value or "").strip().casefold() in {"inclusive", "yes", "true", "1", "y"}


def _validate_extra_sheets(job, extra_sheets, preview) -> list[dict]:
    errors = []
    if not extra_sheets:
        return errors
    from inventory.item_stock import match_warehouse

    preview_by_sku = {
        (row.get("sku") or "").strip().casefold(): row for row in preview if row.get("sku")
    }
    preview_skus = set(preview_by_sku)
    lots = extra_sheets.get("opening_lots") or []
    serials = extra_sheets.get("opening_serials") or []
    serial_skus = set(_extra_serial_counts(extra_sheets))
    serial_skus.update(
        (row.get("sku") or "").strip().casefold()
        for row in preview
        if _tracking_from_row(row)[1]
    )
    batched_skus = set()
    for raw in lots:
        row = {str(k).strip().lower(): v for k, v in (raw or {}).items()}
        sku = str(row.get("sku") or row.get("item code") or "").strip().casefold()
        if str(row.get("batch_no") or row.get("batch no") or row.get("batch") or "").strip():
            batched_skus.add(sku)
    for index, raw in enumerate(lots, start=2):
        row = _map_master_row(raw) if not isinstance(raw, dict) else {str(k).strip().lower(): v for k, v in raw.items()}
        sku = str(row.get("sku") or row.get("item code") or "").strip()
        qty = str(row.get("quantity") or row.get("qty") or "").strip()
        batch_no = str(row.get("batch_no") or row.get("batch no") or row.get("batch") or "").strip()
        expiry = str(row.get("expiry_date") or row.get("expiry date") or row.get("expiry") or "").strip()
        godown = str(row.get("godown") or row.get("warehouse") or "").strip()
        row_errors = []
        preview_row = preview_by_sku.get(sku.casefold()) if sku else None
        item_type = str((preview_row or {}).get("product_type") or "").strip().casefold()
        if item_type in {"service", "services"}:
            row_errors.append(f"Service items cannot have opening stock (row {index}).")
        if not sku:
            row_errors.append("sku is required")
        elif sku.casefold() not in preview_skus:
            row_errors.append(f"unknown sku '{sku}' on opening_lots")
        if not qty:
            row_errors.append("quantity is required")
        if sku.casefold() in batched_skus and (not batch_no or not expiry):
            row_errors.append("Batch tracking is on but this row has no Batch no or Expiry.")
        if sku.casefold() in serial_skus:
            row_errors.append("Serial-tracked items cannot use opening_lots. List serial numbers instead.")
        if godown:
            try:
                match_warehouse(job.company, godown)
            except BusinessRuleError as exc:
                row_errors.append(str(exc.detail if hasattr(exc, "detail") else exc))
        if row_errors:
            errors.append({"row": index, "sheet": "opening_lots", "errors": row_errors, "data": row})
    for index, raw in enumerate(serials, start=2):
        row = {str(k).strip().lower(): v for k, v in (raw or {}).items()}
        sku = str(row.get("sku") or row.get("item code") or "").strip()
        serial = str(row.get("serial_no") or row.get("serial") or "").strip()
        row_errors = []
        if not sku:
            row_errors.append("sku is required")
        elif sku.casefold() not in preview_skus:
            row_errors.append(f"unknown sku '{sku}' on opening_serials")
        if not serial:
            row_errors.append("serial no is required")
        if row_errors:
            errors.append({"row": index, "sheet": "opening_serials", "errors": row_errors, "data": row})
    return errors


class ImportService:
    @staticmethod
    def validate(job: ImportJob):
        """Parse + validate the uploaded CSV/XLSX; store preview and error report."""
        if job.kind == ImportJob.Kind.PURCHASE_BILL:
            raise BusinessRuleError("Purchase bill imports use LLM extraction, not CSV validate.")
        if job.kind in ImportJob.BILL_KINDS:
            raise BusinessRuleError("Bill imports use a separate extraction pipeline.")
        with job.file.file.open("rb") as handle:
            raw = handle.read()
        job.file_sha256 = hashlib.sha256(raw).hexdigest()
        if (
            job.kind == ImportJob.Kind.PRODUCTS
            and job.file_sha256
            and ImportJob.objects.filter(
                company=job.company,
                kind=ImportJob.Kind.PRODUCTS,
                file_sha256=job.file_sha256,
                status=ImportJob.Status.COMMITTED,
            ).exclude(pk=job.pk).exists()
        ):
            raise BusinessRuleError("This file was already committed. Re-upload will not post opening stock again.")
        filename = job.file.original_name or job.file.file.name or ""
        name_lower = filename.lower()
        extra_sheets = {}
        extra_sheet_notes = []
        if name_lower.endswith((".xlsx", ".xlsm")):
            if job.kind == ImportJob.Kind.PRODUCTS:
                named = _read_named_sheets(raw)
                preferred_item_sheets = ("items", "products", "bulk_upload", "sheet1")
                items_key = next((k for k in preferred_item_sheets if k in named and named[k]), None)
                if items_key:
                    rows = named[items_key]
                elif any(k in named for k in ("items", "products")):
                    raise BusinessRuleError(
                        "No item rows found. Add at least one product on the items sheet."
                    )
                else:
                    skip = {"notes", "opening_lots", "opening_serials", "openinglots", "openingserials"}
                    rows = next((v for k, v in named.items() if v and k not in skip), [])
                extra_sheets = {
                    "opening_lots": named.get("opening_lots") or named.get("openinglots") or [],
                    "opening_serials": named.get("opening_serials") or named.get("openingserials") or [],
                }
                extra_sheet_notes = _prune_unmatched_extra_sheets(extra_sheets, rows)
            else:
                rows = _read_structured_rows(raw, filename)
            if not rows:
                raise BusinessRuleError("CSV file has no header row.")
            fieldnames = list(rows[0].keys()) if rows else []
            if not fieldnames:
                raise BusinessRuleError("CSV file has no header row.")
        else:
            text = _decode_csv_text(raw)
            reader = csv.DictReader(io.StringIO(text))
            if not reader.fieldnames:
                raise BusinessRuleError("CSV file has no header row.")
            fieldnames = [f.strip().lower() for f in reader.fieldnames]
            rows = [{(k or "").strip().lower(): v for k, v in raw_row.items()} for raw_row in reader]

        # Required columns after alias expansion: accept either canonical or any alias.
        required = REQUIRED_COLUMNS[job.kind]
        available = set(fieldnames)
        for aliases in MASTER_COLUMN_ALIASES.values():
            available.update(aliases)
        # Also treat present aliased headers as satisfying canonical requirements.
        present_canonical = set(fieldnames)
        for field, aliases in MASTER_COLUMN_ALIASES.items():
            if any(a in fieldnames for a in aliases):
                present_canonical.add(field)
        missing = [c for c in required if c not in present_canonical]
        if missing:
            raise BusinessRuleError(f"Missing required columns: {', '.join(missing)}.")

        # Bulk prefetch for duplicate / existence checks (avoid per-row queries).
        existing_skus = {
            (s or "").casefold()
            for s in Product.objects.filter(company=job.company).exclude(sku="")
            .values_list("sku", flat=True)
        }
        existing_barcodes = {
            (b or "").casefold()
            for b in Product.objects.filter(company=job.company).exclude(barcode="")
            .values_list("barcode", flat=True)
        }
        products_by_sku = {
            (p.sku or "").casefold(): p
            for p in Product.objects.filter(company=job.company).exclude(sku="")
        }
        skus_with_opening = set(
            StockMovement.objects.filter(
                company=job.company,
                movement_type=MovementType.OPENING_STOCK,
            )
            .exclude(reference_type="import_voided")
            .values_list("product_id", flat=True)
        )

        extra_serial_counts = _extra_serial_counts(extra_sheets)

        header_map = {}
        cf_defs = []
        if job.kind == ImportJob.Kind.PRODUCTS:
            from masters.custom_fields import active_defs, resolve_import_columns

            cf_defs = active_defs(job.company)
            header_map, cf_header_errors = resolve_import_columns(list(fieldnames), cf_defs)
            if cf_header_errors:
                raise BusinessRuleError(cf_header_errors[0])
            existing_barcodes -= _released_barcodes(rows, products_by_sku)

        preview, errors = [], []
        seen_skus, seen_barcodes = set(), set()
        for index, raw_row in enumerate(rows, start=2):  # header is row 1
            row = _map_master_row(raw_row)
            row_errors = _validate_row(
                job.kind, row, job.company,
                seen_skus=seen_skus, seen_barcodes=seen_barcodes,
                existing_skus=existing_skus, existing_barcodes=existing_barcodes,
                products_by_sku=products_by_sku, skus_with_opening=skus_with_opening,
                extra_serial_counts=extra_serial_counts,
            )
            if job.kind == ImportJob.Kind.PRODUCTS and header_map:
                row_errors.extend(_custom_field_row_errors(row, header_map, cf_defs))
            if row_errors:
                errors.append({"row": index, "errors": row_errors, "data": row})
            else:
                preview.append(row)

        extra_sheets = extra_sheets or {}
        inline_lots = _lots_from_item_rows(preview)
        if inline_lots:
            extra_sheets["opening_lots"] = list(extra_sheets.get("opening_lots") or []) + inline_lots
        inline_serials = _serials_from_item_rows(preview)
        if inline_serials:
            extra_sheets["opening_serials"] = list(extra_sheets.get("opening_serials") or []) + inline_serials
        extra_errors = _validate_extra_sheets(job, extra_sheets, preview)
        errors.extend(extra_errors)
        job.total_rows = len(preview) + len(errors)
        job.valid_rows = len(preview)
        job.error_rows = len(errors)
        job.preview = preview
        job.errors = errors
        job.column_mappings = resolve_master_column_mappings(fieldnames) + extra_sheet_notes
        job.extra_sheets = extra_sheets
        job.custom_field_defs_snapshot = cf_defs
        job.custom_field_header_map = header_map
        job.status = ImportJob.Status.PREVIEWED
        job.save()
        return job

    @staticmethod
    @transaction.atomic
    def commit(job: ImportJob, user):
        """
        All-or-nothing write of previewed rows (duplicates caught at validate).
        Locks the ImportJob row and re-checks PREVIEWED before writing.
        """
        if job.kind in ImportJob.BILL_KINDS:
            return BillImportService.commit(job, user)

        job = ImportJob.objects.select_for_update().get(pk=job.pk)
        if job.status != ImportJob.Status.PREVIEWED:
            raise BusinessRuleError("Import must be previewed before commit.")
        # R4-015 (reverted): CUSTOMERS / SUPPLIERS / OPENING_STOCK imports commit
        # PARTIALLY by design (write valid rows, report the rest — see
        # test_commit_writes_only_valid_rows / test_opening_stock_import). Only
        # PRODUCTS is all-or-nothing. Invalid rows are still skipped, never
        # written, so bad data does not land regardless.
        if job.kind == ImportJob.Kind.PRODUCTS and job.error_rows:
            raise BusinessRuleError(
                "Fix all errors before commit. Any invalid row blocks the entire job.",
                code=HelpCode.IMPORT_INVALID_ROWS,
            )
        preview = job.preview if isinstance(job.preview, list) else []
        if job.kind == ImportJob.Kind.CUSTOMERS:
            created = ImportService._commit_customers(job, preview, user)
        elif job.kind == ImportJob.Kind.SUPPLIERS:
            created = ImportService._commit_suppliers(job, preview, user)
        elif job.kind == ImportJob.Kind.PRODUCTS:
            created = ImportService._commit_products(job, preview, user)
        elif job.kind == ImportJob.Kind.OPENING_STOCK:
            created = ImportService._commit_opening_stock(job, preview, user)
        else:
            raise BusinessRuleError(f"Unsupported import kind '{job.kind}'.")

        job.preview = preview
        job.status = ImportJob.Status.COMMITTED
        job.committed_at = timezone.now()
        job.updated_by = user
        job.save()
        AuditService.log(
            company=job.company, user=user, action="IMPORT",
            entity_type="ImportJob", entity_id=job.pk,
            description=f"Imported {created} {job.kind.lower()} rows",
            metadata={"created": created, "errors": job.error_rows},
        )
        return created

    @staticmethod
    def _post_opening_items(job, items, user):
        if not items:
            return
        movements = InventoryService.post_opening_movements_batch(
            company=job.company,
            items=items,
            reference_type="import",
            reference_id=job.pk,
            user=user,
        )
        if job.company.accounting_enabled:
            from accounting.services import PostingService

            for movement in movements:
                PostingService.post_opening_stock(movement, user)

    @staticmethod
    def _commit_customers(job, preview, user):
        now = timezone.now()
        created = 0
        existing_gstin = {
            (c.gstin or "").strip().upper()
            for c in Customer.objects.filter(company=job.company).exclude(gstin="")
        }
        existing_phone = {
            (c.phone or "").strip()
            for c in Customer.objects.filter(company=job.company).exclude(phone="")
        }
        existing_name = {
            (c.name or "").strip().lower()
            for c in Customer.objects.filter(company=job.company)
        }
        rows = []
        for row in preview:
            gstin = (row.get("gstin") or "").strip().upper()
            phone = (row.get("phone") or "").strip()
            name = (row.get("name") or "").strip()
            if gstin and gstin in existing_gstin:
                continue
            if phone and phone in existing_phone:
                continue
            if name.lower() in existing_name:
                continue
            rows.append(
                Customer(
                    company=job.company,
                    name=row["name"],
                    phone=row.get("phone", "") or "",
                    email=row.get("email", "") or "",
                    gstin=row.get("gstin", "") or "",
                    state=row.get("state", "") or "",
                    billing_address=row.get("address", "") or "",
                    created_by=user,
                    updated_by=user,
                    created_at=now,
                    updated_at=now,
                )
            )
            if gstin:
                existing_gstin.add(gstin)
            if phone:
                existing_phone.add(phone)
            existing_name.add(name.lower())
        if rows:
            Customer.objects.bulk_create(rows, batch_size=BULK_BATCH)
            created = len(rows)
        return created

    @staticmethod
    def _commit_suppliers(job, preview, user):
        now = timezone.now()
        existing_gstin = {
            (s.gstin or "").strip().upper()
            for s in Supplier.objects.filter(company=job.company).exclude(gstin="")
        }
        existing_phone = {
            (s.phone or "").strip()
            for s in Supplier.objects.filter(company=job.company).exclude(phone="")
        }
        existing_name = {
            (s.name or "").strip().lower()
            for s in Supplier.objects.filter(company=job.company)
        }
        rows = []
        for row in preview:
            gstin = (row.get("gstin") or "").strip().upper()
            phone = (row.get("phone") or "").strip()
            name = (row.get("name") or "").strip()
            if gstin and gstin in existing_gstin:
                continue
            if phone and phone in existing_phone:
                continue
            if name.lower() in existing_name:
                continue
            rows.append(
                Supplier(
                    company=job.company,
                    name=row["name"],
                    phone=row.get("phone", "") or "",
                    email=row.get("email", "") or "",
                    gstin=row.get("gstin", "") or "",
                    state=row.get("state", "") or "",
                    address=row.get("address", "") or "",
                    created_by=user,
                    updated_by=user,
                    created_at=now,
                    updated_at=now,
                )
            )
            if gstin:
                existing_gstin.add(gstin)
            if phone:
                existing_phone.add(phone)
            existing_name.add(name.lower())
        if rows:
            Supplier.objects.bulk_create(rows, batch_size=BULK_BATCH)
        return len(rows)

    @staticmethod
    def _resolve_units(company, preview, user):
        wanted = {}
        for row in preview:
            for field in ("unit", "unit_name", "alternate_unit"):
                unit_str = (row.get(field) or "").strip()
                if unit_str:
                    wanted[unit_str.casefold()] = unit_str
        if not wanted:
            return {}
        existing = list(Unit.objects.filter(company=company))
        by_key = {}
        for unit in existing:
            by_key[unit.name.casefold()] = unit
            by_key[unit.short_name.casefold()] = unit
        now = timezone.now()
        missing = []
        missing_keys = []
        for key, orig in wanted.items():
            if key not in by_key:
                missing_keys.append(key)
                missing.append(
                    Unit(
                        company=company,
                        name=orig.upper(),
                        short_name=orig.upper()[:10],
                        uqc_code=orig.upper()[:8],
                        created_by=user,
                        updated_by=user,
                        created_at=now,
                        updated_at=now,
                    )
                )
        if missing:
            created = Unit.objects.bulk_create(missing, batch_size=BULK_BATCH)
            for unit in created:
                by_key[unit.name.casefold()] = unit
                by_key[unit.short_name.casefold()] = unit
        return by_key

    @staticmethod
    def _resolve_categories(company, preview, user):
        wanted = {}
        for row in preview:
            name = (row.get("category") or "").strip()
            if name:
                wanted[name.casefold()] = name
        if not wanted:
            return {}
        by_key = {
            (c.name or "").casefold(): c
            for c in Category.objects.filter(company=company)
        }
        now = timezone.now()
        missing = [
            Category(
                company=company, name=orig, created_by=user, updated_by=user,
                created_at=now, updated_at=now,
            )
            for key, orig in wanted.items()
            if key not in by_key
        ]
        if missing:
            created = Category.objects.bulk_create(missing, batch_size=BULK_BATCH)
            for category in created:
                by_key[category.name.casefold()] = category
        return by_key

    @staticmethod
    def _resolve_brands(company, preview):
        wanted = {
            (row.get("brand_code") or "").strip().casefold()
            for row in preview
            if (row.get("brand_code") or "").strip()
        }
        if not wanted:
            return {}
        return {
            (b.name or "").casefold(): b
            for b in Brand.objects.filter(company=company)
            if (b.name or "").casefold() in wanted
        }

    @staticmethod
    def _commit_products(job, preview, user):
        now = timezone.now()
        units = ImportService._resolve_units(job.company, preview, user)
        categories = ImportService._resolve_categories(job.company, preview, user)
        brands = ImportService._resolve_brands(job.company, preview)
        extra_skus = {
            str(r.get("sku") or r.get("item code") or "").strip().casefold()
            for r in (job.extra_sheets or {}).get("opening_lots") or []
        } | {
            str(r.get("sku") or r.get("item code") or "").strip().casefold()
            for r in (job.extra_sheets or {}).get("opening_serials") or []
        }
        existing_by_sku = {
            (p.sku or "").casefold(): p
            for p in Product.objects.filter(company=job.company).exclude(sku="")
        }
        creates = []
        create_opening = []
        updates = []
        update_opening = []
        barcode_rewrites: dict[int, str] = {}
        for row in preview:
            unit_str = (row.get("unit") or row.get("unit_name") or "").strip()
            unit_obj = units.get(unit_str.casefold()) if unit_str else None
            alt_str = (row.get("alternate_unit") or "").strip()
            alt_obj = units.get(alt_str.casefold()) if alt_str else None
            if alt_obj is not None and alt_obj == unit_obj:
                alt_obj = None
            category_str = (row.get("category") or "").strip()
            brand_str = (row.get("brand_code") or "").strip()
            purchase_price = Decimal(row.get("purchase_price") or "0")
            is_service = _is_service_type(row)
            track_batch, track_serial = _tracking_from_row(row)
            track_inventory = _optional_bool(row.get("track_inventory"), default=not is_service)
            if is_service:
                track_inventory = False
                track_batch = False
                track_serial = False
            sku_key = (row.get("sku") or "").strip().casefold()
            existing = existing_by_sku.get(sku_key) if sku_key else None
            opening_stock_raw = (row.get("opening_stock") or row.get("quantity") or "").strip()
            opening_qty = Decimal("0")
            if opening_stock_raw:
                try:
                    opening_qty = Decimal(opening_stock_raw)
                except InvalidOperation:
                    opening_qty = Decimal("0")
            if is_service or sku_key in extra_skus or track_serial:
                opening_spec = (False, None, None)
            elif opening_qty > 0:
                unit_cost_val = (
                    Decimal(row["unit_cost"])
                    if (row.get("unit_cost") or "").strip()
                    else (purchase_price if purchase_price > 0 else None)
                )
                opening_spec = (True, opening_qty, unit_cost_val)
            else:
                opening_spec = (False, None, None)
            if existing is not None:
                existing.name = row["name"]
                barcode = (row.get("barcode") or "").strip()
                if barcode:
                    old_barcode = (existing.barcode or "").strip()
                    if barcode.casefold() != old_barcode.casefold():
                        barcode_rewrites[existing.pk] = barcode
                        existing.barcode = ""
                    else:
                        existing.barcode = barcode
                if (row.get("hsn_code") or "").strip():
                    existing.hsn_code = row.get("hsn_code") or ""
                if (row.get("description") or "").strip():
                    existing.description = row.get("description") or ""
                if category_str:
                    existing.category = categories.get(category_str.casefold())
                if brand_str:
                    existing.brand = brands.get(brand_str.casefold())
                if unit_obj is not None:
                    existing.unit = unit_obj
                if alt_obj is not None:
                    existing.alternate_unit = alt_obj
                if (row.get("conversion_rate") or "").strip():
                    existing.conversion_rate = Decimal(row.get("conversion_rate") or "1")
                if (row.get("gst_rate") or "").strip():
                    existing.gst_rate = Decimal(row.get("gst_rate") or "0")
                if (row.get("purchase_price") or "").strip():
                    existing.purchase_price = purchase_price
                if (row.get("selling_price") or "").strip():
                    existing.selling_price = Decimal(row.get("selling_price") or "0")
                if (row.get("mrp") or "").strip():
                    existing.mrp = Decimal(row.get("mrp") or "0")
                if (row.get("wholesale_price") or "").strip():
                    existing.wholesale_price = Decimal(row.get("wholesale_price") or "0")
                if (row.get("default_discount_percent") or "").strip():
                    existing.default_discount_percent = Decimal(row.get("default_discount_percent") or "0")
                if (row.get("reorder_level") or "").strip():
                    existing.reorder_level = Decimal(row.get("reorder_level") or "0")
                if row.get("selling_tax_inclusive") not in (None, ""):
                    existing.selling_tax_inclusive = _inclusive_flag(row.get("selling_tax_inclusive"))
                if row.get("purchase_tax_inclusive") not in (None, ""):
                    existing.purchase_tax_inclusive = _inclusive_flag(row.get("purchase_tax_inclusive"))
                existing.custom_fields = _custom_fields_for_commit(job, row, existing.custom_fields)
                existing.updated_by = user
                existing.updated_at = now
                updates.append(existing)
                update_opening.append(opening_spec)
                # R4-013: this row updated a pre-existing product — void must NOT
                # delete/deactivate it.
                row["_import_created"] = False
            else:
                row["_import_created"] = True
                creates.append(
                    Product(
                        company=job.company,
                        name=row["name"],
                        sku=row.get("sku", "") or "",
                        barcode=row.get("barcode", "") or "",
                        hsn_code=row.get("hsn_code", "") or "",
                        description=row.get("description", "") or "",
                        category=categories.get(category_str.casefold()) if category_str else None,
                        brand=brands.get(brand_str.casefold()) if brand_str else None,
                        unit=unit_obj,
                        alternate_unit=alt_obj,
                        conversion_rate=Decimal(row.get("conversion_rate") or "1"),
                        gst_rate=Decimal(row.get("gst_rate") or "0"),
                        purchase_price=purchase_price,
                        selling_price=Decimal(row.get("selling_price") or "0"),
                        mrp=Decimal(row.get("mrp") or "0"),
                        wholesale_price=Decimal(row.get("wholesale_price") or "0"),
                        default_discount_percent=Decimal(row.get("default_discount_percent") or "0"),
                        reorder_level=Decimal(row.get("reorder_level") or "0"),
                        product_type=Product.ProductType.SERVICE if is_service else Product.ProductType.GOODS,
                        track_inventory=bool(track_inventory) and not is_service,
                        track_batch=bool(track_batch) and not is_service,
                        track_serial=bool(track_serial) and not is_service,
                        selling_tax_inclusive=_inclusive_flag(row.get("selling_tax_inclusive")),
                        purchase_tax_inclusive=_inclusive_flag(row.get("purchase_tax_inclusive")),
                        custom_fields=_custom_fields_for_commit(job, row),
                        created_by=user,
                        updated_by=user,
                        created_at=now,
                        updated_at=now,
                    )
                )
                create_opening.append(opening_spec)

        try:
            if updates:
                Product.objects.bulk_update(
                    updates,
                    [
                        "name", "barcode", "hsn_code", "description", "category", "brand",
                        "unit", "alternate_unit", "conversion_rate", "gst_rate",
                        "purchase_price", "selling_price", "mrp", "wholesale_price",
                        "default_discount_percent", "reorder_level",
                        "selling_tax_inclusive", "purchase_tax_inclusive", "custom_fields",
                        "updated_by", "updated_at",
                    ],
                    batch_size=BULK_BATCH,
                )
                if barcode_rewrites:
                    delayed = []
                    for product in updates:
                        new_barcode = barcode_rewrites.get(product.pk)
                        if not new_barcode:
                            continue
                        product.barcode = new_barcode
                        delayed.append(product)
                    if delayed:
                        Product.objects.bulk_update(delayed, ["barcode"], batch_size=BULK_BATCH)
            created_products = Product.objects.bulk_create(creates, batch_size=BULK_BATCH) if creates else []
        except IntegrityError:
            raise BusinessRuleError(
                "Could not import items because a SKU or barcode is no longer unique. Preview the file again."
            ) from None
        opening_items = [
            {"product": product, "quantity": qty, "unit_cost": cost}
            for product, (has_opening, qty, cost) in list(zip(created_products, create_opening)) + list(zip(updates, update_opening))
            if has_opening
        ]
        ImportService._post_opening_items(job, opening_items, user)
        ImportService._post_extra_opening(job, created_products, user)
        return len(preview)


    @staticmethod
    def _post_extra_opening(job, created_products, user):
        extra = job.extra_sheets or {}
        lots = extra.get("opening_lots") or []
        serials = extra.get("opening_serials") or []
        if not lots and not serials:
            return
        from inventory.item_stock import match_warehouse
        from inventory.services import InventoryService
        from masters.models import Product as ProductModel

        by_sku = {(p.sku or "").casefold(): p for p in created_products if p.sku}
        batched_skus = {
            str(r.get("sku") or r.get("item code") or "").strip().casefold()
            for r in lots
            if str(r.get("batch_no") or r.get("batch no") or r.get("batch") or "").strip()
        }
        if batched_skus:
            ProductModel.objects.filter(
                company=job.company, sku__in=[p.sku for p in created_products if (p.sku or "").casefold() in batched_skus]
            ).update(track_batch=True)
            for product in created_products:
                if (product.sku or "").casefold() in batched_skus:
                    product.track_batch = True
        serial_skus = {str(r.get("sku") or r.get("item code") or "").strip().casefold() for r in serials}
        if serial_skus:
            ProductModel.objects.filter(
                company=job.company, sku__in=[p.sku for p in created_products if (p.sku or "").casefold() in serial_skus]
            ).update(track_serial=True)
            for product in created_products:
                if (product.sku or "").casefold() in serial_skus:
                    product.track_serial = True
        for raw in lots:
            row = {str(k).strip().lower(): v for k, v in raw.items()}
            sku = str(row.get("sku") or row.get("item code") or "").strip()
            product = by_sku.get(sku.casefold())
            if product is None:
                continue
            if product.track_serial:
                raise BusinessRuleError("Serial numbers are required for serial-tracked opening stock.")
            warehouse = match_warehouse(job.company, str(row.get("godown") or row.get("warehouse") or ""))
            qty = Decimal(str(row.get("quantity") or row.get("qty") or "0"))
            expiry_raw = str(row.get("expiry_date") or row.get("expiry") or "").strip()
            mfg_raw = str(row.get("manufacturing_date") or row.get("mfg") or "").strip()
            as_of_raw = str(row.get("as_of") or row.get("as of") or "").strip()
            expiry = parse_date(expiry_raw[:10]) if expiry_raw else None
            mfg = parse_date(mfg_raw[:10]) if mfg_raw else None
            as_of = parse_date(as_of_raw[:10]) if as_of_raw else None
            cost = row.get("unit_cost") or row.get("unit cost")
            InventoryService.post_opening(
                company=job.company,
                product=product,
                quantity=qty,
                unit_cost=cost if cost not in (None, "") else None,
                warehouse=warehouse,
                batch_no=str(row.get("batch_no") or row.get("batch") or ""),
                expiry_date=expiry,
                manufacturing_date=mfg,
                as_of=as_of,
                user=user,
                reference_type="import",
                reference_id=job.pk,
            )
        grouped = {}
        for raw in serials:
            row = {str(k).strip().lower(): v for k, v in raw.items()}
            sku = str(row.get("sku") or row.get("item code") or "").strip().casefold()
            godown = str(row.get("godown") or row.get("warehouse") or "").strip().casefold()
            grouped.setdefault((sku, godown), []).append(row)
        for (sku, _godown), rows in grouped.items():
            product = by_sku.get(sku)
            if product is None:
                continue
            warehouse = match_warehouse(job.company, str(rows[0].get("godown") or rows[0].get("warehouse") or ""))
            numbers = [str(r.get("serial_no") or r.get("serial") or "").strip() for r in rows]
            cost = rows[0].get("unit_cost")
            InventoryService.post_opening(
                company=job.company,
                product=product,
                quantity=len(numbers),
                unit_cost=cost,
                warehouse=warehouse,
                serial_numbers=numbers,
                user=user,
                reference_type="import",
                reference_id=job.pk,
            )
            # R4-014: tag the units this import created for a precise void.
            from inventory.models import SerialNumber as _SN

            _SN.objects.filter(
                company=job.company, product=product,
                serial_number__in=[n for n in numbers if n], import_job_ref__isnull=True,
            ).update(import_job_ref=job.pk)

    @staticmethod
    def _commit_opening_stock(job, preview, user):
        needed = {(row.get("sku") or "").strip().casefold() for row in preview}
        products_by_sku = {
            (p.sku or "").casefold(): p
            for p in Product.objects.filter(company=job.company).exclude(sku="")
            if (p.sku or "").casefold() in needed
        }
        items = []
        for row in preview:
            sku = (row.get("sku") or "").strip()
            product = products_by_sku.get(sku.casefold())
            if not product:
                raise BusinessRuleError(f"No product with sku '{sku}' found.")
            quantity = Decimal(row["quantity"])
            unit_cost = Decimal(row["unit_cost"]) if (row.get("unit_cost") or "").strip() else None
            serials = _parse_serial_numbers(row.get("serial_no"))
            if product.track_serial or serials:
                movement = InventoryService.post_opening(
                    company=job.company,
                    product=product,
                    quantity=quantity,
                    unit_cost=unit_cost,
                    serial_numbers=serials,
                    user=user,
                    reference_type="import",
                    reference_id=job.pk,
                )
                if serials:
                    # R4-014: tag the units this import created for a precise void.
                    from inventory.models import SerialNumber

                    SerialNumber.objects.filter(
                        company=job.company, product=product,
                        serial_number__in=serials, import_job_ref__isnull=True,
                    ).update(import_job_ref=job.pk)
                if job.company.accounting_enabled:
                    from accounting.services import PostingService

                    PostingService.post_opening_stock(movement, user)
            else:
                items.append({
                    "product": product,
                    "quantity": quantity,
                    "unit_cost": unit_cost,
                })
        ImportService._post_opening_items(job, items, user)
        return len(preview)

    @staticmethod
    def _movement_is_unused(movement) -> str:
        """Return an error string if this opening movement cannot be reversed, else ''."""
        from inventory.models import InventoryCostLayer

        qty = abs(Decimal(str(movement.quantity or 0)))
        if qty <= 0:
            return ""
        layers = list(
            InventoryCostLayer.objects.select_for_update().filter(source_movement=movement)
        )
        if layers:
            remaining = sum((Decimal(str(layer.qty_remaining or 0)) for layer in layers), Decimal("0"))
            if remaining < qty:
                return (
                    f"Cannot void: stock from '{movement.product.name}' has already been "
                    "used (sold/issued). Void unused rows individually, or use Adjust Stock "
                    "to correct quantities."
                )
        available = InventoryService.available_quantity(
            movement.company, movement.product, warehouse=movement.warehouse, batch=movement.batch
        )
        if available < qty:
            return (
                f"Cannot void: insufficient available stock for '{movement.product.name}' "
                "to reverse the opening quantity. Use Adjust Stock to correct quantities."
            )
        return ""

    @staticmethod
    def _reverse_import_movement(job, movement, user):
        qty = abs(Decimal(str(movement.quantity or 0)))
        if qty <= 0:
            StockMovement.objects.filter(pk=movement.pk).update(reference_type="import_voided")
            return
        InventoryService.post_movement(
            company=job.company,
            product=movement.product,
            movement_type=MovementType.ADJUSTMENT,
            quantity=-qty,
            reason=f"Void import #{job.pk}",
            reference_type="import_void",
            reference_id=job.pk,
            user=user,
            warehouse=movement.warehouse,
            batch=movement.batch,
            skip_negative_check=False,
        )
        if job.company.accounting_enabled:
            from accounting.models import JournalEntry
            from accounting.services import PostingService

            entries = JournalEntry.objects.filter(
                company=job.company,
                source_type="STOCK_MOVEMENT",
                source_id=movement.id,
                purpose="OPENING_STOCK",
                status=JournalEntry.Status.POSTED,
            )
            for entry in entries:
                PostingService.reverse(entry, user=user)
        StockMovement.objects.filter(pk=movement.pk).update(reference_type="import_voided")

    @staticmethod
    def _cleanup_imported_product(job, product, user):
        still_referenced = product.is_referenced()
        if still_referenced:
            product.sku = ""
            product.barcode = ""
            product.status = Product.Status.INACTIVE
            product.updated_by = user
            product.save(update_fields=["sku", "barcode", "status", "updated_by", "updated_at"])
        else:
            product.delete()

    @staticmethod
    def _find_imported_product(job, sku="", name=""):
        product = None
        sku = (sku or "").strip()
        name = (name or "").strip()
        if sku:
            product = Product.objects.filter(company=job.company, sku__iexact=sku).first()
        if product is None and name:
            product = Product.objects.filter(company=job.company, name__iexact=name).first()
        return product

    @staticmethod
    @transaction.atomic
    def void(job: ImportJob, user):
        """
        Reverse a committed PRODUCTS / OPENING_STOCK import when stock layers
        are still intact (nothing sold from the imported opening qty).
        """
        job = ImportJob.objects.select_for_update().get(pk=job.pk)
        if job.kind not in (ImportJob.Kind.PRODUCTS, ImportJob.Kind.OPENING_STOCK):
            raise BusinessRuleError("Only product or opening-stock imports can be voided.")
        if job.status != ImportJob.Status.COMMITTED:
            raise BusinessRuleError("Only a committed import can be voided.")

        movements = list(
            StockMovement.objects.select_for_update()
            .filter(
                company=job.company,
                reference_type="import",
                reference_id=str(job.pk),
                movement_type=MovementType.OPENING_STOCK,
            )
            .select_related("product")
        )

        for movement in movements:
            blocked = ImportService._movement_is_unused(movement)
            if blocked:
                raise BusinessRuleError(blocked)
            ImportService._reverse_import_movement(job, movement, user)

        # R4-014: serial-tracked opening stock received SerialNumber rows that
        # the movement reversal alone doesn't touch — scrap exactly the units
        # this import created (tagged with import_job_ref at commit time), plus a
        # best-effort fallback for pre-tag imports, so nothing orphans AVAILABLE.
        from inventory.models import SerialNumber

        SerialNumber.objects.filter(
            company=job.company,
            import_job_ref=job.pk,
            status=SerialNumber.Status.AVAILABLE,
        ).update(status=SerialNumber.Status.SCRAPPED, updated_by=user)

        _void_serials: list[str] = []
        for _r in (job.preview if isinstance(job.preview, list) else []):
            _void_serials += _parse_serial_numbers(_r.get("serial_no") or _r.get("serial_numbers"))
        for _r in (job.extra_sheets or {}).get("opening_serials") or []:
            _void_serials += _parse_serial_numbers(
                _r.get("serial_no") or _r.get("serial number") or _r.get("serial_numbers")
            )
        _void_serials = [s for s in {s.strip() for s in _void_serials} if s]
        if _void_serials:
            SerialNumber.objects.filter(
                company=job.company,
                serial_number__in=_void_serials,
                import_job_ref__isnull=True,
                status=SerialNumber.Status.AVAILABLE,
            ).update(status=SerialNumber.Status.SCRAPPED, updated_by=user)

        if job.kind == ImportJob.Kind.PRODUCTS:
            preview_rows = job.preview if isinstance(job.preview, list) else []
            for row in preview_rows:
                # R4-013: never delete/deactivate a product this import only
                # *updated*. Rows committed before this marker existed fall back
                # to the old (best-effort) behaviour.
                if row.get("_import_created") is False:
                    continue
                product = ImportService._find_imported_product(
                    job, sku=row.get("sku", ""), name=row.get("name", "")
                )
                if product is None:
                    continue
                ImportService._cleanup_imported_product(job, product, user)

        job.status = ImportJob.Status.VOIDED
        job.failure_reason = f"Voided by user at {timezone.now().isoformat()}"
        job.voided_rows = [
            {"sku": (m.product.sku or ""), "name": m.product.name}
            for m in movements
        ]
        job.updated_by = user
        job.save(update_fields=["status", "failure_reason", "voided_rows", "updated_by", "updated_at"])
        AuditService.log(
            company=job.company, user=user, action="VOID",
            entity_type="ImportJob", entity_id=job.pk,
            description=f"Voided import job #{job.pk} ({job.kind})",
            metadata={"movements_reversed": len(movements)},
        )
        return job

    @staticmethod
    @transaction.atomic
    def void_rows(job: ImportJob, user, *, skus: list[str]):
        """Reverse OPENING_STOCK (and PRODUCTS cleanup) for specific SKUs only."""
        job = ImportJob.objects.select_for_update().get(pk=job.pk)
        if job.kind not in (ImportJob.Kind.PRODUCTS, ImportJob.Kind.OPENING_STOCK):
            raise BusinessRuleError("Only product or opening-stock imports can be voided.")
        if job.status != ImportJob.Status.COMMITTED:
            raise BusinessRuleError("Only a committed import can be voided.")
        wanted = [(s or "").strip() for s in skus if (s or "").strip()]
        if not wanted:
            raise BusinessRuleError("Provide a non-empty skus array.")
        movements = list(
            StockMovement.objects.select_for_update()
            .filter(
                company=job.company,
                reference_type="import",
                reference_id=str(job.pk),
                movement_type=MovementType.OPENING_STOCK,
            )
            .select_related("product")
        )
        from collections import defaultdict

        by_sku: dict[str, list] = defaultdict(list)
        for m in movements:
            by_sku[(m.product.sku or "").casefold()].append(m)

        voided = list(job.voided_rows or [])
        blocked = []
        reversed_count = 0
        for sku in wanted:
            key = sku.casefold()
            lots = by_sku.get(key) or []
            product = ImportService._find_imported_product(job, sku=sku)
            if not lots:
                if job.kind == ImportJob.Kind.PRODUCTS and product is not None:
                    ImportService._cleanup_imported_product(job, product, user)
                    voided.append({"sku": sku, "name": product.name})
                    continue
                blocked.append({"sku": sku, "reason": f"No imported opening-stock row found for sku '{sku}'."})
                continue
            lot_blocked = False
            for movement in lots:
                reason = ImportService._movement_is_unused(movement)
                if reason:
                    blocked.append({"sku": sku, "reason": reason})
                    lot_blocked = True
                    break
            if lot_blocked:
                continue
            for movement in lots:
                ImportService._reverse_import_movement(job, movement, user)
                reversed_count += 1
            if job.kind == ImportJob.Kind.PRODUCTS and product is not None:
                ImportService._cleanup_imported_product(job, product, user)
            voided.append({
                "sku": sku,
                "name": getattr(product or getattr(lots[0], "product", None), "name", ""),
            })

        job.voided_rows = voided
        remaining = (
            StockMovement.objects.filter(
                company=job.company,
                reference_type="import",
                reference_id=str(job.pk),
                movement_type=MovementType.OPENING_STOCK,
            ).exists()
        )
        # Full VOIDED only when every requested sku succeeded and nothing remains.
        if not remaining and not blocked and job.kind in (
            ImportJob.Kind.OPENING_STOCK, ImportJob.Kind.PRODUCTS,
        ):
            job.status = ImportJob.Status.VOIDED
            job.failure_reason = f"Voided by user at {timezone.now().isoformat()}"
        job.updated_by = user
        job.save(update_fields=["status", "failure_reason", "voided_rows", "updated_by", "updated_at"])
        AuditService.log(
            company=job.company, user=user, action="VOID",
            entity_type="ImportJob", entity_id=job.pk,
            description=f"Voided {reversed_count} row(s) of import job #{job.pk}",
            metadata={"skus": wanted, "reversed": reversed_count, "blocked": blocked},
        )
        if reversed_count == 0 and blocked:
            raise BusinessRuleError(blocked[0]["reason"])
        return {"voided": [v["sku"] for v in voided], "blocked": blocked, "job": job}


class BillImportService:
    @staticmethod
    def start_extraction(job: ImportJob):
        if job.kind not in ImportJob.BILL_KINDS:
            raise BusinessRuleError("Only purchase/sales bill jobs can be extracted.")
        allowed = {ImportJob.Status.UPLOADED, ImportJob.Status.FAILED}
        # B3-005: a hard worker kill (OOM / SIGKILL / lost broker message)
        # strands the job in EXTRACTING with no recovery path. Treat an
        # EXTRACTING job whose last update predates the task hard limit (plus a
        # margin) as retryable rather than permanently wedged.
        if job.status == ImportJob.Status.EXTRACTING and job.updated_at is not None:
            from django.utils import timezone as _tz

            stale_after = timedelta(seconds=480 + 120)
            if _tz.now() - job.updated_at > stale_after:
                allowed.add(ImportJob.Status.EXTRACTING)
        if job.status not in allowed:
            raise BusinessRuleError(
                "Extraction can only start from UPLOADED or FAILED "
                "(or a stalled EXTRACTING job)."
            )
        job.status = ImportJob.Status.EXTRACTING
        job.failure_reason = ""
        job.errors = []
        job.save(update_fields=["status", "failure_reason", "errors", "updated_at"])
        from django.conf import settings
        from imports.tasks import extract_purchase_bill_task

        pk, company_id = job.pk, job.company_id
        if getattr(settings, "CELERY_TASK_ALWAYS_EAGER", False):
            extract_purchase_bill_task(pk, company_id=company_id)
            job.refresh_from_db()
        else:
            from django.db import transaction as db_transaction
            db_transaction.on_commit(
                lambda: extract_purchase_bill_task.delay(pk, company_id=company_id)
            )
        return job

    @staticmethod
    def parse_structured_file(job: ImportJob):
        """CSV/XLSX bill-import path (§4.1) — deterministic, no LLM, no
        clarification loop (there's no OCR ambiguity to resolve)."""
        if job.kind not in ImportJob.BILL_KINDS:
            raise BusinessRuleError("Structured bill parsing is only for bill imports.")
        with job.file.file.open("rb") as handle:
            raw = handle.read()
        filename = job.file.original_name or job.file.file.name or ""
        rows, meta = _read_structured_bill(raw, filename)
        if not rows:
            raise BusinessRuleError("The file has no data rows.")

        raw_lines = [_map_structured_row(row) for row in rows]
        preview_lines, errors = BillImportService._build_preview_lines(raw_lines)
        answers = _infer_qty_answers(preview_lines, tolerance=Decimal("0.50")) or {}
        if not answers.get("qty_formula"):
            from imports.qty_formula import _union_keys
            keys = _union_keys(preview_lines)
            if "cs" in keys and "upc" in keys:
                answers = {"qty_formula": "cs*upc+quantity"}
            elif "cs" in keys and "quantity" in keys:
                answers = {"qty_formula": "cs+quantity"}
        formula_key = _apply_cross_check(
            preview_lines, answers, tolerance=Decimal("0.50"), reconcile_print=False
        )
        errors = _refresh_line_validity(preview_lines)
        preview = {
            "supplier_name": meta.get("supplier_name") or "",
            "supplier_gstin": meta.get("supplier_gstin") or "",
            "buyer_name": meta.get("buyer_name") or "",
            "buyer_gstin": meta.get("buyer_gstin") or "",
            "bill_number": meta.get("bill_number") or "",
            "bill_date": meta.get("bill_date") or "",
            "extraction_confidence": 1.0,
            "low_confidence_accepted": True,
            "column_headers": list(rows[0].keys()) if rows else [],
            "printed_line_count": len(preview_lines),
            "resolved_formula": formula_key,
            "lines": preview_lines,
        }
        included = [ln for ln in preview_lines if ln.get("include")]
        job.preview = preview
        job.errors = errors
        job.total_rows = len(preview_lines)
        job.valid_rows = len(included)
        job.error_rows = len(errors)
        job.clarifications = []
        job.status = ImportJob.Status.PREVIEWED
        job.failure_reason = ""
        job.save()
        return job

    @staticmethod
    def _build_preview_lines(raw_lines: list[dict]) -> tuple[list[dict], list[dict]]:
        preview_lines, errors = [], []
        rate_warnings: list[str] = []
        for index, raw_line in enumerate(raw_lines, start=1):
            line = _preview_bill_line(raw_line, index=index, rate_warnings=rate_warnings)
            if line.get("include"):
                line_errors = _bill_line_errors(line)
                if line_errors:
                    errors.append({"row": index, "errors": line_errors, "data": line})
                    line["include"] = False
            preview_lines.append(line)
        return preview_lines, errors

    @staticmethod
    def apply_extraction(job: ImportJob, payload: dict):
        raw_lines = list(payload.get("lines") or [])
        column_headers = list(payload.get("column_headers") or [])
        gstin = (
            str(payload.get("supplier_gstin") or "").strip()
            if job.kind == ImportJob.Kind.PURCHASE_BILL
            else str(payload.get("buyer_gstin") or "").strip()
        )
        # For a sales bill the "vendor" whose layout we're learning is still
        # the document's issuer (the seller) — that's who prints the columns.
        template_gstin = str(payload.get("supplier_gstin") or "").strip() or gstin
        template = _match_bill_template(job.company, template_gstin, column_headers)

        preview = {
            "supplier_name": str(payload.get("supplier_name") or "").strip(),
            "supplier_gstin": str(payload.get("supplier_gstin") or "").strip(),
            "buyer_name": str(payload.get("buyer_name") or "").strip(),
            "buyer_gstin": str(payload.get("buyer_gstin") or "").strip(),
            "bill_number": str(payload.get("bill_number") or "").strip(),
            "bill_date": str(payload.get("bill_date") or "").strip(),
            "extraction_confidence": _parse_extraction_confidence(payload.get("confidence")),
            "low_confidence_accepted": False,
            "column_headers": column_headers,
            "printed_line_count": payload.get("printed_line_count"),
            "lines": [],
        }
        direction_warning = _infer_direction_warning(job.kind, payload, job.company)
        if direction_warning:
            preview["direction_warning"] = direction_warning

        preview_lines, errors = BillImportService._build_preview_lines(raw_lines)
        if _skip_simple_qty_template(template, preview_lines):
            template = None
        tolerance = template.rounding_tolerance if template is not None else Decimal("0.50")

        if template is not None:
            job.bill_template = template
            formula_key = _apply_cross_check(
                preview_lines, _template_answers(template), tolerance=tolerance
            )
            errors = _refresh_line_validity(preview_lines)
            preview["resolved_formula"] = formula_key
            preview["lines"] = preview_lines
            job.status = ImportJob.Status.PREVIEWED
            job.clarifications = []
        else:
            inferred = _infer_qty_answers(preview_lines, tolerance=tolerance)
            if inferred is not None:
                formula_key = _apply_cross_check(preview_lines, inferred, tolerance=tolerance)
                errors = _refresh_line_validity(preview_lines)
                preview["resolved_formula"] = formula_key
                preview["resolved_answers"] = inferred
                preview["qty_formula"] = inferred.get("qty_formula")
                preview["lines"] = preview_lines
                job.clarifications = []
                job.status = ImportJob.Status.PREVIEWED
            else:
                clarifications = _detect_clarifications({
                    "column_headers": column_headers,
                    "lines": preview_lines,
                })
                if clarifications:
                    # Cross-check can't run correctly yet (quantity may still need
                    # recombining) — hold in NEEDS_CLARIFICATION until answered.
                    preview["lines"] = preview_lines
                    job.clarifications = clarifications
                    job.status = ImportJob.Status.NEEDS_CLARIFICATION
                else:
                    formula_key = _apply_cross_check(preview_lines, {}, tolerance=tolerance)
                    errors = _refresh_line_validity(preview_lines)
                    preview["resolved_formula"] = formula_key
                    preview["lines"] = preview_lines
                    job.clarifications = []
                    job.status = ImportJob.Status.PREVIEWED

        included = [ln for ln in preview["lines"] if ln.get("include")]
        job.preview = preview
        job.errors = errors
        job.total_rows = len(preview["lines"])
        job.valid_rows = len(included)
        job.error_rows = len(errors)
        job.failure_reason = ""
        job.save()
        return job

    @staticmethod
    def answer_clarifications(job: ImportJob, answers: dict, user=None):
        """Apply document-level clarification answers (§4.3), recompute
        quantities/cross-check, and move the job from NEEDS_CLARIFICATION to
        PREVIEWED."""
        if job.kind not in ImportJob.BILL_KINDS:
            raise BusinessRuleError("Clarifications are only for bill imports.")
        if job.status != ImportJob.Status.NEEDS_CLARIFICATION:
            raise BusinessRuleError("This job has no pending clarifications.")
        if not isinstance(answers, dict):
            raise BusinessRuleError("answers must be an object of {field: value}.")

        preview = dict(job.preview or {})
        clarifications = list(job.clarifications or [])
        resolved_answers = {k: v for k, v in answers.items() if v not in (None, "")}
        for item in clarifications:
            field = item.get("field")
            if field in answers:
                item["answer"] = answers[field]
            if item.get("answer"):
                resolved_answers[field] = item["answer"]
        job.clarifications = clarifications

        preview_lines = list(preview.get("lines") or [])
        formula_key = _apply_cross_check(preview_lines, resolved_answers, tolerance=Decimal("0.50"))
        errors = _refresh_line_validity(preview_lines)
        preview["resolved_formula"] = formula_key
        preview["resolved_answers"] = resolved_answers
        preview["lines"] = preview_lines

        job.preview = preview
        job.errors = errors
        job.status = ImportJob.Status.PREVIEWED
        job.valid_rows = len([ln for ln in preview_lines if ln.get("include")])
        job.error_rows = len(errors)
        if user is not None:
            job.updated_by = user
        job.save()
        return job

    @staticmethod
    def mark_failed(job: ImportJob, reason: str):
        job.status = ImportJob.Status.FAILED
        job.failure_reason = (reason or "Extraction failed.")[:2000]
        job.save(update_fields=["status", "failure_reason", "updated_at"])
        return job

    @staticmethod
    def update_preview(job: ImportJob, data: dict, user=None):
        if job.kind not in ImportJob.BILL_KINDS:
            raise BusinessRuleError("Preview update is only for bill imports.")
        if job.status != ImportJob.Status.PREVIEWED:
            raise BusinessRuleError("Only previewed bill jobs can be edited.")

        preview = dict(job.preview or {})
        if not isinstance(preview, dict):
            preview = {"lines": []}

        if "supplier_name" in data:
            preview["supplier_name"] = str(data.get("supplier_name") or "").strip()
        if "supplier_gstin" in data:
            preview["supplier_gstin"] = str(data.get("supplier_gstin") or "").strip()
        if "customer_name" in data:
            preview["customer_name"] = str(data.get("customer_name") or "").strip()
        if "bill_number" in data:
            preview["bill_number"] = str(data.get("bill_number") or "").strip()
        if "bill_date" in data:
            preview["bill_date"] = str(data.get("bill_date") or "").strip()
        if "low_confidence_accepted" in data:
            preview["low_confidence_accepted"] = bool(data.get("low_confidence_accepted"))

        if "lines" in data:
            lines_in = data.get("lines") or []
            if not isinstance(lines_in, list):
                raise BusinessRuleError("lines must be an array.")
            lines, errors = [], []
            rate_warnings: list[str] = []
            for index, raw_line in enumerate(lines_in, start=1):
                if not isinstance(raw_line, dict):
                    continue
                line = _preview_bill_line(raw_line, index=index, rate_warnings=rate_warnings)
                # User edits may toggle include even with previously blank fields.
                if "include" in raw_line:
                    line["include"] = bool(raw_line.get("include"))
                line_errors = _bill_line_errors(line) if line.get("include") else []
                if line_errors:
                    errors.append({"row": index, "errors": line_errors, "data": line})
                    line["include"] = False
                lines.append(line)
            preview["lines"] = lines
            if rate_warnings:
                preview["warnings"] = rate_warnings
            else:
                preview.pop("warnings", None)
            job.errors = errors
            job.total_rows = len(lines)
            job.valid_rows = len([ln for ln in lines if ln.get("include")])
            job.error_rows = len(errors)

        job.preview = preview

        supplier_id = data.get("supplier_id", data.get("supplier"))
        if supplier_id is not None and supplier_id != "":
            try:
                supplier = Supplier.objects.get(pk=supplier_id, company=job.company)
            except (Supplier.DoesNotExist, ValueError, TypeError):
                raise BusinessRuleError("Invalid supplier.")
            job.supplier = supplier

        customer_id = data.get("customer_id", data.get("customer"))
        if customer_id is not None and customer_id != "":
            try:
                customer = Customer.objects.get(pk=customer_id, company=job.company)
            except (Customer.DoesNotExist, ValueError, TypeError):
                raise BusinessRuleError("Invalid customer.")
            job.customer = customer

        if user is not None:
            job.updated_by = user
        job.save()
        return job

    @staticmethod
    def _match_or_create_product(company, line: dict, user, *, direction: str = "PURCHASE") -> tuple[Product, bool]:
        sku = str(line.get("sku") or "").strip()
        name = str(line.get("name") or "").strip()
        product = None
        if sku:
            product = Product.objects.filter(company=company).filter(
                Q(sku__iexact=sku) | Q(barcode__iexact=sku)
            ).first()
        if product is None and name:
            product = Product.objects.filter(company=company, name__iexact=name).first()

        unit_price = _as_decimal(line.get("unit_price"), "0")
        gst_rate = _normalize_gst_rate(line.get("gst_rate"))
        mrp = _as_decimal(line.get("mrp"), "0")
        hsn = str(line.get("hsn_code") or "").strip()
        # A purchase bill's unit price is a cost; a sales bill's is a selling
        # price — never conflate the two on the product master.
        price_field = "purchase_price" if direction == "PURCHASE" else "selling_price"

        if product is None:
            product = Product.objects.create(
                company=company,
                name=name,
                sku=sku,
                hsn_code=hsn,
                gst_rate=gst_rate,
                purchase_price=unit_price if direction == "PURCHASE" else Decimal("0"),
                selling_price=unit_price if direction == "SALES" else Decimal("0"),
                mrp=mrp,
                created_by=user,
                updated_by=user,
            )
            return product, True

        updates = []
        # Matched products: only fill empty HSN; never overwrite gst/price from OCR noise.
        if hsn and not product.hsn_code:
            product.hsn_code = hsn
            updates.append("hsn_code")
        if line.get("update_master"):
            if unit_price > 0 and getattr(product, price_field) != unit_price:
                setattr(product, price_field, unit_price)
                updates.append(price_field)
            if product.gst_rate != gst_rate:
                product.gst_rate = gst_rate
                updates.append("gst_rate")
            if mrp > 0 and product.mrp != mrp:
                product.mrp = mrp
                updates.append("mrp")
        if updates:
            product.updated_by = user
            updates.extend(["updated_by", "updated_at"])
            product.save(update_fields=updates)
        return product, False

    @staticmethod
    def _resolve_supplier(job: ImportJob, user) -> Supplier:
        if job.supplier_id:
            return job.supplier
        preview = job.preview if isinstance(job.preview, dict) else {}
        gstin = str(preview.get("supplier_gstin") or "").strip()
        name = str(preview.get("supplier_name") or "").strip()
        supplier = None
        if gstin:
            supplier = Supplier.objects.filter(company=job.company, gstin__iexact=gstin).first()
        if supplier is None and name:
            supplier = Supplier.objects.filter(company=job.company, name__iexact=name).first()
        if supplier is None:
            if not name:
                raise BusinessRuleError("Select a supplier before committing the purchase bill.")
            supplier = Supplier.objects.create(
                company=job.company,
                name=name,
                gstin=gstin if gstin and GSTIN_RE.match(gstin) else "",
                created_by=user,
                updated_by=user,
            )
        job.supplier = supplier
        job.save(update_fields=["supplier", "updated_at"])
        return supplier

    @staticmethod
    def _resolve_customer(job: ImportJob, user) -> Customer:
        if job.customer_id:
            return job.customer
        preview = job.preview if isinstance(job.preview, dict) else {}
        gstin = str(preview.get("buyer_gstin") or "").strip()
        name = (
            str(preview.get("customer_name") or "").strip()
            or str(preview.get("buyer_name") or "").strip()
        )
        customer = None
        if gstin:
            customer = Customer.objects.filter(company=job.company, gstin__iexact=gstin).first()
        if customer is None and name:
            customer = Customer.objects.filter(company=job.company, name__iexact=name).first()
        if customer is None:
            if not name:
                raise BusinessRuleError("Select a customer before committing the sales bill.")
            customer = Customer.objects.create(
                company=job.company,
                name=name,
                gstin=gstin if gstin and GSTIN_RE.match(gstin) else "",
                created_by=user,
                updated_by=user,
            )
        job.customer = customer
        job.save(update_fields=["customer", "updated_at"])
        return customer

    @staticmethod
    def _save_bill_template(job: ImportJob, preview: dict, user):
        """Persist the confirmed column mapping / formula as this vendor's
        SupplierBillTemplate (§4.4) — learned once, applied silently on every
        later bill from the same GSTIN."""
        gstin = str(preview.get("supplier_gstin") or "").strip()
        if not gstin or not GSTIN_RE.match(gstin):
            return
        formula_key = preview.get("resolved_formula") or SupplierBillTemplate.LineTotalFormula.SIMPLE
        column_headers = preview.get("column_headers") or []
        mapping = dict(preview.get("column_mapping") or {})
        answers = preview.get("resolved_answers") if isinstance(preview.get("resolved_answers"), dict) else {}
        qty_formula = answers.get("qty_formula") or mapping.get("qty_formula")
        if qty_formula:
            mapping["qty_formula"] = qty_formula
        SupplierBillTemplate.objects.update_or_create(
            company=job.company,
            gstin=gstin,
            defaults={
                "party_name": str(preview.get("supplier_name") or "")[:255],
                "line_total_formula": formula_key,
                "column_mapping": mapping,
                "column_signature": column_headers,
                "confirmed_by": user,
                "confirmed_at": timezone.now(),
                "updated_by": user,
            },
        )

    @staticmethod
    def _parse_bill_date(value: str, *, required: bool = True):
        value = (value or "").strip()
        if not value:
            if required:
                raise BusinessRuleError(
                    "Bill date is missing or could not be parsed. Set bill date before committing."
                )
            return timezone.localdate()
        parsed = parse_date(value)
        if parsed:
            return parsed
        for fmt in (
            "%d-%m-%Y",
            "%d/%m/%Y",
            "%d.%m.%Y",
            "%Y/%m/%d",
            "%d-%b-%Y",
            "%d-%B-%Y",
            "%d %b %Y",
            "%d %B %Y",
        ):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        raise BusinessRuleError(
            f"Could not parse bill date '{value}'. Use YYYY-MM-DD or DD/MM/YYYY."
        )

    @staticmethod
    @transaction.atomic
    def commit(job: ImportJob, user):
        if job.kind not in ImportJob.BILL_KINDS:
            raise BusinessRuleError("Not a bill import.")
        job = ImportJob.objects.select_for_update().get(pk=job.pk)
        if job.status != ImportJob.Status.PREVIEWED:
            raise BusinessRuleError("Import must be previewed before commit.")

        preview = job.preview if isinstance(job.preview, dict) else {}
        confidence = _parse_extraction_confidence(preview.get("extraction_confidence", 1))
        if confidence < OCR_BILL_MIN_CONFIDENCE and not preview.get("low_confidence_accepted"):
            raise BusinessRuleError(
                f"OCR confidence ({confidence:.2f}) is below {OCR_BILL_MIN_CONFIDENCE:.2f}. "
                "Review extracted fields and set low_confidence_accepted on preview before commit."
            )

        lines = [ln for ln in (preview.get("lines") or []) if ln.get("include")]
        if not lines:
            raise BusinessRuleError("Select at least one valid line to commit.")

        if job.kind == ImportJob.Kind.SALES_BILL:
            result = BillImportService._commit_sales(job, preview, lines, user)
        else:
            result = BillImportService._commit_purchase(job, preview, lines, user)

        # Learn this vendor's layout for next time (§4.4) — only once we have
        # a resolved formula (i.e. any clarification was actually answered,
        # or the bill needed no clarification and the SIMPLE formula applied).
        if job.bill_template_id is None:
            BillImportService._save_bill_template(job, preview, user)

        job.status = ImportJob.Status.COMMITTED
        job.committed_at = timezone.now()
        job.updated_by = user
        job.save()
        return result

    @staticmethod
    def _commit_purchase(job: ImportJob, preview: dict, lines: list[dict], user) -> dict:
        supplier = BillImportService._resolve_supplier(job, user)
        items_data = []
        products_created = 0
        for line in lines:
            product, created = BillImportService._match_or_create_product(
                job.company, line, user, direction="PURCHASE"
            )
            if created:
                products_created += 1
            if line.get("quantity") in (None, ""):
                raise BusinessRuleError("Quantity is required for each imported purchase line.")
            if line.get("gst_rate") in (None, ""):
                raise BusinessRuleError("GST rate is required for each imported purchase line.")
            items_data.append({
                "product": product,
                "description": product.name,
                "quantity": _as_decimal(line.get("quantity")),
                "unit_price": _as_decimal(line.get("unit_price"), "0"),
                "discount_percent": Decimal("0"),
                "gst_rate": _normalize_gst_rate(line.get("gst_rate"), required=True),
            })

        invoice = PurchaseInvoice.objects.create(
            company=job.company,
            supplier=supplier,
            company_gstin=_resolve_import_company_gstin(job.company, preview, kind=ImportJob.Kind.PURCHASE_BILL),
            purchase_type=PurchaseInvoice.PurchaseType.GST,
            invoice_date=BillImportService._parse_bill_date(
                str(preview.get("bill_date") or ""),
                required=True,
            ),
            supplier_bill_number=str(preview.get("bill_number") or "")[:64],
            notes="Created from purchase bill upload",
            attachment=job.file,
            created_by=user,
            updated_by=user,
        )
        PurchaseService.set_items(invoice, items_data, user)

        job.purchase_invoice = invoice
        AuditService.log(
            company=job.company, user=user, action="IMPORT",
            entity_type="ImportJob", entity_id=job.pk,
            description=f"Imported purchase bill → draft purchase #{invoice.pk}",
            metadata={
                "products_created": products_created,
                "lines": len(items_data),
                "purchase_invoice_id": invoice.pk,
            },
        )
        return {
            "created": len(items_data),
            "products_created": products_created,
            "purchase_invoice_id": invoice.pk,
        }

    @staticmethod
    def _commit_sales(job: ImportJob, preview: dict, lines: list[dict], user) -> dict:
        customer = BillImportService._resolve_customer(job, user)
        items_data = []
        products_created = 0
        for line in lines:
            product, created = BillImportService._match_or_create_product(
                job.company, line, user, direction="SALES"
            )
            if created:
                products_created += 1
            if line.get("quantity") in (None, ""):
                raise BusinessRuleError("Quantity is required for each imported sales line.")
            if line.get("gst_rate") in (None, ""):
                raise BusinessRuleError("GST rate is required for each imported sales line.")
            items_data.append({
                "product": product,
                "description": product.name,
                "quantity": _as_decimal(line.get("quantity")),
                "unit_price": _as_decimal(line.get("unit_price"), "0"),
                "discount_percent": Decimal("0"),
                "gst_rate": _normalize_gst_rate(line.get("gst_rate"), required=True),
            })

        invoice = SalesInvoice.objects.create(
            company=job.company,
            customer=customer,
            company_gstin=_resolve_import_company_gstin(job.company, preview, kind=ImportJob.Kind.SALES_BILL),
            invoice_type=SalesInvoice.InvoiceType.GST,
            invoice_date=BillImportService._parse_bill_date(
                str(preview.get("bill_date") or ""),
                required=True,
            ),
            notes=f"Created from sales bill upload (bill #{str(preview.get('bill_number') or '')[:64]})",
            created_by=user,
            updated_by=user,
        )
        SalesService.set_items(invoice, items_data, user)

        job.sales_invoice = invoice
        AuditService.log(
            company=job.company, user=user, action="IMPORT",
            entity_type="ImportJob", entity_id=job.pk,
            description=f"Imported sales bill → draft sales invoice #{invoice.pk}",
            metadata={
                "products_created": products_created,
                "lines": len(items_data),
                "sales_invoice_id": invoice.pk,
            },
        )
        return {
            "created": len(items_data),
            "products_created": products_created,
            "sales_invoice_id": invoice.pk,
        }
