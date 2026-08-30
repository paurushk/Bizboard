"""Purchase bill upload — LLM extract → preview → draft purchase commit."""

from decimal import Decimal
from unittest.mock import patch

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile

from imports.models import SupplierBillTemplate
from inventory.models import StockBalance, StockMovement
from masters.models import Product, Supplier
from purchases.models import PurchaseInvoice
from sales.models import SalesInvoice
from tests.conftest import make_customer, make_product, make_supplier

pytestmark = pytest.mark.django_db

FAKE_EXTRACT = {
    "supplier_name": "Acme Distributors",
    "supplier_gstin": "",
    "bill_number": "PB-100",
    "bill_date": "2026-07-01",
    "lines": [
        {
            "name": "Ariel Powder 500g",
            "sku": "ARIEL-500",
            "hsn_code": "3402",
            "quantity": "2",
            "unit_price": "120.50",
            "gst_rate": "18",
            "mrp": "150",
        },
        {
            "name": "Baby Rub 10ml",
            "sku": "",
            "hsn_code": "3004",
            "quantity": "5",
            "unit_price": "45",
            "gst_rate": "12",
            "mrp": "0",
        },
    ],
}


def _png_bytes():
    # Minimal 1x1 PNG
    return (
        b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
        b"\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDATx\x9cc\xf8\x0f\x00"
        b"\x00\x01\x01\x00\x05\x18\xd8N\x00\x00\x00\x00IEND\xaeB`\x82"
    )


def _upload_bill(tenant, content=None, name="bill.png", content_type="image/png", supplier_id=None):
    data = {
        "kind": "PURCHASE_BILL",
        "file": SimpleUploadedFile(name, content or _png_bytes(), content_type=content_type),
    }
    if supplier_id is not None:
        data["supplier_id"] = supplier_id
    return tenant.client.post("/api/v1/imports/", data, format="multipart")


@patch("core.services.llm.extract_purchase_bill", return_value={
    "supplier_name": "VTC TRADEWINGS PVT",
    "supplier_gstin": "09AAPCS3897R1ZX",
    "bill_number": "VTAGR-26-1038635",
    "bill_date": "2026-06-11",
    "confidence": 0.82,
    "lines": [
        {
            "name": "Olay NA IGF 40gm",
            "sku": "",
            "hsn_code": "33049990",
            "quantity": "3",
            "unit_price": "140.50",
            "gst_rate": "18",
            "mrp": "199",
            "include": True,
        },
        {
            "name": "Blurry Item",
            "sku": "",
            "hsn_code": "",
            "quantity": "",
            "unit_price": "10",
            "gst_rate": "",
            "mrp": "0",
            "include": False,
        },
    ],
})
def test_purchase_bill_keeps_blank_qty_gst_without_inventing(mock_llm, tenant_a):
    """Unread qty/GST must stay blank — never invent 1 / 18% (BB-000557/692)."""
    resp = _upload_bill(tenant_a)
    assert resp.status_code == 201, resp.data
    lines = resp.data["preview"]["lines"]
    assert lines[0]["quantity"] == "3"
    assert lines[0]["gst_rate"] == "18"
    assert lines[0]["include"] is True
    assert lines[1]["quantity"] == ""
    assert lines[1]["gst_rate"] == ""
    assert lines[1]["include"] is False
    assert resp.data["valid_rows"] == 1


@patch("core.services.llm.extract_purchase_bill", return_value=FAKE_EXTRACT)
def test_purchase_bill_extract_preview(mock_llm, tenant_a):
    resp = _upload_bill(tenant_a)
    assert resp.status_code == 201, resp.data
    assert resp.data["status"] == "PREVIEWED"
    assert resp.data["valid_rows"] == 2
    preview = resp.data["preview"]
    assert preview["bill_number"] == "PB-100"
    assert len(preview["lines"]) == 2
    mock_llm.assert_called_once()


@patch("core.services.llm.extract_purchase_bill", return_value=FAKE_EXTRACT)
def test_purchase_bill_commit_creates_products_and_draft(mock_llm, tenant_a):
    supplier = make_supplier(tenant_a.company, name="Existing Supplier")
    job = _upload_bill(tenant_a, supplier_id=supplier.id).data
    assert job["status"] == "PREVIEWED"

    resp = tenant_a.client.post(f"/api/v1/imports/{job['id']}/commit/")
    assert resp.status_code == 200, resp.data
    assert resp.data["created"] == 2
    assert resp.data["products_created"] == 2
    assert resp.data["purchase_invoice_id"]

    assert Product.objects.filter(company=tenant_a.company, sku="ARIEL-500").exists()
    assert Product.objects.filter(company=tenant_a.company, name="Baby Rub 10ml").exists()

    invoice = PurchaseInvoice.objects.get(pk=resp.data["purchase_invoice_id"])
    assert invoice.status == PurchaseInvoice.Status.DRAFT
    assert invoice.supplier_id == supplier.id
    assert invoice.items.count() == 2
    assert invoice.supplier_bill_number == "PB-100"

    # Stock unchanged until Complete
    assert StockMovement.objects.filter(company=tenant_a.company).count() == 0
    assert StockBalance.objects.filter(company=tenant_a.company).count() == 0


@patch("core.services.llm.extract_purchase_bill", return_value=FAKE_EXTRACT)
def test_purchase_bill_matches_existing_product(mock_llm, tenant_a):
    make_product(tenant_a.company, name="Ariel Powder 500g", sku="ARIEL-500", purchase_price="100")
    supplier = make_supplier(tenant_a.company)
    job = _upload_bill(tenant_a, supplier_id=supplier.id).data

    resp = tenant_a.client.post(f"/api/v1/imports/{job['id']}/commit/")
    assert resp.status_code == 200
    assert resp.data["products_created"] == 1  # only Baby Rub is new
    product = Product.objects.get(company=tenant_a.company, sku="ARIEL-500")
    # Matched products keep master purchase_price (OCR must not overwrite).
    assert product.purchase_price == Decimal("100")


@patch("core.services.llm.extract_purchase_bill", return_value=FAKE_EXTRACT)
def test_purchase_bill_preview_edit_excludes_line(mock_llm, tenant_a):
    supplier = make_supplier(tenant_a.company)
    job = _upload_bill(tenant_a, supplier_id=supplier.id).data
    preview = job["preview"]
    preview["lines"][1]["include"] = False

    patched = tenant_a.client.patch(
        f"/api/v1/imports/{job['id']}/preview/",
        {"lines": preview["lines"], "supplier_id": supplier.id},
        format="json",
    )
    assert patched.status_code == 200
    assert patched.data["valid_rows"] == 1

    resp = tenant_a.client.post(f"/api/v1/imports/{job['id']}/commit/")
    assert resp.status_code == 200
    assert resp.data["created"] == 1
    invoice = PurchaseInvoice.objects.get(pk=resp.data["purchase_invoice_id"])
    assert invoice.items.count() == 1


@patch("core.services.llm.extract_purchase_bill", side_effect=Exception("provider down"))
def test_purchase_bill_extraction_failure(mock_llm, tenant_a):
    resp = _upload_bill(tenant_a)
    assert resp.status_code == 201
    assert resp.data["status"] == "FAILED"
    assert "provider down" in (resp.data.get("failure_reason") or "")


def test_purchase_bill_rejects_unsupported_file_type(tenant_a):
    resp = tenant_a.client.post("/api/v1/imports/", {
        "kind": "PURCHASE_BILL",
        "file": SimpleUploadedFile("data.txt", b"not a bill", content_type="text/plain"),
    }, format="multipart")
    assert resp.status_code == 400


def test_purchase_bill_csv_export_parses_deterministically(tenant_a):
    """Bill Import Redesign Plan §7 Phase 3: a CSV/XLSX export from the
    supplier's own system skips LLM extraction entirely and is parsed
    straight to PREVIEWED — no clarification loop, since there's no OCR
    ambiguity to resolve."""
    csv_body = (
        b"name,sku,hsn_code,quantity,unit_price,gst_rate,mrp\n"
        b"Ariel Powder 500g,ARIEL-500,3402,2,120.50,18,150\n"
        b"Baby Rub 10ml,,3004,5,45,12,0\n"
    )
    resp = tenant_a.client.post("/api/v1/imports/", {
        "kind": "PURCHASE_BILL",
        "file": SimpleUploadedFile("bill.csv", csv_body, content_type="text/csv"),
    }, format="multipart")
    assert resp.status_code == 201
    assert resp.data["status"] == "PREVIEWED"
    assert resp.data["total_rows"] == 2
    assert resp.data["valid_rows"] == 2
    assert resp.data["preview"]["lines"][0]["name"] == "Ariel Powder 500g"


# --- Clarification loop + per-vendor bill templates (Bill Import Redesign Plan §4.3/§4.4) ---

DMS_GSTIN = "09AAPCS3897R1ZX"
DMS_COLUMN_HEADERS = ["HSN", "PCode", "Item Description", "MRP", "Cs", "Pcs", "UPC", "Pc Price", "Gross Amt", "GST %"]


def _dms_extract(*, cs="2", upc="6", pcs="3", price="10", gross="150", headers=True):
    line = {
        "name": "Olay NA IGF 40gm",
        "hsn_code": "33049990",
        "quantity": pcs,
        "unit_price": price,
        "gst_rate": "18",
        "mrp": "199",
        "cs": cs,
        "upc": upc,
    }
    if gross is not None:
        line["printed_gross_amt"] = gross
    return {
        "supplier_name": "VTC TRADEWINGS PVT",
        "supplier_gstin": DMS_GSTIN,
        "bill_number": "VTAGR-1",
        "bill_date": "2026-06-11",
        "confidence": 0.9,
        "column_headers": DMS_COLUMN_HEADERS if headers else [],
        "lines": [line],
    }


@patch("core.services.llm.extract_purchase_bill", return_value=_dms_extract(gross=None))
def test_bill_upload_asks_upc_and_cs_clarifications(mock_llm, tenant_a):
    """A DMS bill with a 'UPC' column and a nonzero 'Cs' on some line is
    exactly the documented ambiguity — must ask, not guess, when the printed
    Gross Amt is missing so the formula cannot be inferred (§4.2/§4.3)."""
    resp = _upload_bill(tenant_a)
    assert resp.status_code == 201, resp.data
    assert resp.data["status"] == "NEEDS_CLARIFICATION"
    fields = {c["field"] for c in resp.data["clarifications"]}
    assert fields == {"qty_formula"}
    option_values = {o["value"] for o in resp.data["clarifications"][0]["options"]}
    assert "quantity" in option_values
    assert "cs*upc+quantity" in option_values


@patch("core.services.llm.extract_purchase_bill", return_value=_dms_extract(gross=None))
def test_bill_clarification_recombines_quantity_and_learns_template(mock_llm, tenant_a):
    """Answering the derived qty formula must recompute quantity as
    (Cs x UPC) + Pcs = 15 and persist a SupplierBillTemplate so the next
    bill from this GSTIN skips the question entirely."""
    job = _upload_bill(tenant_a).data
    assert job["status"] == "NEEDS_CLARIFICATION"

    resp = tenant_a.client.post(
        f"/api/v1/imports/{job['id']}/clarify/",
        {"answers": {"qty_formula": "cs*upc+quantity"}},
        format="json",
    )
    assert resp.status_code == 200, resp.data
    assert resp.data["status"] == "PREVIEWED"
    line = resp.data["preview"]["lines"][0]
    assert line["quantity"] == "15"
    assert line["flags"] == []  # 15 x 10 == printed 150, no mismatch

    commit = tenant_a.client.post(f"/api/v1/imports/{job['id']}/commit/")
    assert commit.status_code == 200, commit.data
    invoice = PurchaseInvoice.objects.get(pk=commit.data["purchase_invoice_id"])
    assert invoice.items.first().quantity == Decimal("15")

    template = SupplierBillTemplate.objects.get(company=tenant_a.company, gstin=DMS_GSTIN)
    assert template.line_total_formula == SupplierBillTemplate.LineTotalFormula.CASE_UNITS_PLUS_LOOSE

    # Second bill from the same vendor: template auto-applies, no questions.
    mock_llm.return_value = _dms_extract(cs="1", upc="6", pcs="0", price="10", gross="60")
    job2 = _upload_bill(tenant_a).data
    assert job2["status"] == "PREVIEWED"
    assert job2["clarifications"] == []
    assert job2["preview"]["lines"][0]["quantity"] == "6"  # (1x6)+0


@patch("core.services.llm.extract_purchase_bill", return_value=_dms_extract(gross="999"))
def test_bill_cross_check_flags_mismatched_printed_total(mock_llm, tenant_a):
    """A printed Gross Amt that doesn't match qty x price once quantity is
    resolved must be flagged for review, not silently trusted (§4.2)."""
    job = _upload_bill(tenant_a).data
    resp = tenant_a.client.post(
        f"/api/v1/imports/{job['id']}/clarify/",
        {"answers": {"qty_formula": "cs*upc+quantity"}},
        format="json",
    )
    line = resp.data["preview"]["lines"][0]
    assert line["quantity"] == "15"
    assert len(line["flags"]) == 1
    assert "999" in line["flags"][0]


@patch("core.services.llm.extract_purchase_bill", return_value=_dms_extract())
def test_bill_infers_case_units_from_printed_gross(mock_llm, tenant_a):
    """When Gross Amt uniquely matches (Cs×UPC)+Pcs, skip the question and
    recompute quantity — the bill's own total is the answer key (§4.2)."""
    resp = _upload_bill(tenant_a)
    assert resp.status_code == 201, resp.data
    assert resp.data["status"] == "PREVIEWED"
    assert resp.data["clarifications"] == []
    line = resp.data["preview"]["lines"][0]
    assert line["quantity"] == "15"
    assert line["include"] is True
    assert line["flags"] == []
    assert resp.data["preview"]["resolved_formula"] == SupplierBillTemplate.LineTotalFormula.CASE_UNITS_PLUS_LOOSE


@patch(
    "core.services.llm.extract_purchase_bill",
    return_value=_dms_extract(cs="5", upc="24", pcs="0", price="107.14", gross="12856.80"),
)
def test_bill_reincludes_zero_pcs_case_row_after_qty_recombine(mock_llm, tenant_a):
    """Full-carton rows print Pcs=0; after qty becomes (5×24)+0=120 they must
    be included, not left as 'quantity must be > 0' errors."""
    resp = _upload_bill(tenant_a)
    assert resp.status_code == 201, resp.data
    assert resp.data["status"] == "PREVIEWED"
    line = resp.data["preview"]["lines"][0]
    assert line["quantity"] == "120"
    assert line["include"] is True
    assert resp.data["valid_rows"] == 1
    assert resp.data["error_rows"] == 0


@patch(
    "core.services.llm.extract_purchase_bill",
    return_value=_dms_extract(gross=None, headers=False),
)
def test_bill_asks_clarifications_from_line_fields_without_headers(mock_llm, tenant_a):
    """If the model omits column_headers, Cs/UPC on the lines is enough to
    detect the DMS layout instead of silently using Pcs as quantity."""
    resp = _upload_bill(tenant_a)
    assert resp.status_code == 201, resp.data
    assert resp.data["status"] == "NEEDS_CLARIFICATION"
    fields = {c["field"] for c in resp.data["clarifications"]}
    assert fields == {"qty_formula"}


@patch("core.services.llm.extract_purchase_bill", return_value={
    "supplier_name": "Tally Supplier",
    "supplier_gstin": "29AAGCB1234A1Z5",
    "bill_number": "T-1",
    "bill_date": "2026-08-01",
    "confidence": 0.92,
    "column_headers": ["Particulars", "HSN", "Qty", "Rate", "Amount", "GST%"],
    "lines": [{
        "name": "Notebook A4",
        "hsn_code": "4820",
        "quantity": "12",
        "unit_price": "25",
        "gst_rate": "18",
        "printed_gross_amt": "300",
    }],
})
def test_simple_qty_rate_bill_does_not_ask_or_recombine(mock_llm, tenant_a):
    """A Tally-style Qty × Rate bill has no pack-size columns — leave qty as
    printed and skip the clarification loop."""
    resp = _upload_bill(tenant_a)
    assert resp.status_code == 201, resp.data
    assert resp.data["status"] == "PREVIEWED"
    assert resp.data["clarifications"] == []
    assert resp.data["preview"]["lines"][0]["quantity"] == "12"
    assert resp.data["preview"]["resolved_formula"] == SupplierBillTemplate.LineTotalFormula.SIMPLE


@patch("core.services.llm.extract_purchase_bill", return_value={
    "supplier_name": "Local Wholesaler",
    "supplier_gstin": "27AABCU9603R1ZM",
    "bill_number": "B-9",
    "bill_date": "2026-08-01",
    "confidence": 0.9,
    "column_headers": ["Item", "Boxes", "Pack", "Rate", "Amount", "GST%"],
    "lines": [{
        "name": "Tea 100g",
        "quantity": "2",
        "unit_price": "10",
        "gst_rate": "5",
        "raw_columns": {"boxes": "2", "pack": "12"},
        "printed_gross_amt": "240",
    }],
})
def test_bill_infers_box_times_pack_from_printed_amount(mock_llm, tenant_a):
    """A Boxes × Pack layout is a different vendor format from Cs/UPC — the
    same printed-amount scorer must pick boxes*pack without a code change."""
    resp = _upload_bill(tenant_a)
    assert resp.status_code == 201, resp.data
    assert resp.data["status"] == "PREVIEWED"
    assert resp.data["clarifications"] == []
    assert resp.data["preview"]["lines"][0]["quantity"] == "24"


def test_structured_xlsx_uses_line_items_sheet_and_recombines_qty(tenant_a):
    """A multi-sheet DMS/ChatGPT workbook must read Line Items (not the
    cover sheet) and compute qty as (Cs×UPC)+Pcs from printed Gross Amt."""
    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    cover = wb.active
    cover.title = "Invoice Details"
    cover["A1"] = "TAX INVOICE"
    cover["A2"] = "Registered Name"
    cover["B2"] = "VTC TRADEWINGS PVT"
    items = wb.create_sheet("Line Items")
    items.append([
        "Sl", "HSN", "PCode", "Item Description", "MRP", "Cs", "Pcs", "UPC",
        "Pcs Price", "Gross Amt", "GST %",
    ])
    items.append([1, "33049990", "80741159", "Olay NA IGF 40gm", 199, 0, 3, 24, 146.65, 439.95, 18])
    items.append([6, "33032100", "80895672", "OB WHLSALE SHINY CS", 198, 5, 0, 24, 107.14, 12856.80, 5])
    buf = BytesIO()
    wb.save(buf)
    resp = tenant_a.client.post("/api/v1/imports/", {
        "kind": "PURCHASE_BILL",
        "file": SimpleUploadedFile(
            "VTAGR-26-1038635.xlsx", buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
    }, format="multipart")
    assert resp.status_code == 201, resp.data
    assert resp.data["status"] == "PREVIEWED"
    assert resp.data["total_rows"] == 2
    lines = resp.data["preview"]["lines"]
    assert lines[0]["name"] == "Olay NA IGF 40gm"
    assert lines[0]["quantity"] == "3"
    assert lines[1]["name"] == "OB WHLSALE SHINY CS"
    assert lines[1]["quantity"] == "120"
    assert lines[1]["include"] is True


def test_structured_xlsx_keeps_pack_qty_when_gross_disagrees(tenant_a):
    """XLSX Cs/Pcs/UPC are the source of truth; Gross Amt only flags a mismatch."""
    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    items = wb.active
    items.title = "Line Items"
    items.append([
        "Sl", "Item Description", "Cs", "Pcs", "UPC", "Pcs Price", "Gross Amt", "GST %",
    ])
    items.append([7, "OB WHLSALE SHINYM CS", 5, 0, 24, 107.14, 5142.72, 5])
    buf = BytesIO()
    wb.save(buf)
    resp = tenant_a.client.post("/api/v1/imports/", {
        "kind": "PURCHASE_BILL",
        "file": SimpleUploadedFile(
            "pack-vs-gross.xlsx", buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
    }, format="multipart")
    assert resp.status_code == 201, resp.data
    line = resp.data["preview"]["lines"][0]
    assert line["quantity"] == "120"
    assert str(line.get("cs") or "") == "5"
    assert line["flags"]


def test_extraction_continuation_merges_remaining_si_rows():
    from core.services.llm import (
        merge_extraction_line_payloads,
        needs_extraction_continuation,
        _normalize_payload,
    )

    first = _normalize_payload({
        "printed_line_count": "30",
        "lines": [
            {"si": "1", "name": "Olay NA IGF 40gm", "quantity": "3", "gst_rate": "18"},
            {"si": "20", "name": "H&S 7in1 Rs3", "quantity": "19", "gst_rate": "18"},
        ],
    })
    assert first["printed_line_count"] == 30
    assert needs_extraction_continuation(first, None) is True
    assert needs_extraction_continuation(first, "length") is True
    extra = _normalize_payload({
        "printed_line_count": "30",
        "lines": [
            {"si": "20", "name": "H&S 7in1 Rs3", "quantity": "19", "gst_rate": "18"},
            {"si": "30", "name": "Whspr Ultra XL+ 15s CFC", "quantity": "2", "gst_rate": "0"},
        ],
    })
    merged = merge_extraction_line_payloads(first, extra)
    names = [ln["name"] for ln in merged["lines"]]
    assert names == ["Olay NA IGF 40gm", "H&S 7in1 Rs3", "Whspr Ultra XL+ 15s CFC"]
    assert merged["printed_line_count"] == 30
    shuffled = merge_extraction_line_payloads(
        _normalize_payload({
            "printed_line_count": "5",
            "lines": [
                {"si": "1", "name": "Olay NA IGF 40gm", "quantity": "3", "gst_rate": "18"},
                {"si": "5", "name": "Guard Cream 125 gm", "quantity": "6", "gst_rate": "18"},
                {"si": "2", "name": "Olay NA IGF 20gm", "quantity": "4", "gst_rate": "18"},
            ],
        }),
        _normalize_payload({
            "printed_line_count": "5",
            "lines": [
                {"si": "3", "name": "Gillette TSG", "quantity": "6", "gst_rate": "18"},
                {"si": "4", "name": "Guard Cream Small 25 gm", "quantity": "12", "gst_rate": "18"},
            ],
        }),
    )
    assert [ln["name"] for ln in shuffled["lines"]] == [
        "Olay NA IGF 40gm",
        "Olay NA IGF 20gm",
        "Gillette TSG",
        "Guard Cream Small 25 gm",
        "Guard Cream 125 gm",
    ]
    complete = _normalize_payload({
        "printed_line_count": "2",
        "lines": [
            {"si": "1", "name": "Olay NA IGF 40gm", "quantity": "3", "gst_rate": "18"},
            {"si": "2", "name": "H&S 7in1 Rs3", "quantity": "19", "gst_rate": "18"},
        ],
    })
    assert needs_extraction_continuation(complete, "stop") is False


@patch("core.services.llm.extract_purchase_bill", return_value={
    "supplier_name": "Retail Buyer LLP",
    "buyer_name": "Retail Buyer LLP",
    "buyer_gstin": "",
    "bill_number": "SB-1",
    "bill_date": "2026-07-05",
    "confidence": 0.9,
    "lines": [
        {
            "name": "Ariel Powder 500g",
            "hsn_code": "3402",
            "quantity": "4",
            "unit_price": "150",
            "gst_rate": "18",
            "mrp": "180",
        },
    ],
})
def test_sales_bill_commit_creates_draft_sales_invoice(mock_llm, tenant_a):
    customer = make_customer(tenant_a.company, name="Retail Buyer LLP")
    resp = tenant_a.client.post("/api/v1/imports/", {
        "kind": "SALES_BILL",
        "customer_id": customer.id,
        "file": SimpleUploadedFile("sale.png", _png_bytes(), content_type="image/png"),
    }, format="multipart")
    assert resp.status_code == 201, resp.data
    assert resp.data["status"] == "PREVIEWED"
    job_id = resp.data["id"]

    commit = tenant_a.client.post(f"/api/v1/imports/{job_id}/commit/")
    assert commit.status_code == 200, commit.data
    assert commit.data["sales_invoice_id"]
    invoice = SalesInvoice.objects.get(pk=commit.data["sales_invoice_id"])
    assert invoice.status == SalesInvoice.Status.DRAFT
    assert invoice.customer_id == customer.id
    assert invoice.items.count() == 1
    # A sales bill's unit price is a selling price, never written to purchase_price.
    product = Product.objects.get(company=tenant_a.company, name="Ariel Powder 500g")
    assert product.selling_price == Decimal("150")
    assert product.purchase_price == Decimal("0")


@patch("core.services.llm.extract_purchase_bill", return_value=FAKE_EXTRACT)
def test_purchase_bill_creates_supplier_from_extracted_name(mock_llm, tenant_a):
    job = _upload_bill(tenant_a).data
    resp = tenant_a.client.post(f"/api/v1/imports/{job['id']}/commit/")
    assert resp.status_code == 200, resp.data
    assert Supplier.objects.filter(company=tenant_a.company, name="Acme Distributors").exists()


@patch("core.services.llm.extract_purchase_bill", return_value=FAKE_EXTRACT)
def test_purchase_bill_rejects_unparseable_date(mock_llm, tenant_a):
    supplier = make_supplier(tenant_a.company)
    job = _upload_bill(tenant_a, supplier_id=supplier.id).data
    preview = job["preview"]
    preview["bill_date"] = "not-a-date"
    patched = tenant_a.client.patch(
        f"/api/v1/imports/{job['id']}/preview/",
        {"bill_date": "not-a-date", "lines": preview["lines"], "supplier_id": supplier.id},
        format="json",
    )
    assert patched.status_code == 200
    resp = tenant_a.client.post(f"/api/v1/imports/{job['id']}/commit/")
    assert resp.status_code == 400
    assert "bill date" in str(resp.data).lower() or "parse" in str(resp.data).lower()


@patch("core.services.llm.extract_purchase_bill", side_effect=Exception("provider down"))
def test_purchase_bill_retry_extract(mock_fail, tenant_a):
    job = _upload_bill(tenant_a).data
    assert job["status"] == "FAILED"
    with patch("core.services.llm.extract_purchase_bill", return_value=FAKE_EXTRACT) as mock_ok:
        resp = tenant_a.client.post(f"/api/v1/imports/{job['id']}/retry-extract/")
        assert resp.status_code == 200, resp.data
        assert resp.data["status"] == "PREVIEWED"
        mock_ok.assert_called_once()


def test_split_bill_image_emits_top_and_bottom_for_tall_photo():
    from io import BytesIO

    from PIL import Image

    from core.services.bill_images import TARGET_LONG_EDGE, split_bill_image

    image = Image.new("RGB", (400, 900), color=(255, 255, 255))
    buf = BytesIO()
    image.save(buf, format="JPEG")
    views = split_bill_image(buf.getvalue())
    assert "full" in views and "top" in views and "bottom" in views
    full = Image.open(BytesIO(views["full"]))
    assert max(full.size) >= TARGET_LONG_EDGE


def test_should_probe_remaining_rows_when_model_claims_complete_at_20():
    from core.services.llm import needs_extraction_continuation, should_probe_remaining_rows

    truncated = {
        "printed_line_count": 20,
        "lines": [{"si": str(i), "name": f"Item {i}"} for i in range(1, 21)],
    }
    assert needs_extraction_continuation(truncated, "stop") is False
    assert should_probe_remaining_rows(truncated, "stop") is True
    short = {
        "printed_line_count": 5,
        "lines": [{"si": str(i), "name": f"Item {i}"} for i in range(1, 6)],
    }
    assert should_probe_remaining_rows(short, "stop") is False


@patch("core.services.llm._extract_openai_compatible")
@patch("core.services.llm._bill_model_for", return_value="gpt-4o")
@patch("core.services.llm._provider", return_value="openai")
def test_extract_purchase_bill_fetches_si_21_30_after_20_row_stop(mock_provider, mock_model, mock_oa):
    from core.services.llm import extract_purchase_bill

    def fake(*, prompt, **kwargs):
        if "between 1 and 15" in prompt:
            lines = [
                {"si": str(i), "name": f"Item {i}", "quantity": "1", "gst_rate": "18"}
                for i in range(1, 21)
            ]
            return {"printed_line_count": "20", "lines": lines, "column_headers": ["SI"]}, "stop"
        if "between 21 and" in prompt:
            lines = [
                {"si": str(i), "name": f"Item {i}", "quantity": "1", "gst_rate": "18"}
                for i in range(21, 31)
            ]
            return {"printed_line_count": "30", "lines": lines}, "stop"
        return {"printed_line_count": "30", "lines": []}, "stop"

    mock_oa.side_effect = fake
    payload = extract_purchase_bill([b"\xff\xd8\xff dummy"])
    assert len(payload["lines"]) == 30
    assert payload["printed_line_count"] == 30
    assert mock_oa.call_count == 2


@patch("core.services.llm._extract_openai_compatible")
@patch("core.services.llm._bill_model_for", return_value="gpt-4o")
@patch("core.services.llm._provider", return_value="openai")
def test_extract_purchase_bill_stops_on_short_table(mock_provider, mock_model, mock_oa):
    from core.services.llm import extract_purchase_bill

    mock_oa.return_value = (
        {
            "printed_line_count": "5",
            "lines": [
                {"si": str(i), "name": f"Item {i}", "quantity": "1", "gst_rate": "18"}
                for i in range(1, 6)
            ],
        },
        "stop",
    )
    payload = extract_purchase_bill([b"\xff\xd8\xff dummy"])
    assert len(payload["lines"]) == 5
    assert mock_oa.call_count == 1


@patch("core.services.llm.extract_purchase_bill", return_value=_dms_extract())
def test_simple_template_ignored_when_bill_has_pack_columns(mock_llm, tenant_a):
    """A saved SIMPLE layout from a prior bad extract must not pin Pcs as qty
    when this bill actually has Cs/UPC columns."""
    SupplierBillTemplate.objects.create(
        company=tenant_a.company,
        gstin=DMS_GSTIN,
        party_name="VTC TRADEWINGS PVT",
        line_total_formula=SupplierBillTemplate.LineTotalFormula.SIMPLE,
        column_mapping={"qty_formula": "quantity"},
        column_signature=DMS_COLUMN_HEADERS,
    )
    resp = _upload_bill(tenant_a)
    assert resp.status_code == 201, resp.data
    assert resp.data["status"] == "PREVIEWED"
    assert resp.data["preview"]["lines"][0]["quantity"] == "15"


@patch("core.services.llm.extract_purchase_bill", return_value=FAKE_EXTRACT)
def test_bill_reupload_same_idempotency_key_starts_new_job(mock_llm, tenant_a):
    """Re-uploading the same photo must not replay a prior preview/commit."""
    headers = {"HTTP_IDEMPOTENCY_KEY": "import-upload-PURCHASE_BILL-same.jpeg-1-1"}
    first = tenant_a.client.post(
        "/api/v1/imports/",
        {
            "kind": "PURCHASE_BILL",
            "file": SimpleUploadedFile("same.jpeg", _png_bytes(), content_type="image/jpeg"),
        },
        format="multipart",
        **headers,
    )
    assert first.status_code == 201, first.data
    second = tenant_a.client.post(
        "/api/v1/imports/",
        {
            "kind": "PURCHASE_BILL",
            "file": SimpleUploadedFile("same.jpeg", _png_bytes(), content_type="image/jpeg"),
        },
        format="multipart",
        **headers,
    )
    assert second.status_code == 201, second.data
    assert second.data["id"] != first.data["id"]


def test_qty_uses_printed_gross_over_wrong_pcs_ocr():
    """Circled Pcs=3 misread as 2 must still become billed qty 3 from Gross Amt."""
    from imports.qty_formula import apply_qty_formula

    lines = [{
        "name": "Olay NA IGF 40gm",
        "quantity": "2",
        "unit_price": "146.65",
        "printed_gross_amt": "439.95",
        "extras": {"upc": "24"},
    }]
    apply_qty_formula(lines, {"qty_formula": "quantity"}, tolerance=Decimal("0.50"))
    assert lines[0]["quantity"] == "3"
    assert lines[0]["flags"] == []


def test_qty_case_row_from_gross_when_cs_missing():
    from imports.qty_formula import apply_qty_formula

    lines = [{
        "name": "OB WHLSALE SHINYS CS",
        "quantity": "0",
        "unit_price": "107.14",
        "printed_gross_amt": "12856.80",
        "extras": {"upc": "24"},
    }]
    apply_qty_formula(lines, {"qty_formula": "quantity"}, tolerance=Decimal("0.50"))
    assert lines[0]["quantity"] == "120"
    assert lines[0]["flags"] == []


def test_qty_does_not_follow_garbage_printed_gross():
    """A nonsense Gross Amt must not replace a formula qty that was already resolved."""
    from imports.qty_formula import apply_qty_formula

    lines = [{
        "name": "Olay NA IGF 40gm",
        "quantity": "3",
        "unit_price": "10",
        "printed_gross_amt": "999",
        "cs": "2",
        "upc": "6",
        "extras": {"cs": "2", "upc": "6"},
    }]
    apply_qty_formula(lines, {"qty_formula": "cs*upc+quantity"}, tolerance=Decimal("0.50"))
    assert lines[0]["quantity"] == "15"
    assert len(lines[0]["flags"]) == 1


def test_qty_recovers_missing_cs_from_gross_and_keeps_pcs():
    from imports.qty_formula import apply_qty_formula

    lines = [{
        "name": "OB WHLSALE SHINYS CS",
        "quantity": "0",
        "unit_price": "107.14",
        "printed_gross_amt": "12856.80",
        "extras": {"upc": "24"},
    }]
    apply_qty_formula(lines, {"qty_formula": "cs*upc+quantity"}, tolerance=Decimal("0.50"))
    assert lines[0]["quantity"] == "120"
    assert lines[0]["cs"] == "5"
    assert lines[0]["pcs"] == "0"
    assert lines[0]["flags"] == []


def test_qty_gst_inclusive_amount_still_resolves_pieces():
    from imports.qty_formula import apply_qty_formula

    lines = [{
        "name": "Olay NA IGF 40gm",
        "quantity": "2",
        "unit_price": "146.65",
        "gst_rate": "18",
        "printed_gross_amt": "519.14",
        "extras": {"upc": "24"},
    }]
    apply_qty_formula(lines, {"qty_formula": "quantity"}, tolerance=Decimal("0.50"))
    assert lines[0]["quantity"] == "3"
    assert lines[0]["pcs"] == "3"
    assert lines[0]["cs"] == "0"
