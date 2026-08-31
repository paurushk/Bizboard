"""Item type, opening lots/serials, godown lifecycle, FEFO, import extras."""

from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db.models import Sum
from openpyxl import Workbook

from inventory.models import BatchLot, MovementType, StockBalance, StockMovement, Warehouse
from inventory.services import InventoryService, InventoryValuationService
from masters.models import Product
from tests.conftest import add_stock, create_draft_invoice, create_draft_purchase, make_customer, make_product, make_supplier

pytestmark = pytest.mark.django_db


def _opening(tenant, **payload):
    return tenant.client.post("/api/v1/inventory/opening-stock/", payload, format="json")


def test_service_item_rejects_opening_stock(tenant_a):
    product = make_product(tenant_a.company, name="Internet 30MBPS", sku="NET-30")
    product.product_type = Product.ProductType.SERVICE
    product.track_inventory = False
    product.save()
    resp = _opening(tenant_a, product=product.id, quantity="10")
    assert resp.status_code == 400


def test_serial_opening_requires_serial_numbers(tenant_a):
    product = make_product(tenant_a.company, sku="SN-REQ", track_serial=True)
    missing = _opening(tenant_a, product=product.id, quantity="2")
    assert missing.status_code == 400
    assert "Serial numbers are required for serial-tracked opening stock." in str(missing.data)

    mismatch = _opening(tenant_a, product=product.id, quantity="2", serial_numbers=["ONLY-ONE"])
    assert mismatch.status_code == 400

    ok = _opening(tenant_a, product=product.id, quantity="2", serial_numbers=["SN-1", "SN-2"])
    assert ok.status_code == 201, ok.data
    from inventory.models import SerialNumber, StockBalance

    assert StockBalance.objects.get(product=product).on_hand == Decimal("2")
    assert SerialNumber.objects.filter(product=product).count() == 2


def test_opening_two_godown_lots(tenant_a):
    product = make_product(tenant_a.company, name="Milk", sku="MILK1")
    product.track_batch = True
    product.save()
    default = InventoryService.default_warehouse(tenant_a.company)
    cold = Warehouse.objects.create(
        company=tenant_a.company, name="Cold room", code="COLD", is_default=False,
    )
    expiry_a = (date.today() + timedelta(days=20)).isoformat()
    expiry_b = (date.today() + timedelta(days=40)).isoformat()
    a = _opening(
        tenant_a, product=product.id, quantity="10", warehouse=default.id,
        batch_no="LOT-A", expiry_date=expiry_a, unit_cost="1.00",
    )
    b = _opening(
        tenant_a, product=product.id, quantity="15", warehouse=cold.id,
        batch_no="LOT-B", expiry_date=expiry_b, unit_cost="1.00",
    )
    assert a.status_code == 201, a.data
    assert b.status_code == 201, b.data
    assert BatchLot.objects.filter(product=product).count() == 2
    assert StockBalance.objects.filter(product=product, on_hand__gt=0).count() == 2


def test_expired_opening_lot_rejected(tenant_a):
    product = make_product(tenant_a.company, sku="EXP-1")
    product.track_batch = True
    product.save()
    resp = _opening(
        tenant_a, product=product.id, quantity="5",
        batch_no="OLD", expiry_date=(date.today() - timedelta(days=1)).isoformat(),
    )
    assert resp.status_code == 400


def test_cannot_enable_batch_after_movements(tenant_a):
    product = make_product(tenant_a.company, sku="LIVE-1")
    add_stock(tenant_a, product, "5")
    resp = tenant_a.client.patch(f"/api/v1/products/{product.id}/", {"track_batch": True}, format="json")
    assert resp.status_code == 400


def test_deactivate_godown_with_stock_blocked(tenant_a):
    product = make_product(tenant_a.company)
    add_stock(tenant_a, product, "4")
    warehouse = Warehouse.objects.get(company=tenant_a.company, is_default=True)
    other = Warehouse.objects.create(company=tenant_a.company, name="Spare", code="SPARE")
    resp = tenant_a.client.patch(
        f"/api/v1/inventory/warehouses/{other.id}/",
        {"is_active": False},
        format="json",
    )
    assert resp.status_code == 200
    blocked = tenant_a.client.patch(
        f"/api/v1/inventory/warehouses/{warehouse.id}/",
        {"is_active": False},
        format="json",
    )
    assert blocked.status_code == 400


def test_fefo_sale_splits_lots(tenant_a):
    product = make_product(tenant_a.company, sku="FEFO-1")
    product.track_batch = True
    product.save()
    warehouse = InventoryService.default_warehouse(tenant_a.company)
    early = (date.today() + timedelta(days=5)).isoformat()
    later = (date.today() + timedelta(days=25)).isoformat()
    assert _opening(
        tenant_a, product=product.id, quantity="10", warehouse=warehouse.id,
        batch_no="A", expiry_date=early, unit_cost="10",
    ).status_code == 201
    assert _opening(
        tenant_a, product=product.id, quantity="20", warehouse=warehouse.id,
        batch_no="B", expiry_date=later, unit_cost="12",
    ).status_code == 201
    customer = make_customer(tenant_a.company)
    inv = create_draft_invoice(tenant_a, customer, [
        {"product": product.id, "quantity": "25", "unit_price": "40"},
    ])
    complete = tenant_a.client.post(f"/api/v1/sales/invoices/{inv['id']}/complete/")
    assert complete.status_code == 200, complete.data
    sales = list(
        StockMovement.objects.filter(product=product, movement_type=MovementType.SALE).order_by("id")
    )
    assert len(sales) == 2
    qty = sorted(abs(m.quantity) for m in sales)
    assert qty == [Decimal("10"), Decimal("15")]


def test_inclusive_opening_stores_exclusive_cost(tenant_a):
    product = make_product(tenant_a.company, sku="INC-1", gst_rate="18")
    product.purchase_tax_inclusive = True
    product.purchase_price = Decimal("118")
    product.save()
    resp = _opening(tenant_a, product=product.id, quantity="2", unit_cost="118")
    assert resp.status_code == 201, resp.data
    move = StockMovement.objects.get(product=product, movement_type=MovementType.OPENING_STOCK)
    assert move.unit_cost == Decimal("100.00")


def test_products_import_item_name_star_alias(tenant_a):
    csv_content = (
        b"Item Name*,Item code,GST Tax Rate(%),Sales Price,Item type,Current stock\n"
        b"Milk,MILK1,5,40,Product,100\n"
        b"Internet 30MBPS,NET30,0,499,Service,\n"
    )
    upload = tenant_a.client.post("/api/v1/imports/", {
        "kind": "PRODUCTS",
        "file": SimpleUploadedFile("bulk.csv", csv_content, content_type="text/csv"),
    }, format="multipart")
    assert upload.status_code == 201, upload.data
    assert upload.data["error_rows"] == 0
    commit = tenant_a.client.post(f"/api/v1/imports/{upload.data['id']}/commit/")
    assert commit.status_code == 200, commit.data
    milk = Product.objects.get(company=tenant_a.company, sku="MILK1")
    net = Product.objects.get(company=tenant_a.company, sku="NET30")
    assert milk.product_type == Product.ProductType.GOODS
    assert net.product_type == Product.ProductType.SERVICE
    assert StockBalance.objects.filter(product=milk).exists()
    assert not StockBalance.objects.filter(product=net).exists()


def test_duplicate_committed_file_rejected(tenant_a):
    csv_content = b"name,sku\nSoap,SOAP-DUP\n"
    f1 = SimpleUploadedFile("a.csv", csv_content, content_type="text/csv")
    job = tenant_a.client.post("/api/v1/imports/", {"kind": "PRODUCTS", "file": f1}, format="multipart")
    assert job.status_code == 201
    tenant_a.client.post(f"/api/v1/imports/{job.data['id']}/commit/")
    f2 = SimpleUploadedFile("a.csv", csv_content, content_type="text/csv")
    again = tenant_a.client.post("/api/v1/imports/", {"kind": "PRODUCTS", "file": f2}, format="multipart")
    assert again.status_code == 400


def test_opening_lots_sheet_posts_per_godown_opening(tenant_a):
    default = InventoryService.default_warehouse(tenant_a.company)
    cold = Warehouse.objects.create(
        company=tenant_a.company, name="Cold room", code="COLD", is_default=False,
    )
    wb = Workbook()
    items = wb.active
    items.title = "items"
    items.append(["Item Name*", "Item code", "GST Tax Rate(%)", "Item type", "Current stock"])
    items.append(["Milk", "MILK-LOTS", "5", "Product", "999"])
    lots = wb.create_sheet("opening_lots")
    lots.append(["sku", "godown", "quantity", "batch_no", "expiry_date", "unit_cost"])
    exp = (date.today() + timedelta(days=20)).isoformat()
    lots.append(["MILK-LOTS", default.name, "10", "LOT-A", exp, "1"])
    lots.append(["MILK-LOTS", cold.name, "15", "LOT-B", (date.today() + timedelta(days=40)).isoformat(), "1"])
    buf = BytesIO()
    wb.save(buf)
    upload = tenant_a.client.post(
        "/api/v1/imports/",
        {
            "kind": "PRODUCTS",
            "file": SimpleUploadedFile(
                "lots.xlsx",
                buf.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
        format="multipart",
    )
    assert upload.status_code == 201, upload.data
    commit = tenant_a.client.post(f"/api/v1/imports/{upload.data['id']}/commit/")
    assert commit.status_code == 200, commit.data
    product = Product.objects.get(company=tenant_a.company, sku="MILK-LOTS")
    assert product.track_batch is True
    assert BatchLot.objects.filter(product=product).count() == 2
    qty = StockBalance.objects.filter(product=product).aggregate(total=Sum("on_hand"))["total"]
    assert qty == Decimal("25")
    assert not StockMovement.objects.filter(
        product=product, movement_type=MovementType.OPENING_STOCK, quantity=Decimal("999"),
    ).exists()


def test_opening_serials_sheet_posts_serial_opening(tenant_a):
    wb = Workbook()
    items = wb.active
    items.title = "items"
    items.append(["name", "sku", "gst_rate", "track_serial"])
    items.append(["Phone", "PH-XLSX", "18", "yes"])
    serials = wb.create_sheet("opening_serials")
    serials.append(["sku", "godown", "serial_no", "as_of", "unit_cost"])
    serials.append(["PH-XLSX", "", "IMEI-X1", "", "9000"])
    serials.append(["PH-XLSX", "", "IMEI-X2", "", "9000"])
    buf = BytesIO()
    wb.save(buf)
    upload = tenant_a.client.post(
        "/api/v1/imports/",
        {
            "kind": "PRODUCTS",
            "file": SimpleUploadedFile(
                "serials.xlsx",
                buf.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
        format="multipart",
    )
    assert upload.status_code == 201, upload.data
    assert upload.data["error_rows"] == 0, upload.data
    commit = tenant_a.client.post(f"/api/v1/imports/{upload.data['id']}/commit/")
    assert commit.status_code == 200, commit.data
    from inventory.models import SerialNumber

    product = Product.objects.get(company=tenant_a.company, sku="PH-XLSX")
    assert product.track_serial is True
    assert StockBalance.objects.get(product=product).on_hand == Decimal("2")
    assert set(
        SerialNumber.objects.filter(product=product).values_list("serial_number", flat=True)
    ) == {"IMEI-X1", "IMEI-X2"}


def test_xlsx_excel_dates_and_misfilled_serials_do_not_500(tenant_a):
    """Numeric SKUs + Excel dates must save; a serials sheet of product names is ignored."""
    wb = Workbook()
    items = wb.active
    items.title = "items"
    items.append(["name", "sku", "gst_rate", "opening_stock"])
    items.append(["H&S Basic Cool", 80689219, 5, 60])
    lots = wb.create_sheet("opening_lots")
    lots.append(
        ["sku", "godown", "quantity", "as_of", "batch_no", "expiry_date", "manufacturing_date", "unit_cost"]
    )
    lots.append([
        80689219, None, 60,
        datetime(2026, 8, 20), "Lot 1", datetime(2027, 9, 22), "2026-04-01", 23.29,
    ])
    serials = wb.create_sheet("opening_serials")
    serials.append(["sku", "godown", "serial_no", "as_of", "unit_cost"])
    serials.append(["H&S Basic Cool", None, 80689219, datetime(2026, 8, 20), 23.29])
    buf = BytesIO()
    wb.save(buf)
    upload = tenant_a.client.post(
        "/api/v1/imports/",
        {
            "kind": "PRODUCTS",
            "file": SimpleUploadedFile(
                "products_import_template.xlsx",
                buf.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
        format="multipart",
    )
    assert upload.status_code == 201, upload.data
    assert upload.data["error_rows"] == 0, upload.data
    assert upload.data["valid_rows"] == 1
    assert any(
        (m.get("source") == "opening_serials") for m in (upload.data.get("column_mappings") or [])
    )
    commit = tenant_a.client.post(f"/api/v1/imports/{upload.data['id']}/commit/")
    assert commit.status_code == 200, commit.data
    product = Product.objects.get(company=tenant_a.company, sku="80689219")
    assert product.track_batch is True
    assert BatchLot.objects.filter(product=product).count() == 1
    qty = StockBalance.objects.filter(product=product).aggregate(total=Sum("on_hand"))["total"]
    assert qty == Decimal("60")


def test_empty_items_sheet_does_not_treat_lots_as_products(tenant_a):
    wb = Workbook()
    items = wb.active
    items.title = "items"
    items.append(["name", "sku", "gst_rate"])
    lots = wb.create_sheet("opening_lots")
    lots.append(["sku", "godown", "quantity", "batch_no", "expiry_date", "unit_cost"])
    lots.append(["MILK1", "", "4000", "LOT-B", "2026-09-22", "0.01"])
    buf = BytesIO()
    wb.save(buf)
    upload = tenant_a.client.post(
        "/api/v1/imports/",
        {
            "kind": "PRODUCTS",
            "file": SimpleUploadedFile(
                "blank.xlsx",
                buf.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
        format="multipart",
    )
    assert upload.status_code == 400
    assert "item rows" in str(upload.data).lower()


def test_stock_count_posts_adjustment_variance(tenant_a):
    product = make_product(tenant_a.company, sku="CNT-1")
    add_stock(tenant_a, product, "10")
    warehouse = InventoryService.default_warehouse(tenant_a.company)
    session = tenant_a.client.post(
        "/api/v1/inventory/stock-counts/",
        {"warehouse": warehouse.id},
        format="json",
    )
    assert session.status_code == 201, session.data
    lines = session.data.get("lines") or []
    assert lines, session.data
    line = lines[0]
    patch = tenant_a.client.patch(
        f"/api/v1/inventory/stock-counts/{session.data['id']}/",
        {"lines": [{"id": line["id"], "counted_qty": "7"}]},
        format="json",
    )
    assert patch.status_code == 200, patch.data
    posted = tenant_a.client.post(f"/api/v1/inventory/stock-counts/{session.data['id']}/post/")
    assert posted.status_code == 200, posted.data
    movements = StockMovement.objects.filter(
        product=product, movement_type=MovementType.ADJUSTMENT, reason="STOCK_COUNT",
    )
    assert movements.exists()
    assert StockBalance.objects.get(product=product, warehouse=warehouse, batch__isnull=True).on_hand == Decimal("7")


def test_products_template_download(tenant_a):
    from imports.services import PRODUCTS_ITEM_COLUMNS, PRODUCTS_ITEM_SAMPLE_ROWS
    from openpyxl import load_workbook

    assert all(len(row) == len(PRODUCTS_ITEM_COLUMNS) for row in PRODUCTS_ITEM_SAMPLE_ROWS)

    resp = tenant_a.client.get("/api/v1/imports/template/", {"kind": "PRODUCTS"})
    assert resp.status_code == 200
    assert "spreadsheetml" in resp["Content-Type"]
    wb = load_workbook(BytesIO(resp.content))
    headers = [cell.value for cell in next(wb["items"].iter_rows(max_row=1))]
    for col in ("name", "description", "category", "mrp", "product_type", "godown", "batch_no", "expiry_date"):
        assert col in headers
    assert headers == list(PRODUCTS_ITEM_COLUMNS)
    assert "opening_lots" in wb.sheetnames
    csv_resp = tenant_a.client.get("/api/v1/imports/template/", {"kind": "PRODUCTS", "as": "csv"})
    assert csv_resp.status_code == 200
    csv_header = csv_resp.content.decode().splitlines()[0]
    assert csv_header.split(",") == list(PRODUCTS_ITEM_COLUMNS)
    assert PRODUCTS_ITEM_SAMPLE_ROWS == [[""] * len(PRODUCTS_ITEM_COLUMNS)]
    data_line = csv_resp.content.decode().splitlines()[1]
    assert data_line and all(not cell for cell in data_line.split(","))
    assert "SOAP" not in csv_resp.content.decode()


def test_products_csv_extended_columns_persist(tenant_a):
    from masters.custom_fields import validate_definitions
    from masters.models import Category

    tenant_a.company.item_custom_field_defs = validate_definitions([], [
        {"key": "brandCode", "label": "Brand code", "type": "text", "active": True},
    ])
    tenant_a.company.save(update_fields=["item_custom_field_defs"])

    expiry = (date.today() + timedelta(days=20)).isoformat()
    csv_content = (
        "name,sku,description,category,hsn_code,gst_rate,purchase_price,purchase_tax_inclusive,"
        "selling_price,selling_tax_inclusive,mrp,wholesale_price,reorder_level,unit,product_type,"
        "track_batch,godown,opening_stock,unit_cost,batch_no,expiry_date,brand_code\n"
        f"Milk,MILK-EXT,Milk Boxes,Dairy,0402,5,30,Exclusive,40,Exclusive,45,35,1000,ML,Goods,"
        f"Yes,,12,0.50,LOT-A,{expiry},AMUL\n"
    ).encode()
    upload = tenant_a.client.post("/api/v1/imports/", {
        "kind": "PRODUCTS",
        "file": SimpleUploadedFile("extended.csv", csv_content, content_type="text/csv"),
    }, format="multipart")
    assert upload.status_code == 201, upload.data
    assert upload.data["error_rows"] == 0, upload.data
    commit = tenant_a.client.post(f"/api/v1/imports/{upload.data['id']}/commit/")
    assert commit.status_code == 200, commit.data
    product = Product.objects.get(company=tenant_a.company, sku="MILK-EXT")
    assert product.description == "Milk Boxes"
    assert product.mrp == Decimal("45")
    assert product.wholesale_price == Decimal("35")
    assert product.product_type == Product.ProductType.GOODS
    assert product.track_batch is True
    assert product.custom_fields.get("brandCode") == "AMUL"
    assert Category.objects.filter(company=tenant_a.company, name="Dairy").exists()
    lot = BatchLot.objects.get(product=product, batch_no="LOT-A")
    assert lot.expiry_date.isoformat() == expiry
    assert StockBalance.objects.get(product=product, batch=lot).on_hand == Decimal("12")


def test_legacy_ten_column_products_csv_still_imports(tenant_a):
    csv_content = (
        b"name,sku,barcode,hsn_code,gst_rate,purchase_price,selling_price,reorder_level,unit,opening_stock\n"
        b"Soap,SOAP-LEGACY,,3401,18,40,55,5,PCS,25\n"
    )
    upload = tenant_a.client.post("/api/v1/imports/", {
        "kind": "PRODUCTS",
        "file": SimpleUploadedFile("legacy.csv", csv_content, content_type="text/csv"),
    }, format="multipart")
    assert upload.status_code == 201, upload.data
    assert upload.data["error_rows"] == 0, upload.data
    commit = tenant_a.client.post(f"/api/v1/imports/{upload.data['id']}/commit/")
    assert commit.status_code == 200, commit.data
    product = Product.objects.get(company=tenant_a.company, sku="SOAP-LEGACY")
    assert product.track_batch is False
    assert StockBalance.objects.get(product=product, batch__isnull=True).on_hand == Decimal("25")


def test_alternate_unit_sale_converts_to_base(tenant_a):
    from masters.models import Unit

    product = make_product(tenant_a.company, sku="BOX-1")
    box = Unit.objects.create(company=tenant_a.company, name="Box", short_name="BOX", uqc_code="BOX")
    product.alternate_unit = box
    product.conversion_rate = Decimal("10")
    product.save(update_fields=["alternate_unit", "conversion_rate"])
    add_stock(tenant_a, product, "100")
    customer = make_customer(tenant_a.company)
    inv = create_draft_invoice(tenant_a, customer, [
        {"product": product.id, "quantity": "2", "unit_price": "50", "unit_name": "BOX"},
    ])
    complete = tenant_a.client.post(f"/api/v1/sales/invoices/{inv['id']}/complete/")
    assert complete.status_code == 200, complete.data
    sold = abs(sum(
        m.quantity for m in StockMovement.objects.filter(
            product=product, movement_type=MovementType.SALE,
        )
    ))
    assert sold == Decimal("20")


def test_sales_return_unidentified_lot_refused(tenant_a, monkeypatch):
    from sales.cogs_service import CogsService

    product = make_product(tenant_a.company, sku="UID-1")
    add_stock(tenant_a, product, "5")
    customer = make_customer(tenant_a.company)
    inv = create_draft_invoice(tenant_a, customer, [
        {"product": product.id, "quantity": "2", "unit_price": "100"},
    ])
    assert tenant_a.client.post(f"/api/v1/sales/invoices/{inv['id']}/complete/").status_code == 200
    monkeypatch.setattr(CogsService, "invoice_sale_moves", staticmethod(lambda invoice, product=None: []))
    ret = tenant_a.client.post("/api/v1/sales/returns/", {
        "customer": customer.id, "sales_invoice": inv["id"],
        "items": [{"product": product.id, "quantity": "2", "unit_price": "100"}],
    }, format="json")
    assert ret.status_code == 201, ret.data
    resp = tenant_a.client.post(f"/api/v1/sales/returns/{ret.data['id']}/complete/")
    assert resp.status_code == 400


def test_purchase_return_damaged_posts_adjustment(tenant_a):
    product = make_product(tenant_a.company, sku="DMG-1")
    supplier = make_supplier(tenant_a.company)
    pur = create_draft_purchase(tenant_a, supplier, [
        {"product": product.id, "quantity": "5", "unit_price": "80"},
    ])
    assert tenant_a.client.post(f"/api/v1/purchases/invoices/{pur['id']}/complete/").status_code == 200
    ret = tenant_a.client.post("/api/v1/purchases/returns/", {
        "supplier": supplier.id, "purchase_invoice": pur["id"],
        "items": [{"product": product.id, "quantity": "2", "unit_price": "80", "condition": "DAMAGED"}],
    }, format="json")
    assert ret.status_code == 201, ret.data
    done = tenant_a.client.post(f"/api/v1/purchases/returns/{ret.data['id']}/complete/")
    assert done.status_code == 200, done.data
    assert StockMovement.objects.filter(
        product=product, movement_type=MovementType.ADJUSTMENT, reason="DAMAGED",
    ).exists()
    assert not StockMovement.objects.filter(
        product=product, movement_type=MovementType.PURCHASE_RETURN,
    ).exists()


def test_opening_as_of_closed_gst_period_blocked(tenant_a):
    from reporting.models import GstReturnPeriod

    product = make_product(tenant_a.company, sku="LOCK-1")
    period = f"{date.today().year:04d}-{date.today().month:02d}"
    GstReturnPeriod.objects.create(
        company=tenant_a.company, period=period, status=GstReturnPeriod.Status.CLOSED,
    )
    resp = _opening(tenant_a, product=product.id, quantity="2", as_of=date.today().isoformat())
    assert resp.status_code == 400


def test_opening_as_of_before_current_fy_blocked(tenant_a):
    tenant_a.company.fy_start_month = 4
    tenant_a.company.save(update_fields=["fy_start_month"])
    product = make_product(tenant_a.company, sku="FY-1")
    today = date.today()
    fy_start = date(today.year - 1 if today.month < 4 else today.year, 4, 1)
    resp = _opening(
        tenant_a, product=product.id, quantity="2", as_of=(fy_start - timedelta(days=1)).isoformat(),
    )
    assert resp.status_code == 400
    ok = _opening(tenant_a, product=product.id, quantity="2", as_of=fy_start.isoformat())
    assert ok.status_code == 201, ok.data


def test_sales_return_damaged_batched_posts_lot_adjustment(tenant_a):
    product = make_product(tenant_a.company, sku="DMG-B")
    product.track_batch = True
    product.save()
    expiry = (date.today() + timedelta(days=30)).isoformat()
    opened = _opening(
        tenant_a, product=product.id, quantity="5",
        batch_no="LOT-A", expiry_date=expiry, unit_cost="10",
    )
    assert opened.status_code == 201, opened.data
    customer = make_customer(tenant_a.company)
    inv = create_draft_invoice(tenant_a, customer, [
        {"product": product.id, "quantity": "2", "unit_price": "100"},
    ])
    assert tenant_a.client.post(f"/api/v1/sales/invoices/{inv['id']}/complete/").status_code == 200
    ret = tenant_a.client.post("/api/v1/sales/returns/", {
        "customer": customer.id, "sales_invoice": inv["id"],
        "items": [{"product": product.id, "quantity": "2", "unit_price": "100", "condition": "DAMAGED"}],
    }, format="json")
    assert ret.status_code == 201, ret.data
    done = tenant_a.client.post(f"/api/v1/sales/returns/{ret.data['id']}/complete/")
    assert done.status_code == 200, done.data
    adj = StockMovement.objects.get(
        product=product, movement_type=MovementType.ADJUSTMENT, reason="DAMAGED",
    )
    assert adj.batch_id is not None
    warehouse = InventoryService.default_warehouse(tenant_a.company)
    lot = BatchLot.objects.get(product=product, batch_no="LOT-A")
    assert StockBalance.objects.get(product=product, warehouse=warehouse, batch=lot).on_hand == Decimal("3")


def test_low_stock_uses_warehouse_reorder(tenant_a):
    from inventory.models import WarehouseReorderLevel

    product = make_product(tenant_a.company, sku="LS-1", reorder_level="100")
    add_stock(tenant_a, product, "10")
    warehouse = InventoryService.default_warehouse(tenant_a.company)
    WarehouseReorderLevel.objects.create(
        company=tenant_a.company, warehouse=warehouse, product=product, reorder_level=Decimal("5"),
    )
    resp = tenant_a.client.get("/api/v1/inventory/alerts/")
    assert resp.status_code == 200
    ids = [row["product"] for row in resp.data.get("items") or []]
    assert product.id not in ids
    InventoryService.post_movement(
        company=tenant_a.company, product=product, movement_type=MovementType.ADJUSTMENT,
        quantity=Decimal("-8"), reason="TEST", user=tenant_a.owner, warehouse=warehouse,
    )
    resp = tenant_a.client.get("/api/v1/inventory/alerts/")
    ids = [row["product"] for row in resp.data.get("items") or []]
    assert product.id in ids
    row = next(item for item in resp.data["items"] if item["product"] == product.id)
    assert Decimal(str(row["reorder_level"])) == Decimal("5")


def test_low_stock_includes_default_zero_reorder_when_out_of_stock(tenant_a):
    product = make_product(tenant_a.company, sku="LS-0", reorder_level="0")
    add_stock(tenant_a, product, "1")
    InventoryService.post_movement(
        company=tenant_a.company, product=product, movement_type=MovementType.ADJUSTMENT,
        quantity=Decimal("-1"), reason="TEST", user=tenant_a.owner,
    )
    resp = tenant_a.client.get("/api/v1/inventory/alerts/")
    assert resp.status_code == 200
    ids = [row["product"] for row in resp.data.get("items") or []]
    assert product.id in ids


def test_low_stock_sums_across_godowns_without_override(tenant_a):
    product = make_product(tenant_a.company, sku="LS-SUM", reorder_level="15")
    source = InventoryService.default_warehouse(tenant_a.company)
    dest = Warehouse.objects.create(company=tenant_a.company, name="G2", code="G2X")
    InventoryService.post_movement(
        company=tenant_a.company, product=product, warehouse=source,
        movement_type=MovementType.OPENING_STOCK, quantity="10", unit_cost="10",
        user=tenant_a.owner,
    )
    InventoryService.post_movement(
        company=tenant_a.company, product=product, warehouse=dest,
        movement_type=MovementType.OPENING_STOCK, quantity="10", unit_cost="10",
        user=tenant_a.owner,
    )
    resp = tenant_a.client.get("/api/v1/inventory/alerts/")
    ids = [row["product"] for row in resp.data.get("items") or []]
    assert product.id not in ids
    product.reorder_level = Decimal("25")
    product.save(update_fields=["reorder_level"])
    resp = tenant_a.client.get("/api/v1/inventory/alerts/")
    ids = [row["product"] for row in resp.data.get("items") or []]
    assert product.id in ids
    row = next(item for item in resp.data["items"] if item["product"] == product.id)
    assert Decimal(str(row["available"])) == Decimal("20")


def test_alternate_unit_purchase_cost_is_per_base_unit(tenant_a):
    from masters.models import Unit

    product = make_product(tenant_a.company, sku="BOX-P")
    box = Unit.objects.create(company=tenant_a.company, name="Box", short_name="BOX", uqc_code="BOX")
    product.alternate_unit = box
    product.conversion_rate = Decimal("10")
    product.save(update_fields=["alternate_unit", "conversion_rate"])
    supplier = make_supplier(tenant_a.company)
    pur = create_draft_purchase(tenant_a, supplier, [
        {"product": product.id, "quantity": "1", "unit_price": "100", "unit_name": "BOX"},
    ])
    complete = tenant_a.client.post(f"/api/v1/purchases/invoices/{pur['id']}/complete/")
    assert complete.status_code == 200, complete.data
    move = StockMovement.objects.get(product=product, movement_type=MovementType.PURCHASE)
    assert move.quantity == Decimal("10")
    assert Decimal(str(move.unit_cost)) == Decimal("10.00")


def test_alternate_unit_sale_return_restores_base_qty(tenant_a):
    from masters.models import Unit

    product = make_product(tenant_a.company, sku="BOX-SR")
    box = Unit.objects.create(company=tenant_a.company, name="Box", short_name="BOX", uqc_code="BOX")
    product.alternate_unit = box
    product.conversion_rate = Decimal("10")
    product.save(update_fields=["alternate_unit", "conversion_rate"])
    add_stock(tenant_a, product, "100")
    customer = make_customer(tenant_a.company)
    inv = create_draft_invoice(tenant_a, customer, [
        {"product": product.id, "quantity": "2", "unit_price": "50", "unit_name": "BOX"},
    ])
    assert tenant_a.client.post(f"/api/v1/sales/invoices/{inv['id']}/complete/").status_code == 200
    ret = tenant_a.client.post("/api/v1/sales/returns/", {
        "customer": customer.id, "sales_invoice": inv["id"],
        "items": [{"product": product.id, "quantity": "1", "unit_price": "50"}],
    }, format="json")
    assert ret.status_code == 201, ret.data
    done = tenant_a.client.post(f"/api/v1/sales/returns/{ret.data['id']}/complete/")
    assert done.status_code == 200, done.data
    restored = sum(
        m.quantity for m in StockMovement.objects.filter(
            product=product, movement_type=MovementType.SALES_RETURN,
        )
    )
    assert restored == Decimal("10")


def test_purchase_return_damaged_cancel_restores_writeoff(tenant_a):
    product = make_product(tenant_a.company, sku="DMG-C")
    supplier = make_supplier(tenant_a.company)
    pur = create_draft_purchase(tenant_a, supplier, [
        {"product": product.id, "quantity": "5", "unit_price": "80"},
    ])
    assert tenant_a.client.post(f"/api/v1/purchases/invoices/{pur['id']}/complete/").status_code == 200
    ret = tenant_a.client.post("/api/v1/purchases/returns/", {
        "supplier": supplier.id, "purchase_invoice": pur["id"],
        "items": [{"product": product.id, "quantity": "2", "unit_price": "80", "condition": "DAMAGED"}],
    }, format="json")
    assert ret.status_code == 201, ret.data
    assert tenant_a.client.post(f"/api/v1/purchases/returns/{ret.data['id']}/complete/").status_code == 200
    after_return = StockBalance.objects.get(product=product, batch__isnull=True).on_hand
    cancelled = tenant_a.client.post(f"/api/v1/purchases/returns/{ret.data['id']}/cancel/")
    assert cancelled.status_code == 200, cancelled.data
    restored = StockBalance.objects.get(product=product, batch__isnull=True).on_hand
    assert restored == after_return + Decimal("2")


def test_stock_count_ignores_client_system_qty(tenant_a):
    product = make_product(tenant_a.company, sku="CNT-Q")
    add_stock(tenant_a, product, "4")
    warehouse = InventoryService.default_warehouse(tenant_a.company)
    session = tenant_a.client.post(
        "/api/v1/inventory/stock-counts/",
        {
            "warehouse": warehouse.id,
            "lines": [{
                "product": product.id,
                "system_qty": "999",
                "counted_qty": "4",
            }],
        },
        format="json",
    )
    assert session.status_code == 201, session.data
    line = session.data["lines"][0]
    assert Decimal(str(line["system_qty"])) == Decimal("4")


def test_stock_count_cancel_blocks_post(tenant_a):
    product = make_product(tenant_a.company, sku="CNT-C")
    add_stock(tenant_a, product, "4")
    warehouse = InventoryService.default_warehouse(tenant_a.company)
    session = tenant_a.client.post(
        "/api/v1/inventory/stock-counts/",
        {"warehouse": warehouse.id},
        format="json",
    )
    assert session.status_code == 201, session.data
    cancelled = tenant_a.client.post(f"/api/v1/inventory/stock-counts/{session.data['id']}/cancel/")
    assert cancelled.status_code == 200, cancelled.data
    assert cancelled.data["status"] == "CANCELLED"
    posted = tenant_a.client.post(f"/api/v1/inventory/stock-counts/{session.data['id']}/post/")
    assert posted.status_code == 400


def test_barcode_image_svg(tenant_a):
    resp = tenant_a.client.get("/api/v1/products/barcode-image/", {"code": "ABC-123"})
    assert resp.status_code == 200
    assert "svg" in resp["Content-Type"]


def test_expiry_alert_writeoff_posts_adjustment(tenant_a):
    product = make_product(tenant_a.company, sku="C03-WO")
    product.track_batch = True
    product.save()
    expiry = (date.today() + timedelta(days=3)).isoformat()
    opened = _opening(
        tenant_a, product=product.id, quantity="4",
        batch_no="NEAR", expiry_date=expiry, unit_cost="8",
    )
    assert opened.status_code == 201, opened.data
    warehouse = InventoryService.default_warehouse(tenant_a.company)
    lot = BatchLot.objects.get(product=product, batch_no="NEAR")
    listed = tenant_a.client.get("/api/v1/inventory/alerts/expiry/", {"days": 7})
    assert listed.status_code == 200
    items = listed.data.get("items") or listed.data
    assert any(str(row.get("batch") or row.get("id")) == str(lot.id) for row in items)
    wo = tenant_a.client.post(
        "/api/v1/inventory/alerts/expiry/",
        {
            "product": product.id,
            "warehouse": warehouse.id,
            "batch": lot.id,
            "quantity": "4",
        },
        format="json",
    )
    assert wo.status_code == 201, wo.data
    move = StockMovement.objects.get(product=product, reference_type="expiry_write_off")
    assert move.reason == "EXPIRED"
    assert move.movement_type == MovementType.ADJUSTMENT
    assert StockBalance.objects.get(product=product, warehouse=warehouse, batch=lot).on_hand == Decimal("0")
